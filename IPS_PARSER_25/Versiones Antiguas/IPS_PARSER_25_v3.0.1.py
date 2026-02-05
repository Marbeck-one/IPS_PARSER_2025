import pandas as pd
import os
import sys
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v3.0.0 - ESTRATEGIA GLOBAL Y AUTO-LIMPIEZA
# =============================================================================

class IPSParserV300:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_V3.0.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.new_indicator_count = 1
        
        # Configuración
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible' # Por defecto
        
        # Memoria
        self.memory_skip = set()      
        self.memory_generate = False
        self.memory_skip_empty = False

        # Lista negra automática (Cosas que siempre son basura)
        self.blacklist_auto = ["NÚMERO", "NUMERO", "N°", "NO", "Nº"]

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA v3.0")
        print("="*60)
        
        # 1. Porcentajes
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        
        # 2. Estrategia Global de Filas Ocultas
        print("\n2. ¿Cómo manejar filas OCULTAS en todos los archivos?")
        print("   [v] Automático: Procesar SOLO VISIBLES (Recomendado/Seguro).")
        print("   [t] Automático: Procesar TODO (Incluido oculto).")
        print("   [i] Interactivo: Preguntar caso a caso (Permite saltar hoja).")
        resp_h = input("   >> Elija opción (Enter=v): ").lower().strip()
        
        if resp_h == 't': self.opt_hidden_strategy = 'all'
        elif resp_h == 'i': self.opt_hidden_strategy = 'interactive'
        else: self.opt_hidden_strategy = 'visible'

        print(f"\n[OK] Configuración guardada. Estrategia Ocultos: {self.opt_hidden_strategy.upper()}")
        print("-" * 60)

    def print_summary_and_exit(self):
        print("\n" + "="*60)
        print("   RESUMEN FINAL")
        print("="*60)
        print(f"  * Registros Extraídos:  {len(self.flat_data)}")
        print("-" * 60)
        
        if self.flat_data:
            self.export_excel()
        else:
            print("[AVISO] No se generó archivo de salida (sin datos).")
        sys.exit()

    def ask_hidden_interactive(self, count, sheet_name):
        print(f"\n[DECISIÓN] Se detectaron {count} filas OCULTAS en '{sheet_name}'.")
        print("  [v] Procesar SOLO VISIBLES.")
        print("  [t] Procesar TODAS.")
        print("  [s] Saltar esta hoja.")
        
        while True:
            choice = input("  >> Elija (v/t/s): ").lower().strip()
            if choice == 'v': return 'visible'
            if choice == 't': return 'all'
            if choice == 's': return 'skip'

    def ask_column_action(self, missing_cols, sheet_name):
        print(f"\n[ALERTA] En hoja '{sheet_name}' faltan columnas: {missing_cols}")
        print("  Opciones:")
        print("   [c] Continuar (Rellenar con 'No aplica').")
        print("   [s] Saltar esta hoja.")
        print("   [f] Saltar este ARCHIVO completo.")
        print("   [d] Detener y Guardar.")
        
        while True:
            choice = input("  >> Elija opción (c/s/f/d): ").lower().strip()
            if choice == 'c': return 'continue'
            if choice == 's': return 'skip_sheet'
            if choice == 'f': return 'skip_file'
            if choice == 'd': self.print_summary_and_exit()

    def ask_weird_row_action(self, row_idx, content):
        clean = str(content).strip().upper()
        if clean in self.memory_skip: return 'skip'
        if "NUEVO" in clean and self.memory_generate: return 'auto'
        if content == "[VACÍO]" and self.memory_skip_empty: return 'skip'
        
        print(f"\n[FILA RARA #{row_idx}] NÚMERO dice: '{content}'")
        print("  [c]  Procesar (Generar código auto).")
        print("  [ca] Procesar SIEMPRE (Auto para todos los NUEVOS).")
        print("  [s]  Saltar fila.")
        print("  [x]  Saltar SIEMPRE filas con este texto.")
        print("  [d]  Detener y Guardar.")

        while True:
            choice = input("  >> Elija: ").lower().strip()
            if choice == 'c': return 'auto'
            if choice == 'ca': 
                self.memory_generate = True
                return 'auto'
            if choice == 's': return 'skip'
            if choice == 'x':
                if content == "[VACÍO]": self.memory_skip_empty = True
                else: self.memory_skip.add(clean)
                print(f"     -> Ignorando '{content}' siempre.")
                return 'skip'
            if choice == 'd': self.print_summary_and_exit()

    def transform_percentage(self, val, col_name):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or val == "No aplica": return val
        try:
            num = float(val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    def get_excel_files(self):
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$") and "IPS_CONSOLIDADO" not in f]
        if not valid_files:
            print(f"[ERROR] Carpeta vacía o sin Excel: {self.folder_path}")
            sys.exit()
        print(f"\n[INFO] Se encontraron {len(valid_files)} archivos.")
        return valid_files

    def get_hidden_rows(self, file_path, sheet_name):
        try:
            wb = load_workbook(file_path, read_only=False, data_only=True)
            if sheet_name not in wb.sheetnames: return set()
            ws = wb[sheet_name]
            hidden = set()
            for row_idx, row_dim in ws.row_dimensions.items():
                if row_dim.hidden:
                    hidden.add(row_idx - 1) 
            wb.close()
            return hidden
        except: return set()

    def process_folder(self):
        files = self.get_excel_files()
        self.configure()
        
        for idx_file, file_path in enumerate(files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            skip_file_flag = False
            
            print(f"\n>>> PROCESANDO ({idx_file + 1}/{len(files)}): {file_name}")
            
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except Exception as e:
                print(f"  [ERROR] Archivo corrupto: {e}")
                continue

            for sheet in sheet_names:
                if skip_file_flag: break

                # 1. GESTIÓN DE OCULTOS (GLOBAL vs INTERACTIVO)
                hidden_rows = self.get_hidden_rows(file_path, sheet)
                ignored_rows = set()
                
                if hidden_rows:
                    if self.opt_hidden_strategy == 'visible':
                        ignored_rows = hidden_rows
                    elif self.opt_hidden_strategy == 'all':
                        ignored_rows = set()
                    else: # Interactive
                        action = self.ask_hidden_interactive(len(hidden_rows), sheet)
                        if action == 'skip': continue
                        elif action == 'visible': ignored_rows = hidden_rows
                        elif action == 'all': ignored_rows = set()

                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except: continue

                # 2. BUSCAR ENCABEZADO
                h_idx = None
                for idx, row in df.iterrows():
                    if idx in ignored_rows: continue
                    row_vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    if any(x in row_vals for x in ["NÚMERO", "NUMERO", "N°"]):
                        h_idx = idx
                        break
                
                if h_idx is None:
                    continue 

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                def find_c(names):
                    for i, h in enumerate(headers):
                        h_clean = " ".join(str(h).split()).lower()
                        for n in names:
                            if n.lower() in h_clean: return i
                    return None

                c_map = {
                    "num": find_c(["NÚMERO", "NUMERO", "N°"]),
                    "prod": find_c(["PRODUCTO"]),
                    "ind": find_c(["INDICADOR"]),
                    "form": find_c(["FORMULA", "FÓRMULA"]),
                    "uni": find_c(["UNIDAD"]),
                    "resp": find_c(["RESPONSABLE"]),
                    "gest": find_c(["GESTOR"]),
                    "sup": find_c(["SUPERVISORES"]),
                    "meta": find_c(["Meta 2025", "Meta 2026", "Meta"]),
                    "pond": find_c(["Ponderador"]),
                    "op_desc": find_c(["Operandos"]), 
                    "op_est": find_c(["Operandos Estimados", "Estimados Meta", "Estimados"]),
                    "proy": find_c(["Cumplimiento Proyectado", "Proyectado"]),
                    "cump_meta": find_c(["% Cumplimiento"]),
                    "medios": find_c(["Medios"]),
                    "control": find_c(["Control de Cambios"]),
                    "inst": find_c(["Instrumentos"])
                }

                missing = [k for k, v in c_map.items() if v is None and k not in ["pond", "control"]]
                if missing:
                    action = self.ask_column_action(missing, sheet)
                    if action == 'skip_sheet': continue
                    if action == 'skip_file': 
                        skip_file_flag = True
                        break

                months = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                          "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                          "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                month_map = {m: find_c([m]) for m in months}

                sheet_rows = []
                count_rows = 0
                
                for i in range(h_idx + 1, len(df)):
                    if i in ignored_rows: continue

                    def get_val(col_idx, row_offset=0):
                        target_row = i + row_offset
                        if col_idx is None: return "No aplica"
                        if target_row >= len(df) or target_row in ignored_rows: return ""
                        val = df.iloc[target_row, col_idx]
                        return val if pd.notna(val) else ""

                    raw_num = get_val(c_map["num"])
                    str_num = str(raw_num).strip()

                    # 3. FILTRO AUTOMÁTICO DE BASURA (Mejora v3.0)
                    if str_num.upper() in self.blacklist_auto:
                        continue # Saltar sin preguntar

                    final_code = None
                    is_new = "NUEVO" in str_num.upper()
                    is_empty = (str_num == "" or str_num.lower() == "nan")
                    
                    if is_empty or is_new:
                        ind_val = get_val(c_map["ind"])
                        if ind_val and str(ind_val).strip() not in ["", "0", "No aplica"]:
                            action = self.ask_weird_row_action(i+1, str_num if str_num else "[VACÍO]")
                            if action == 'skip': continue
                            
                            prefix = file_name.split()[0][:8]
                            clean_s = ''.join(e for e in sheet if e.isalnum())
                            final_code = f"NUEVO_{self.new_indicator_count}_{prefix}_{clean_s}"
                            self.new_indicator_count += 1
                        else: continue 
                    
                    elif not any(c.isdigit() for c in str_num):
                        action = self.ask_weird_row_action(i+1, str_num)
                        if action == 'skip': continue
                        if action == 'auto':
                             prefix = file_name.split()[0][:8]
                             final_code = f"GEN_{self.new_indicator_count}_{prefix}"
                             self.new_indicator_count += 1
                    else:
                        final_code = str_num

                    if not final_code: continue

                    count_rows += 1
                    
                    row_data = {
                        "ARCHIVO": file_name, "HOJA": sheet, "NÚMERO": final_code,
                        "PRODUCTO O PROCESO ESPECÍFICO": get_val(c_map["prod"]),
                        "INDICADOR": get_val(c_map["ind"]),
                        "FORMULA": get_val(c_map["form"]),
                        "UNIDAD": get_val(c_map["uni"]),
                        "RESPONSABLE CENTRO DE RESPONSABILIDAD": get_val(c_map["resp"]),
                        "GESTOR": get_val(c_map["gest"]),
                        "SUPERVISORES": get_val(c_map["sup"]),
                        "Meta 2026": self.transform_percentage(get_val(c_map["meta"]), "Meta"),
                        "Ponderador": self.transform_percentage(get_val(c_map["pond"]), "Ponderador"),
                    }
                    
                    row_data["Descripción Operando 1"] = get_val(c_map["op_desc"], 0)
                    row_data["Descripción Operando 2"] = get_val(c_map["op_desc"], 3)
                    
                    row_data["Meta Operando 1 (Valor)"] = get_val(c_map["op_est"], 3)
                    row_data["Meta Operando 2 (Valor)"] = get_val(c_map["op_est"], 5)
                    
                    for m_name, m_idx in month_map.items():
                        row_data[f"{m_name} Op 1"] = get_val(m_idx, 3)
                        row_data[f"{m_name} Op 2"] = get_val(m_idx, 5)

                    row_data["Cumplimiento Proyectado 2026 Op 1"] = get_val(c_map["proy"], 3)
                    row_data["Cumplimiento Proyectado 2026 Op 2"] = get_val(c_map["proy"], 5)
                    row_data["% Cumplimiento de Meta"] = self.transform_percentage(get_val(c_map["cump_meta"], 3), "% Cump Meta")
                    row_data["Medios de Verificación"] = get_val(c_map["medios"], 0)
                    row_data["Control de Cambios"] = get_val(c_map["control"], 0)
                    row_data["Instrumentos de Gestión Asociados"] = get_val(c_map["inst"], 0)

                    sheet_rows.append(row_data)
                    self.flat_data.append(row_data)

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  -> {count_rows} ok.")

            if skip_file_flag: print("  [SALTO] Archivo omitido.")

        self.print_summary_and_exit()

    def export_excel(self):
        print(f"\n{'='*60}\nGUARDANDO ARCHIVO MAESTRO...\n{'='*60}")
        wb = Workbook()
        ws = wb.active; ws.title = "Carga Bruta"
        ws.append(list(self.flat_data[0].keys()))
        for r in self.flat_data: ws.append(list(r.values()))
        
        ws_style = wb.create_sheet("Planilla Estilizada")
        styles = {
            'file': PatternFill("solid", fgColor="000000"),
            'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"),
            'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True),
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }

        row_idx = 1
        FULL_WIDTH = 64

        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            
            c = ws_style.cell(row=row_idx, column=1, value=f"ARCHIVO: {fname}")
            c.fill = styles['file']; c.font = styles['w_font']; c.alignment = Alignment('center')
            ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
            row_idx += 1
            
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_style.cell(row=row_idx, column=1, value=f"HOJA: {sname}")
                c.fill = styles['sheet']; c.font = styles['w_font']
                ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
                row_idx += 1
                
                keys = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                for c_i, k in enumerate(keys, 1):
                    c = ws_style.cell(row=row_idx, column=c_i, value=k)
                    c.fill = styles['head']; c.font = styles['b_font']; c.border = styles['border']
                row_idx += 1
                
                for r in rows:
                    for c_i, k in enumerate(keys, 1):
                        c = ws_style.cell(row=row_idx, column=c_i, value=r[k])
                        c.border = styles['border']; c.alignment = Alignment(wrapText=True, vertical='top')
                    row_idx += 1
                row_idx += 1

        for i in range(1, FULL_WIDTH + 2):
            ws_style.column_dimensions[get_column_letter(i)].width = 22

        wb.save(self.output_file)
        print(f"[ÉXITO] Archivo generado: {self.output_file}")

if __name__ == "__main__":
    try:
        path = input("Ruta de la carpeta (Enter para actual): ").strip() or os.getcwd()
        if os.path.isdir(path):
            parser = IPSParserV300(path)
            parser.process_folder()
        else: print("Ruta inválida.")
    except Exception as e:
        print(f"Error: {e}")
        input("Enter para salir.")