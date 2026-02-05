import pandas as pd
import os
import sys
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v2.4.0 - SOLUCIÓN DEFINITIVA (Keywords Robustos + Orden Estricto)
# =============================================================================

class IPSParserV240:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_V2.4.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.new_indicator_count = 1

    def ask_user_action(self, context_msg):
        print(f"\n[ATENCIÓN REQUERIDA] {context_msg}")
        print("  Opciones:")
        print("   [s] Saltar esta hoja.")
        print("   [d] Detener proceso.")
        while True:
            choice = input("  >> Elija opción (s/d): ").lower().strip()
            if choice == 's': return 'skip'
            elif choice == 'd': sys.exit()

    def ask_column_action(self, missing_cols, sheet_name):
        print(f"\n[FALTAN DATOS] Hoja '{sheet_name}' sin columnas: {missing_cols}")
        print("   [c] Continuar (Rellenar con 'No aplica').")
        print("   [s] Saltar hoja.")
        print("   [d] Detener.")
        while True:
            choice = input("  >> Elija opción (c/s/d): ").lower().strip()
            if choice == 'c': return 'continue'
            elif choice == 's': return 'skip'
            elif choice == 'd': sys.exit()

    def transform_percentage(self, val, col_name):
        if pd.isna(val) or val == "" or val == "No aplica": return val
        try:
            num = float(val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    def get_excel_files(self):
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$") and "IPS_CONSOLIDADO" not in f]
        if not valid_files:
            print(f"[ERROR] No hay archivos Excel en: {self.folder_path}")
            sys.exit()
        print(f"\n[INFO] Se encontraron {len(valid_files)} archivos.")
        return valid_files

    def get_hidden_rows(self, file_path, sheet_name):
        try:
            wb = load_workbook(file_path, read_only=False, data_only=True)
            if sheet_name not in wb.sheetnames: return set()
            ws = wb[sheet_name]
            hidden = set()
            for row_idx, row_dim in ws.row_dimensions.items():
                if row_dim.hidden:
                    hidden.add(row_idx - 1) 
            wb.close()
            return hidden
        except: return set()

    def process_folder(self):
        files = self.get_excel_files()
        
        for file_path in files:
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            
            print(f"\n{'='*60}")
            print(f"PROCESANDO: {file_name}")
            print(f"{'='*60}")
            
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except: continue

            for sheet in sheet_names:
                hidden_rows = self.get_hidden_rows(file_path, sheet)
                if hidden_rows:
                    print(f"  [VISIÓN HUMANA] {len(hidden_rows)} filas ocultas en '{sheet}'. Ignoradas.")

                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except: continue

                h_idx = None
                for idx, row in df.iterrows():
                    if idx in hidden_rows: continue
                    row_vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    if any(x in row_vals for x in ["NÚMERO", "NUMERO", "N°"]):
                        h_idx = idx
                        break
                
                if h_idx is None:
                    print(f"\n>>> Hoja '{sheet}'")
                    action = self.ask_user_action("No se encontró encabezado 'NÚMERO'.")
                    if action == 'skip': continue

                print(f"\n>>> Analizando Hoja: '{sheet}' (Encabezado Fila {h_idx+1})")
                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                # MAPEO ULTRA ROBUSTO (Normaliza espacios y saltos de línea)
                def find_c(names):
                    for i, h in enumerate(headers):
                        # Convierte "Cumplimiento\nProyectado" o "Operandos  Estimados" en "cumplimiento proyectado" / "operandos estimados"
                        h_clean = " ".join(str(h).split()).lower()
                        for n in names:
                            if n.lower() in h_clean: return i
                    return None

                c_map = {
                    "num": find_c(["NÚMERO", "NUMERO", "N°"]),
                    "prod": find_c(["PRODUCTO"]),
                    "ind": find_c(["INDICADOR"]),
                    "form": find_c(["FORMULA", "FÓRMULA"]),
                    "uni": find_c(["UNIDAD"]),
                    "resp": find_c(["RESPONSABLE"]),
                    "gest": find_c(["GESTOR"]),
                    "sup": find_c(["SUPERVISORES"]),
                    "meta": find_c(["Meta 2025", "Meta 2026", "Meta"]),
                    "pond": find_c(["Ponderador"]),
                    "op_desc": find_c(["Operandos"]), 
                    # Busca "Estimados Meta" o "Operandos Estimados" o simplemente "Estimados"
                    "op_est": find_c(["Operandos Estimados", "Estimados Meta", "Estimados"]),
                    # Busca "Proyectado" para tolerar "Proyectado 2025" o "Cumplimiento Proyectado"
                    "proy": find_c(["Cumplimiento Proyectado", "Proyectado"]),
                    "cump_meta": find_c(["% Cumplimiento"]),
                    "medios": find_c(["Medios"]),
                    "control": find_c(["Control de Cambios"]),
                    "inst": find_c(["Instrumentos"])
                }

                missing = [k for k, v in c_map.items() if v is None and k not in ["pond", "control"]]
                if missing:
                    action = self.ask_column_action(missing, sheet)
                    if action == 'skip': continue

                months = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                          "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                          "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                month_map = {m: find_c([m]) for m in months}

                sheet_rows = []
                count_rows = 0
                
                for i in range(h_idx + 1, len(df)):
                    if i in hidden_rows: continue

                    def get_val(col_idx, row_offset=0):
                        target_row = i + row_offset
                        if col_idx is None or target_row >= len(df) or target_row in hidden_rows: 
                            return "No aplica"
                        val = df.iloc[target_row, col_idx]
                        return val if pd.notna(val) else ""

                    raw_num = get_val(c_map["num"])
                    str_num = str(raw_num).strip()

                    if str_num == "" or str_num.lower() == "nan" or "NUEVO" in str_num.upper():
                        ind_val = get_val(c_map["ind"])
                        if ind_val and str(ind_val).strip() not in ["", "0", "No aplica"]:
                            prefix = file_name.split()[0][:8]
                            clean_s = ''.join(e for e in sheet if e.isalnum())
                            final_code = f"NUEVO_{self.new_indicator_count}_{prefix}_{clean_s}"
                            print(f"    [AVISO] Fila {i+1}: Nuevo detectado: {final_code}")
                            self.new_indicator_count += 1
                        else: continue 
                    else:
                        if not any(c.isdigit() for c in str_num): continue 
                        final_code = str_num

                    count_rows += 1
                    
                    # ---------------------------------------------------------
                    # CONSTRUCCIÓN ORDENADA DE COLUMNAS
                    # ---------------------------------------------------------
                    row_data = {
                        "ARCHIVO": file_name, 
                        "HOJA": sheet, 
                        "NÚMERO": final_code,
                        "PRODUCTO O PROCESO ESPECÍFICO": get_val(c_map["prod"]),
                        "INDICADOR": get_val(c_map["ind"]),
                        "FORMULA": get_val(c_map["form"]),
                        "UNIDAD": get_val(c_map["uni"]),
                        "RESPONSABLE CENTRO DE RESPONSABILIDAD": get_val(c_map["resp"]),
                        "GESTOR": get_val(c_map["gest"]),
                        "SUPERVISORES": get_val(c_map["sup"]),
                        "Meta 2026": self.transform_percentage(get_val(c_map["meta"]), "Meta"),
                        "Ponderador": self.transform_percentage(get_val(c_map["pond"]), "Ponderador"),
                    }
                    
                    # COLUMNAS M y N (Texto)
                    row_data["Descripción Operando 1"] = get_val(c_map["op_desc"], 0)
                    row_data["Descripción Operando 2"] = get_val(c_map["op_desc"], 3)
                    
                    # COLUMNAS O y P (Valores Numéricos - "Estimados Meta")
                    row_data["Meta Operando 1 (Valor)"] = get_val(c_map["op_est"], 3)
                    row_data["Meta Operando 2 (Valor)"] = get_val(c_map["op_est"], 5)
                    
                    # Meses
                    for m_name, m_idx in month_map.items():
                        row_data[f"{m_name} Op 1"] = get_val(m_idx, 3)
                        row_data[f"{m_name} Op 2"] = get_val(m_idx, 5)

                    # Proyección y Cierre
                    row_data["Cumplimiento Proyectado 2026 Op 1"] = get_val(c_map["proy"], 3)
                    row_data["Cumplimiento Proyectado 2026 Op 2"] = get_val(c_map["proy"], 5)
                    row_data["% Cumplimiento de Meta"] = self.transform_percentage(get_val(c_map["cump_meta"], 3), "% Cump Meta")
                    row_data["Medios de Verificación"] = get_val(c_map["medios"], 0)
                    row_data["Control de Cambios"] = get_val(c_map["control"], 0)
                    row_data["Instrumentos de Gestión Asociados"] = get_val(c_map["inst"], 0)

                    sheet_rows.append(row_data)
                    self.flat_data.append(row_data)

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  [OK] {count_rows} indicadores extraídos.")

    def export_excel(self):
        if not self.flat_data:
            print("\n[FIN] No hay datos para exportar.")
            return

        print(f"\n{'='*60}\nCONSOLIDANDO...\n{'='*60}")
        wb = Workbook()
        
        # HOJA 1: BRUTA
        ws = wb.active; ws.title = "Carga Bruta"
        ws.append(list(self.flat_data[0].keys()))
        for r in self.flat_data: ws.append(list(r.values()))
        
        # HOJA 2: ESTILIZADA
        ws = wb.create_sheet("Planilla Estilizada")
        styles = {
            'file': PatternFill("solid", fgColor="000000"),
            'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"),
            'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True),
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }

        row_idx = 1
        max_col = 0
        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            
            c = ws.cell(row=row_idx, column=1, value=f"ARCHIVO: {fname}")
            c.fill = styles['file']; c.font = styles['w_font']; c.alignment = Alignment('center')
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=15)
            row_idx += 1
            
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws.cell(row=row_idx, column=1, value=f"HOJA: {sname}")
                c.fill = styles['sheet']; c.font = styles['w_font']
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=15)
                row_idx += 1
                
                keys = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                max_col = max(max_col, len(keys))
                
                for c_i, k in enumerate(keys, 1):
                    c = ws.cell(row=row_idx, column=c_i, value=k)
                    c.fill = styles['head']; c.font = styles['b_font']; c.border = styles['border']
                row_idx += 1
                
                for r in rows:
                    for c_i, k in enumerate(keys, 1):
                        c = ws.cell(row=row_idx, column=c_i, value=r[k])
                        c.border = styles['border']; c.alignment = Alignment(wrapText=True, vertical='top')
                    row_idx += 1
                row_idx += 1

        for i in range(1, max_col + 2):
            ws.column_dimensions[get_column_letter(i)].width = 22

        wb.save(self.output_file)
        print(f"\n[ÉXITO] Archivo generado: {self.output_file}")

if __name__ == "__main__":
    try:
        path = input("Ruta de la carpeta (Enter para actual): ").strip() or os.getcwd()
        if os.path.isdir(path):
            parser = IPSParserV240(path)
            parser.process_folder()
            parser.export_excel()
        else: print("Ruta inválida.")
    except Exception as e:
        print(f"Error: {e}")
        input("Enter para salir.")