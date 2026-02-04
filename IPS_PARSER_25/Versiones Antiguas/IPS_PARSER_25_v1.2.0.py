import pandas as pd
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# =============================================================================
# IPS_PARSER_v1.3.0 - FASE 1: EXTRACCIÓN Y ESTILIZADO
# =============================================================================

class IPSParserPhase1:
    def __init__(self, master_file):
        self.master_file = master_file
        self.output_file = "IPS_PARSER_v1.3.0_Phase1.xlsx"
        self.sheets_to_process = ["CDC 2025", "PMG 2025", "Riesgos 2025"]
        self.all_extracted_data = []

    def alert_user(self, message):
        """Manejo de errores: Alerta al usuario y permite decidir si continuar."""
        print(f"\n[ALERTA]: {message}")
        choice = input("¿Desea continuar con el proceso? (s/n): ").lower()
        if choice != 's':
            print("Proceso detenido por el usuario.")
            sys.exit()
        return True

    def detect_headers(self, df, sheet_name):
        """Busca la fila de encabezados y valida su posición."""
        header_idx = None
        for idx, row in df.iterrows():
            if "NÚMERO" in [str(x).upper().strip() for x in row.values if pd.notna(x)]:
                header_idx = idx
                break
        
        if header_idx is None:
            self.alert_user(f"No se encontró la columna 'NÚMERO' en la hoja '{sheet_name}'.")
            return None

        # Alerta si los títulos no están en la fila 10 u 11 (índice 0-based)
        if header_idx not in [10, 11]:
            self.alert_user(f"En '{sheet_name}', los títulos se detectaron en la fila {header_idx + 1}, no en la 11 o 12.")
        
        return header_idx

    def get_col_index(self, headers, name):
        """Busca el índice de una columna por nombre (insensible a mayúsculas/minúsculas)."""
        for i, h in enumerate(headers):
            if name.lower() in str(h).lower().replace('\n', ' '):
                return i
        return None

    def process_sheet(self, sheet_name):
        print(f"--- Procesando hoja: {sheet_name} ---")
        try:
            # Intentar leer desde Excel, si falla o es CSV, manejarlo
            if self.master_file.endswith('.csv'):
                df = pd.read_csv(self.master_file, header=None, sep=None, engine='python', encoding='latin-1')
            else:
                df = pd.read_excel(self.master_file, sheet_name=sheet_name, header=None)
        except Exception as e:
            self.alert_user(f"No se pudo cargar la hoja '{sheet_name}': {e}")
            return

        h_idx = self.detect_headers(df, sheet_name)
        if h_idx is None: return

        headers = [str(h).strip() for h in df.iloc[h_idx]]
        
        # Mapeo de columnas requeridas
        cols = {
            "num": self.get_col_index(headers, "NÚMERO"),
            "prod": self.get_col_index(headers, "PRODUCTO O PROCESO ESPECÍFICO"),
            "ind": self.get_col_index(headers, "INDICADOR"),
            "form": self.get_col_index(headers, "FORMULA"),
            "uni": self.get_col_index(headers, "UNIDAD"),
            "resp": self.get_col_index(headers, "RESPONSABLE CENTRO DE RESPONSABILIDAD"),
            "gest": self.get_col_index(headers, "GESTOR"),
            "sup": self.get_col_index(headers, "SUPERVISORES"),
            "meta": self.get_col_index(headers, "Meta 2025"),
            "pond": self.get_col_index(headers, "Ponderador"),
            "op_desc": self.get_col_index(headers, "Operandos"),
            "op_est": self.get_col_index(headers, "Operandos Estimados Meta"),
            "proy": self.get_col_index(headers, "Cumplimiento Proyectado"),
            "cump_meta": self.get_col_index(headers, "% Cumplimiento de Meta"),
            "medios": self.get_col_index(headers, "Medios de Verificación"),
            "inst": self.get_col_index(headers, "Instrumentos de Gestión Asociados")
        }

        # Validar si faltan columnas críticas
        missing = [k for k, v in cols.items() if v is None and k not in ["pond"]] # Ponderador puede no estar en todas
        if missing:
            self.alert_user(f"Faltan columnas en '{sheet_name}': {missing}")

        # Meses M a AH
        months = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                  "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                  "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
        month_map = {m: self.get_col_index(headers, m) for m in months}

        # Extracción de bloques de indicadores (generalmente 6 filas por indicador)
        # La data real empieza típicamente 2 filas después del encabezado (fila 13)
        for i in range(h_idx + 1, len(df)):
            val_num = str(df.iloc[i, cols["num"]]).strip() if cols["num"] is not None else ""
            
            # Identificar fila de inicio de indicador (empieza por número)
            if val_num and val_num[0].isdigit():
                # Definiendo Offsets basados en tu requerimiento
                # i: Fila principal (13)
                # i+1: Fila del Valor Indicador (14) -> SE DEJA NOTA, NO SE USA
                # i+3: Fila del Operando 1 Real (16)
                # i+5: Fila del Operando 2 Real (18)
                
                row_data = {
                    "ORIGEN": sheet_name,
                    "NÚMERO": df.iloc[i, cols["num"]],
                    "PRODUCTO O PROCESO ESPECÍFICO": df.iloc[i, cols["prod"]] if cols["prod"] is not None else "",
                    "INDICADOR": df.iloc[i, cols["ind"]] if cols["ind"] is not None else "",
                    "FORMULA": df.iloc[i, cols["form"]] if cols["form"] is not None else "",
                    "UNIDAD": df.iloc[i, cols["uni"]] if cols["uni"] is not None else "",
                    "RESPONSABLE CENTRO DE RESPONSABILIDAD": df.iloc[i, cols["resp"]] if cols["resp"] is not None else "",
                    "GESTOR": df.iloc[i, cols["gest"]] if cols["gest"] is not None else "",
                    "SUPERVISORES": df.iloc[i, cols["sup"]] if cols["sup"] is not None else "",
                    "Meta 2025": df.iloc[i, cols["meta"]] if cols["meta"] is not None else "",
                    "Ponderador": df.iloc[i, cols["pond"]] if cols["pond"] is not None else "",
                    
                    # Columna K: Descripciones de Operandos
                    "Descripcion Operando 1": df.iloc[i, cols["op_desc"]] if cols["op_desc"] is not None else "",
                    "Descripcion Operando 2": df.iloc[i+3, cols["op_desc"]] if (cols["op_desc"] is not None and i+3 < len(df)) else "",
                    
                    # Columna L: Valores Meta Operandos
                    # NOTA: El valor del indicador está en df.iloc[i+1, cols["op_est"]]
                    "Meta Op 1": df.iloc[i+3, cols["op_est"]] if (cols["op_est"] is not None and i+3 < len(df)) else "",
                    "Meta Op 2": df.iloc[i+5, cols["op_est"]] if (cols["op_est"] is not None and i+5 < len(df)) else "",
                }

                # Columnas M - AH: Valores Mensuales
                for m, col_idx in month_map.items():
                    if col_idx is not None:
                        # NOTA: Valor del indicador mensual en i+1
                        row_data[f"{m} Op 1"] = df.iloc[i+3, col_idx] if i+3 < len(df) else ""
                        row_data[f"{m} Op 2"] = df.iloc[i+5, col_idx] if i+5 < len(df) else ""

                # AI: Cumplimiento Proyectado
                if cols["proy"] is not None:
                    row_data["Cumplimiento Proyectado Op 1"] = df.iloc[i+3, cols["proy"]] if i+3 < len(df) else ""
                    row_data["Cumplimiento Proyectado Op 2"] = df.iloc[i+5, cols["proy"]] if i+5 < len(df) else ""

                # AJ: % Cumplimiento de Meta
                if cols["cump_meta"] is not None:
                    row_data["% Cumplimiento de Meta"] = df.iloc[i+3, cols["cump_meta"]] if i+3 < len(df) else ""

                # AK: Medios de Verificación (Fila 13)
                if cols["medios"] is not None:
                    row_data["Medios de Verificacion"] = df.iloc[i, cols["medios"]]

                # AM: Instrumentos de Gestión (Fila 13)
                if cols["inst"] is not None:
                    row_data["Instrumentos de Gestión Asociados"] = df.iloc[i, cols["inst"]]

                self.all_extracted_data.append(row_data)

    def save_and_style(self):
        """Genera el Excel con dos hojas y aplica el estilo visual solicitado."""
        if not self.all_extracted_data:
            print("No hay datos para guardar.")
            return

        final_df = pd.DataFrame(self.all_extracted_data)

        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name="Carga Bruta", index=False)
            final_df.to_excel(writer, sheet_name="Hoja Estilizada", index=False)
            
            ws = writer.book["Hoja Estilizada"]
            
            # --- ESTILOS ---
            # Colores
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            separator_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Bordes
            thin_border = Side(border_style="thin", color="000000")
            thick_border = Side(border_style="medium", color="000000")
            all_borders = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            
            # Formatear encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(top=thick_border, left=thin_border, right=thin_border, bottom=thick_border)

            # Formatear filas y distinguir entre hojas (separadores horizontales)
            last_origin = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                current_origin = final_df.iloc[row_idx-2]["ORIGEN"]
                
                # Si cambia el origen (ej. de CDC a PMG), aplicamos un estilo de separación
                is_new_section = current_origin != last_origin
                
                for cell in row:
                    cell.border = all_borders
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    if is_new_section:
                        cell.fill = separator_fill
                        cell.border = Border(top=thick_border, left=thin_border, right=thin_border, bottom=thin_border)
                
                last_origin = current_origin

            # Ajustar ancho de columnas (automático básico)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)

        print(f"\n--- PROCESO FINALIZADO ---")
        print(f"Archivo generado con éxito: {self.output_file}")

# =============================================================================
# EJECUCIÓN
# =============================================================================

if __name__ == "__main__":
    # Cambia esto por el nombre de tu archivo maestro
    archivo_entrada = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
    
    if not os.path.exists(archivo_entrada):
        print(f"Error: El archivo '{archivo_entrada}' no existe en el directorio.")
    else:
        parser = IPSParserPhase1(archivo_entrada)
        for sheet in parser.sheets_to_process:
            parser.process_sheet(sheet)
        parser.save_and_style()