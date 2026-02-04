"""
=============================================================================
 PROGRAMA: IPS_PARSER_25
 VERSIÓN:  v1.0.0 (Fase 1: Extracción Maestra)
 FECHA:    Febrero 2026
 
 DESCRIPCIÓN:
 Módulo de extracción pura. Lee el archivo maestro de Proyecciones,
 detecta bloques de indicadores complejos (celdas combinadas, filas relativas)
 y genera dos salidas estandarizadas: Bruta (CSV-like) y Estilizada (Reporte).
=============================================================================
"""

import pandas as pd
import os
import re
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Silenciar alertas de compatibilidad de Excel
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)

# =============================================================================
# 1. CONFIGURACIÓN
# =============================================================================

ARCHIVOS = {
    # Nombre exacto del archivo que me indicaste
    "MAESTRO": "Proyecciones Indicadores 2025 - División Planificación (1).xlsx",
    "ESTILIZADA": "Planilla_Estilizada_2025.xlsx",
    "BRUTA": "Planilla_Bruta_2025.xlsx"
}

# Hojas a buscar dentro del archivo maestro
HOJAS_CONFIG = {
    "CDC 2025": "CDC", 
    "PMG 2025": "PMG", 
    "Riesgos 2025": "Riesgos"
}

# =============================================================================
# 2. UTILS DE LIMPIEZA
# =============================================================================

def limpiar_porcentaje(val):
    """Convierte strings de porcentaje o decimales a número base 100."""
    if pd.isna(val) or str(val).strip() == "": return 0
    s = str(val).replace('%', '').strip()
    # Reemplazo seguro de coma decimal
    s = s.replace(',', '.')
    try:
        num = float(s)
        # Si viene como 0.85 (Excel puro), lo pasamos a 85
        # Si viene como 85 (texto), se queda en 85
        if num < 1.05 and num > -1.05 and num != 0: # Heurística: si es pequeño, es decimal
            return num * 100
        return num
    except: return 0

def limpiar_numero(val):
    """Limpia separadores de miles y deja decimales con punto."""
    if pd.isna(val) or str(val).strip() == "": return 0
    s = str(val).strip()
    # Eliminar puntos de miles si existen comas decimales
    if "," in s: 
        s = s.replace('.', '').replace(',', '.')
    else:
        # Si solo hay puntos, asumir que son decimales si el número es pequeño, 
        # o miles si es grande? Mejor estandarizar: en Chile punto es mil.
        # Eliminamos punto.
        s = s.replace('.', '')
    try: return float(s)
    except: return 0

def limpiar_texto(val):
    if pd.isna(val): return ""
    # Quitar saltos de línea internos y espacios extra
    return str(val).strip().replace("\n", " ").replace("\r", " ")

def limpiar_op_inicio(val):
    """Quita paréntesis al inicio (ej: '(Sumatoria...' -> 'Sumatoria...')"""
    txt = limpiar_texto(val)
    if txt.startswith("("): return txt[1:].strip()
    return txt

def limpiar_op_final(val):
    """Quita fórmula al final (ej: '...total)*100' -> '...total')"""
    txt = limpiar_texto(val)
    # Regex para quitar )*100 al final
    return re.sub(r'\)\s*\*100$', '', txt).strip()

def detectar_encabezados(df):
    """
    Escanea las primeras 30 filas buscando la fila que contiene 'INDICADOR' y 'NÚMERO'.
    Retorna el índice de la fila y la lista de encabezados.
    """
    for i in range(min(30, len(df))):
        try:
            fila = [str(x).upper().strip() for x in df.iloc[i].tolist()]
            # Criterio flexible: Debe tener INDICADOR y (NÚMERO o CODIGO)
            if any("INDICADOR" in x for x in fila) and any(y in fila for y in ["NÚMERO", "NUMERO", "CODIGO"]):
                return i, df.iloc[i].astype(str).tolist()
        except: continue
    return None, None

# =============================================================================
# 3. MOTOR DE EXTRACCIÓN (LÓGICA POSICIONAL + MAPEO)
# =============================================================================

def procesar_hoja_maestra(ruta_archivo, nombre_hoja_excel, etiqueta_log):
    print(f"   -> Analizando hoja: '{nombre_hoja_excel}' ({etiqueta_log})...")
    
    try:
        # Leer hoja completa sin header para buscarlo manualmente
        df_raw = pd.read_excel(ruta_archivo, sheet_name=nombre_hoja_excel, header=None)
    except Exception as e:
        print(f"      [AVISO] No se encontró o leyó la hoja '{nombre_hoja_excel}'.")
        return None

    # 1. Detectar Ancla (Header)
    idx_header, encabezados = detectar_encabezados(df_raw)
    if idx_header is None:
        print("      [ERROR] No se detectó fila de encabezados válida.")
        return None

    # 2. Mapeo Dinámico de Columnas
    # Crea un diccionario { "NOMBRE COLUMNA": indice }
    mapa_cols = {str(txt).upper().strip().replace("\n", " "): i for i, txt in enumerate(encabezados)}
    
    def get_col_idx(claves):
        """Busca el índice de columna por lista de palabras clave."""
        if isinstance(claves, str): claves = [claves]
        for k in claves:
            k_upper = k.upper()
            for nombre_real, idx in mapa_cols.items():
                if k_upper in nombre_real: return idx
        return None

    # Índices Clave
    IDX_NUM = get_col_idx(["NÚMERO", "NUMERO", "CODIGO"])
    IDX_IND = get_col_idx(["INDICADOR"])
    IDX_OP_DESC = get_col_idx(["OPERANDOS"]) # Descripción
    # A veces se llama "Operandos Estimados", a veces "Estimados", a veces "Meta"
    IDX_OP_EST = get_col_idx(["OPERANDOS ESTIMADOS", "ESTIMADOS", "META 2025"]) 
    
    # Metadatos Extra
    idx_prod = get_col_idx("PRODUCTO")
    idx_form = get_col_idx("FORMULA")
    idx_unidad = get_col_idx("UNIDAD")
    idx_resp = get_col_idx("RESPONSABLE")
    idx_gestor = get_col_idx("GESTOR")
    idx_sup = get_col_idx("SUPERVISORES")
    idx_medios = get_col_idx("MEDIOS")
    idx_ponderador = get_col_idx("PONDERADOR")

    # Meses (Búsqueda dinámica)
    meses_std = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEPT", "OCT", "NOV", "DIC"]
    mapa_meses = {}
    for m in meses_std:
        # Buscamos "ENE.", "FEB." o "ENE" solo
        col = get_col_idx([f"{m}.", m])
        if col is not None: mapa_meses[m] = col

    # 3. Iteración y Extracción ("Francotirador")
    extracted_rows = []
    
    # Iteramos desde la fila siguiente al encabezado
    # Usamos len(df) - 5 para asegurar que tenemos espacio para leer i+5
    for i in range(idx_header + 1, len(df_raw) - 5):
        val_num = str(df_raw.iloc[i, IDX_NUM]).strip()
        
        # Detector de Indicador: No vacio, no es "NÚMERO", tiene longitud razonable (ej "3.1")
        if pd.notna(df_raw.iloc[i, IDX_NUM]) and val_num != "NÚMERO" and len(val_num) >= 3:
            
            try:
                # --- EXTRACCIÓN DE DATOS BASE ---
                fila = {}
                fila["TIPO_ORIGEN"] = etiqueta_log
                fila["NÚMERO"] = val_num
                fila["INDICADOR"] = limpiar_texto(df_raw.iloc[i, IDX_IND])
                
                # Metadatos (si existen las columnas)
                fila["PRODUCTO"] = limpiar_texto(df_raw.iloc[i, idx_prod]) if idx_prod else ""
                fila["FORMULA"] = limpiar_texto(df_raw.iloc[i, idx_form]) if idx_form else ""
                fila["UNIDAD"] = limpiar_texto(df_raw.iloc[i, idx_unidad]) if idx_unidad else ""
                fila["RESPONSABLE"] = limpiar_texto(df_raw.iloc[i, idx_resp]) if idx_resp else ""
                fila["GESTOR"] = limpiar_texto(df_raw.iloc[i, idx_gestor]) if idx_gestor else ""
                fila["SUPERVISOR"] = limpiar_texto(df_raw.iloc[i, idx_sup]) if idx_sup else ""
                fila["MEDIOS VERIFICACIÓN"] = limpiar_texto(df_raw.iloc[i, idx_medios]) if idx_medios else ""
                
                # Ponderador (si existe)
                pond_val = df_raw.iloc[i, idx_ponderador] if idx_ponderador else 0
                fila["Ponderador (%)"] = limpiar_porcentaje(pond_val)

                # --- EXTRACCIÓN POSICIONAL (Bloque Vertical) ---
                # Fila i   -> Desc Op1
                # Fila i+1 -> Meta 2025 (en col Estimados)
                # Fila i+3 -> Desc Op2 (en col Operandos), Valor Anual Op1 (en col Estimados)
                # Fila i+5 -> Valor Anual Op2 (en col Estimados)

                fila["Desc. Op1"] = limpiar_op_inicio(df_raw.iloc[i, IDX_OP_DESC])
                
                # Meta
                meta_raw = df_raw.iloc[i+1, IDX_OP_EST]
                fila["Meta 2025 (%)"] = limpiar_porcentaje(meta_raw)
                
                # Op1 Valor y Op2 Desc
                fila["Est. Meta Op1"] = limpiar_numero(df_raw.iloc[i+3, IDX_OP_EST])
                fila["Desc. Op2"] = limpiar_op_final(df_raw.iloc[i+3, IDX_OP_DESC])
                
                # Op2 Valor
                fila["Est. Meta Op2"] = limpiar_numero(df_raw.iloc[i+5, IDX_OP_EST])

                # --- CICLO MENSUAL ---
                for mes_nombre, col_idx in mapa_meses.items():
                    # Fila i+1 -> Valor Indicador %
                    # Fila i+3 -> Valor Op1
                    # Fila i+5 -> Valor Op2
                    fila[f"{mes_nombre} Ind (%)"] = limpiar_porcentaje(df_raw.iloc[i+1, col_idx])
                    fila[f"{mes_nombre} Op1"] = limpiar_numero(df_raw.iloc[i+3, col_idx])
                    fila[f"{mes_nombre} Op2"] = limpiar_numero(df_raw.iloc[i+5, col_idx])

                extracted_rows.append(fila)

            except Exception as e:
                # Si falla una fila, logueamos pero seguimos con la siguiente
                # print(f"      [WARN] Fila {i} saltada: {e}")
                continue

    return pd.DataFrame(extracted_rows)

# =============================================================================
# 4. ESTILOS VISUALES (Salida Estilizada)
# =============================================================================

def aplicar_estilos_planillas(nombre_archivo):
    print(f"   -> Aplicando estilos visuales a {nombre_archivo}...")
    try:
        wb = load_workbook(nombre_archivo)
        
        # Definición de estilos
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
        font_header = Font(bold=True, color="FFFFFF")
        border_thin = Side(border_style="thin", color="000000")
        border_box = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formatear Encabezados
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.border = border_box
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Ajustar anchos
            ws.column_dimensions['B'].width = 12  # Número
            ws.column_dimensions['C'].width = 50  # Indicador
            ws.column_dimensions['D'].width = 20  # Producto
            ws.column_dimensions['H'].width = 30  # Responsable
            ws.column_dimensions['L'].width = 40  # Desc Op1
            ws.column_dimensions['N'].width = 40  # Desc Op2
            
            # Congelar paneles
            ws.freeze_panes = "A2"

        wb.save(nombre_archivo)
    except Exception as e:
        print(f"      [ERROR] Falló el estilizado: {e}")

# =============================================================================
# 5. ORQUESTACIÓN PRINCIPAL
# =============================================================================

def ejecutar_fase_1():
    print("\n=== IPS_PARSER_25 v1.0.0: FASE 1 (EXTRACCIÓN) ===")
    
    if not os.path.exists(ARCHIVOS["MAESTRO"]):
        print(f"[ERROR CRÍTICO] No se encuentra el archivo maestro: '{ARCHIVOS['MAESTRO']}'")
        return

    dataframes_listos = {}
    
    # 1. Procesar cada hoja configurada
    for hoja_excel, etiqueta in HOJAS_CONFIG.items():
        df = procesar_hoja_maestra(ARCHIVOS["MAESTRO"], hoja_excel, etiqueta)
        if df is not None and not df.empty:
            dataframes_listos[etiqueta] = df
            print(f"      [OK] Hoja '{hoja_excel}': {len(df)} indicadores extraídos.")
        else:
            print(f"      [WARN] Hoja '{hoja_excel}': Sin datos válidos.")

    if not dataframes_listos:
        print("[FIN] No se generó información. Revisa el nombre del archivo o las hojas.")
        return

    # 2. Guardar Salida BRUTA (Sin formato, solo datos)
    print(f"   -> Generando: {ARCHIVOS['BRUTA']}")
    with pd.ExcelWriter(ARCHIVOS["BRUTA"], engine='openpyxl') as writer:
        for etiq, df in dataframes_listos.items():
            df.to_excel(writer, sheet_name=etiq, index=False)

    # 3. Guardar Salida ESTILIZADA (Formato legible)
    print(f"   -> Generando: {ARCHIVOS['ESTILIZADA']}")
    with pd.ExcelWriter(ARCHIVOS["ESTILIZADA"], engine='openpyxl') as writer:
        for etiq, df in dataframes_listos.items():
            df.to_excel(writer, sheet_name=etiq, index=False)
    
    aplicar_estilos_planillas(ARCHIVOS["ESTILIZADA"])

    print("\n=== PROCESO COMPLETADO EXITOSAMENTE ===")
    print(f"1. Datos Crudos: {ARCHIVOS['BRUTA']}")
    print(f"2. Datos Visuales: {ARCHIVOS['ESTILIZADA']}")

if __name__ == "__main__":
    ejecutar_fase_1()