import pandas as pd
import os
import sys
import glob
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_ADP_PARSER_v1.1.2 - SIG_DATOS_VARIABLES CON ARCHIVO Y HOJA (12 COLUMNAS)
# =============================================================================

class IPS_ADP_Parser:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "ADP_CONSOLIDADO_v20260226-20-15.xlsx")
        
        self.tree_proy = {} 
        self.tree_sig = {} 
        self.flat_data_proy = [] 
        self.flat_data_sig = [] 
        self.variable_data_sig = [] 
        
        self.opt_format_percent = True
        self.valid_sheet_keywords = ["PROYEC", "SIG"]
        
        self.meses_fijos = [f"{m}-25" for m in ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]] + \
                           [f"{m}-26" for m in ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]]

        self.ordered_keys = [
            "ARCHIVO", "HOJA", "EQUIPO", "TIPO INDICADOR", "NÚMERO",
            "PRODUCTO O PROCESO ESPECIFICO", "INDICADOR", "DIMENSION", "AMBITO",
            "FORMULA", "TIPO FORMULA", "UNIDAD", "RESPONSABLE", "GESTOR", "SUPERVISORES",
            "META", "PONDERACIÓN",
            "Descripción Operando 1", "Descripción Operando 2", "Meta Operando 1", "Meta Operando 2"
        ]
        for mes in self.meses_fijos:
            self.ordered_keys.extend([f"{mes} Op 1", f"{mes} Op 2", f"{mes} Acum Op 1", f"{mes} Acum Op 2"])
        self.ordered_keys.extend([
            "Efectivo Op 1", "Efectivo Op 2", "% Cumplimiento de Meta",
            "Medios de control", "Control de cambios", "Instrumentos de Gestion Asociados"
        ])

        self.mapa_meses_num = {
            "Ene": "01", "Feb": "02", "Mar": "03", "Abr": "04", 
            "May": "05", "Jun": "06", "Jul": "07", "Ago": "08", 
            "Sep": "09", "Oct": "10", "Nov": "11", "Dic": "12"
        }

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA ADP v1.1.2 (Transaccional 12 Cols)")
        print("="*60)
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        print(f"\n[OK] Configuración guardada.")
        print("-" * 60)

    def print_summary_and_exit(self):
        print("\n" + "="*60)
        print("   RESUMEN FINAL")
        print("="*60)
        print(f"  * Registros 'Proyecciones - Bruta': {len(self.flat_data_proy)}")
        print(f"  * Registros 'SIG - Bruta':          {len(self.flat_data_sig)}")
        print(f"  * Registros 'SIG_DATOS_VARIABLES':  {len(self.variable_data_sig)}")
        print("-" * 60)
        if self.flat_data_proy or self.flat_data_sig: 
            self.export_excel()
        else: 
            print("[AVISO] No se generó archivo de salida.")
        sys.exit()

    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "": return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix, core, f_type = "", f_clean, "CUOCIENTE"
        if match:
            suffix, core, f_type = match.group(1), f_clean[:match.start()].strip(), "PORCENTAJE"
        if core.startswith("(") and core.endswith(")"): core = core[1:-1].strip()
        return core + suffix, f_type

    def transform_percentage(self, val):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or str(val).strip() == "": return ""
        try:
            num = float(val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    def extract_month_name(self, val):
        if pd.isna(val): return ""
        val_str = str(val).strip().lower()
        if not val_str: return ""
        try:
            dt = pd.to_datetime(val, errors='raise')
            mes_idx = dt.month - 1
            year = str(dt.year)[-2:]
            meses_base = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
            return f"{meses_base[mes_idx]}-{year}"
        except: pass
        mapa_meses = {
            "ene": "Ene", "feb": "Feb", "mar": "Mar", "abr": "Abr", 
            "may": "May", "jun": "Jun", "jul": "Jul", "ago": "Ago", 
            "sep": "Sep", "oct": "Oct", "nov": "Nov", "dic": "Dic"
        }
        if "-" in val_str:
            parts = val_str.split("-")
            if len(parts) == 2:
                m, y = parts[0].strip(), parts[1].strip()
                if m in mapa_meses and y.isdigit(): return f"{mapa_meses[m]}-{y[-2:]}"
                elif y in mapa_meses and m.isdigit(): return f"{mapa_meses[y]}-{m[-2:]}"
        return val_str

    def get_excel_files(self):
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$") and "ADP_CONSOLIDADO" not in f]
        return valid_files

    def process_folder(self):
        files = self.get_excel_files()
        if not files:
            print("[ERROR] No hay archivos válidos.")
            return
            
        self.configure()
        
        for idx_file, file_path in enumerate(files):
            file_name_correct = os.path.basename(file_path)
            
            self.tree_proy[file_name_correct] = {}
            self.tree_sig[file_name_correct] = {}
            
            print(f"\n>>> PROCESANDO ({idx_file + 1}/{len(files)}): {file_name_correct}")
            
            equipo_name = "No definido"
            if "ADP" in file_name_correct.upper():
                parts = re.split(r'(?i)ADP', file_name_correct, maxsplit=1)
                if len(parts) > 1:
                    suffix = parts[1]
                    suffix = re.sub(r'(?i)\.xlsx?', '', suffix) 
                    suffix = re.sub(r'\d+', '', suffix) 
                    equipo_name = suffix.strip(" -_").title()

            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except: continue

            for sheet in sheet_names:
                s_upper = sheet.upper()
                if "CONSOLIDADO" in s_upper: continue
                if not any(k in s_upper for k in self.valid_sheet_keywords): continue

                is_proy = "PROYEC" in s_upper
                is_sig = "SIG" in s_upper

                df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                
                responsable = "No aplica"
                for r_idx in range(5, 12):
                    if r_idx >= len(df): break
                    for c_idx in range(4):
                        val = str(df.iloc[r_idx, c_idx])
                        if "," in val and "Fecha" not in val and "Elaborado" not in val:
                            parts = val.split(",", 1)
                            if len(parts) > 1:
                                resp_clean = re.sub(r'\(.*?\)', '', parts[1]).strip()
                                responsable = resp_clean.title()
                                break
                    if responsable != "No aplica": break

                h_idx = None
                for idx in range(min(15, len(df))):
                    row_vals = [str(x).upper().strip() for x in df.iloc[idx].values if pd.notna(x)]
                    if ("NUMERO" in row_vals or "NÚMERO" in row_vals) or ("INDICADOR" in row_vals and "FORMULA" in row_vals):
                        h_idx = idx
                        break
                
                if h_idx is None: 
                    print(f"  [!] No se encontró cabecera válida en la hoja: {sheet}. Saltando.")
                    continue 

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                def find_col(keywords):
                    for i, h in enumerate(headers):
                        if any(k in str(h).replace('\n', ' ').upper() for k in keywords): return i
                    return None

                col_num = find_col(["NUMERO", "NÚMERO"])     
                col_ind = find_col(["INDICADOR"])            
                
                if col_num is None:
                    col_num = 0
                    if col_ind == 0:
                        col_ind = 1

                col_form = find_col(["FORMULA"])             
                col_pond = find_col(["PONDER"]) 
                col_meta = find_col(["META"])                
                col_operandos = find_col(["OPERANDOS"])
                col_meta_est = find_col(["ESTIMADOS META"])
                col_efectivo = find_col(["EFECTIVO"]) 
                col_porc_cump = find_col(["% CUMPLIMIENTO", "CUMPLIMIENTO DE META"])

                month_cols = {} 
                month_layout = {m: [] for m in self.meses_fijos} 
                current_m = None
                
                start_month_idx = 7 if len(headers) > 7 else (col_operandos + 2 if col_operandos else 0)
                for c_idx in range(start_month_idx, len(headers)):
                    h_name = headers[c_idx]
                    h_str = str(h_name).replace('\n', ' ').upper().strip()
                    if h_str in ["NAN", "", "NONE"] or "META" in h_str or "EFECTIVO" in h_str or "CUMPLIMIENTO" in h_str: continue
                        
                    if "ACUM" in h_str:
                        if current_m: month_layout[current_m].append(c_idx)
                    else:
                        generic_month = self.extract_month_name(h_name)
                        if generic_month in self.meses_fijos:
                            month_cols[generic_month] = c_idx
                            current_m = generic_month

                sheet_rows = []
                count_rows = 0
                
                def get_v(r_idx, c_idx, default=""):
                    if c_idx is None or r_idx >= len(df): return default
                    val = df.iloc[r_idx, c_idx]
                    return val if pd.notna(val) else default

                i = h_idx + 1
                while i < len(df):
                    raw_num = get_v(i, col_num)
                    str_num = str(raw_num).strip()
                    
                    if str_num != "" and str_num.upper() not in ["NAN", "NONE"]:
                        clean_formula, type_formula = self.analyze_formula(get_v(i, col_form))
                        
                        temp_data = {
                            "ARCHIVO": file_name_correct, 
                            "HOJA": sheet, 
                            "EQUIPO": equipo_name,
                            "TIPO INDICADOR": f"ADP({sheet})",
                            "NÚMERO": str_num,
                            "PRODUCTO O PROCESO ESPECIFICO": "",
                            "INDICADOR": get_v(i, col_ind),
                            "DIMENSION": "",
                            "AMBITO": "",
                            "FORMULA": clean_formula,   
                            "TIPO FORMULA": type_formula,
                            "UNIDAD": equipo_name,
                            "RESPONSABLE": responsable,
                            "GESTOR": "",
                            "SUPERVISORES": "",
                            "META": self.transform_percentage(get_v(i, col_meta)),
                            "PONDERACIÓN": self.transform_percentage(get_v(i, col_pond)), 
                            "Descripción Operando 1": get_v(i + 0, col_operandos),
                            "Descripción Operando 2": get_v(i + 3, col_operandos),
                            "Meta Operando 1": get_v(i + 3, col_meta_est),
                            "Meta Operando 2": get_v(i + 5, col_meta_est),
                            "Medios de control": "",
                            "Control de cambios": "",
                            "Instrumentos de Gestion Asociados": ""
                        }

                        for mes in self.meses_fijos:
                            if mes in month_cols:
                                col_idx = month_cols[mes]
                                val_op1 = get_v(i + 3, col_idx)
                                val_op2 = get_v(i + 5, col_idx)
                                
                                # **LÓGICA TRANSACCIONAL EXCLUSIVA PARA SIG**
                                if is_sig:
                                    m_text, y_text = mes.split("-")
                                    ano_num = f"20{y_text}"
                                    mes_num = self.mapa_meses_num.get(m_text, m_text)

                                    if str(val_op1).strip() != "":
                                        self.variable_data_sig.append({
                                            "AÑO": ano_num, "MES": mes_num, "VARIABLE_COD": f"{str_num}_A",
                                            "CENTRO_RESP_COD": equipo_name, "COD_REGION": 0,
                                            "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "",
                                            "VALOR_TOTAL": val_op1,
                                            "ARCHIVO": file_name_correct, "HOJA": sheet
                                        })
                                    if str(val_op2).strip() != "":
                                        self.variable_data_sig.append({
                                            "AÑO": ano_num, "MES": mes_num, "VARIABLE_COD": f"{str_num}_B",
                                            "CENTRO_RESP_COD": equipo_name, "COD_REGION": 0,
                                            "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "",
                                            "VALOR_TOTAL": val_op2,
                                            "ARCHIVO": file_name_correct, "HOJA": sheet
                                        })
                            else:
                                val_op1, val_op2 = "", ""
                                
                            temp_data[f"{mes} Op 1"] = val_op1
                            temp_data[f"{mes} Op 2"] = val_op2
                            
                            acums = month_layout.get(mes, [])
                            if acums:
                                val_acum_op1 = get_v(i + 3, acums[0])
                                val_acum_op2 = get_v(i + 5, acums[0])
                            else:
                                val_acum_op1, val_acum_op2 = "", ""
                                
                            temp_data[f"{mes} Acum Op 1"] = val_acum_op1
                            temp_data[f"{mes} Acum Op 2"] = val_acum_op2

                        temp_data["Efectivo Op 1"] = get_v(i + 3, col_efectivo)
                        temp_data["Efectivo Op 2"] = get_v(i + 5, col_efectivo)

                        cump_val = ""
                        for offset in range(6):
                            val = get_v(i + offset, col_porc_cump)
                            if str(val).strip() != "":
                                cump_val = val
                                break
                        temp_data["% Cumplimiento de Meta"] = self.transform_percentage(cump_val)

                        sheet_rows.append(temp_data)
                        
                        if is_proy:
                            self.flat_data_proy.append(temp_data)
                        elif is_sig:
                            self.flat_data_sig.append(temp_data)

                        count_rows += 1
                        i += 6 
                    else:
                        i += 1 

                if is_proy:
                    self.tree_proy[file_name_correct][sheet] = sheet_rows
                elif is_sig:
                    self.tree_sig[file_name_correct][sheet] = sheet_rows
                    
                print(f"  -> {count_rows} indicadores procesados [Hoja: {sheet}]")

        self.print_summary_and_exit()

    def _render_estilizada(self, ws, tree_data):
        styles = {
            'file': PatternFill("solid", fgColor="000000"),
            'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"),
            'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True),
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }
        
        estilizada_keys = [k for k in self.ordered_keys if k not in ["ARCHIVO", "HOJA"]]
        row_idx = 1
        FULL_WIDTH = len(estilizada_keys)

        for fname, sheets in tree_data.items():
            if not any(sheets.values()): continue
            c = ws.cell(row=row_idx, column=1, value=f"ARCHIVO: {fname}")
            c.fill = styles['file']; c.font = styles['w_font']; c.alignment = Alignment('center')
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
            row_idx += 1
            
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws.cell(row=row_idx, column=1, value=f"HOJA: {sname}")
                c.fill = styles['sheet']; c.font = styles['w_font']
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
                row_idx += 1
                
                for c_i, k in enumerate(estilizada_keys, 1):
                    c = ws.cell(row=row_idx, column=c_i, value=k)
                    c.fill = styles['head']; c.font = styles['b_font']; c.border = styles['border']
                row_idx += 1
                
                for r in rows:
                    for c_i, k in enumerate(estilizada_keys, 1):
                        c = ws.cell(row=row_idx, column=c_i, value=r.get(k, ""))
                        c.border = styles['border']; c.alignment = Alignment(wrapText=True, vertical='top')
                    row_idx += 1
                row_idx += 1

        for i in range(1, FULL_WIDTH + 2): ws.column_dimensions[get_column_letter(i)].width = 22

    def export_excel(self):
        print(f"\n{'='*60}\nGUARDANDO ARCHIVO: {self.output_file}\n{'='*60}")
        wb = Workbook()
        default_sheet = wb.active
        
        # 1. Proyecciones - Bruta
        ws_proy_bruta = wb.create_sheet("Proyecciones - Bruta")
        ws_proy_bruta.append(self.ordered_keys)
        for r in self.flat_data_proy: ws_proy_bruta.append([r.get(k, "") for k in self.ordered_keys])
            
        # 2. Proyecciones - Estilizada
        ws_proy_est = wb.create_sheet("Proyecciones - Estilizada")
        self._render_estilizada(ws_proy_est, self.tree_proy)
        
        # 3. SIG - Bruta
        ws_sig_bruta = wb.create_sheet("SIG - Bruta")
        ws_sig_bruta.append(self.ordered_keys)
        for r in self.flat_data_sig: ws_sig_bruta.append([r.get(k, "") for k in self.ordered_keys])
            
        # 4. SIG - Estilizada
        ws_sig_est = wb.create_sheet("SIG - Estilizada")
        self._render_estilizada(ws_sig_est, self.tree_sig)

        # 5. SIG_DATOS_VARIABLES (12 Columnas)
        ws_vars = wb.create_sheet("SIG_DATOS_VARIABLES")
        headers_vars = [
            "AÑO", "MES", "VARIABLE_COD", "CENTRO_RESP_COD", "COD_REGION", 
            "VALOR_M", "VALOR_F", "VALOR_S", "VALOR_J", "VALOR_TOTAL", 
            "ARCHIVO", "HOJA"
        ]
        ws_vars.append(headers_vars)
        for row in self.variable_data_sig: 
            ws_vars.append([row.get(k, "") for k in headers_vars])
            
        wb.remove(default_sheet)

        while True:
            try:
                wb.save(self.output_file)
                print(f"[ÉXITO] Archivo generado: {self.output_file}")
                break
            except Exception as e:
                print(f"\n[ERROR AL GUARDAR] {e}")
                input("  >> Cierra el archivo si está abierto y presiona Enter para reintentar...")

if __name__ == "__main__":
    try:
        print("INICIANDO PROCESADOR MASIVO ADP v1.1.2 (SIG Datos Variables - 12 Cols)")
        path = input("Ruta de la carpeta (Enter para actual): ").strip() or os.getcwd()
        if os.path.isdir(path):
            parser = IPS_ADP_Parser(path)
            parser.process_folder()
        else: print("Ruta inválida.")
    except Exception as e:
        print(f"Error fatal: {e}")
        input("Enter para salir.")