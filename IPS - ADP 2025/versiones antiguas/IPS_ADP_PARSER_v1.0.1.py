import pandas as pd
import os
import sys
import glob
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_ADP_PARSER_v1.0.1 - PROCESAMIENTO MASIVO CONVENIOS ADP
# =============================================================================

class IPS_ADP_Parser:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_ADP_CONSOLIDADO_V1.0.1.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.variable_data = [] 
        
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible'
        
        # Hojas a procesar permitidas
        self.valid_sheet_keywords = ["PROYECCI", "SIG"]

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA ADP v1.0.1")
        print("="*60)
        
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        
        print("\n2. ¿Cómo manejar filas OCULTAS en todos los archivos?")
        print("   [v] Automático: Procesar SOLO VISIBLES (Recomendado).")
        print("   [t] Automático: Procesar TODO.")
        resp_h = input("   >> Elija opción (Enter=v): ").lower().strip()
        
        if resp_h == 't': self.opt_hidden_strategy = 'all'
        else: self.opt_hidden_strategy = 'visible'

        print(f"\n[OK] Configuración guardada. Estrategia Ocultos: {self.opt_hidden_strategy.upper()}")
        print("-" * 60)

    def print_summary_and_exit(self):
        print("\n" + "="*60)
        print("   RESUMEN FINAL")
        print("="*60)
        print(f"  * Registros 'Carga Bruta' (Indicadores): {len(self.flat_data)}")
        print(f"  * Registros 'DATOS_VARIABLE' (Meses):    {len(self.variable_data)}")
        print("-" * 60)
        
        if self.flat_data:
            self.export_excel()
        else:
            print("[AVISO] No se generó archivo de salida (sin datos o formato incorrecto).")
        sys.exit()

    def is_fully_enclosed_by_parens(self, text):
        if not text.startswith("(") or not text.endswith(")"): return False
        balance = 0
        for i, char in enumerate(text):
            if char == '(': balance += 1
            elif char == ')': balance -= 1
            if balance == 0 and i < len(text) - 1: return False
        return balance == 0

    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "":
            return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix, core, f_type = "", f_clean, "CUOCIENTE"
        if match:
            suffix = match.group(1)
            core = f_clean[:match.start()].strip()
            f_type = "PORCENTAJE"
        while self.is_fully_enclosed_by_parens(core):
            core = core[1:-1].strip()
        return core + suffix, f_type

    def transform_percentage(self, val):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or str(val).strip() == "": return ""
        try:
            num = float(val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    def format_month(self, val):
        if pd.isna(val): return ""
        val_str = str(val).strip().lower().replace(".", "")
        try:
            dt = pd.to_datetime(val)
            return dt.strftime("%Y-%m")
        except: pass
        
        meses = {"ene": "01", "feb": "02", "mar": "03", "abr": "04", "may": "05", "jun": "06", 
                 "jul": "07", "ago": "08", "sep": "09", "oct": "10", "nov": "11", "dic": "12"}
        if "-" in val_str:
            parts = val_str.split("-")
            if len(parts) == 2:
                m, y = parts[0], parts[1]
                if m in meses and y.isdigit():
                    year = "20" + y if len(y) == 2 else y
                    return f"{year}-{meses[m]}"
                elif y in meses and m.isdigit():
                    year = "20" + m if len(m) == 2 else m
                    return f"{year}-{meses[y]}"
        return str(val).strip()

    def get_excel_files(self):
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$") and "IPS_ADP_CONSOLIDADO" not in f and "IPS_CONSOLIDADO" not in f]
        if not valid_files:
            print(f"[ERROR] Carpeta vacía o sin Excel: {self.folder_path}")
            sys.exit()
        return valid_files

    def get_hidden_rows(self, file_path, sheet_name):
        try:
            wb = load_workbook(file_path, read_only=False, data_only=True)
            if sheet_name not in wb.sheetnames: return set()
            ws = wb[sheet_name]
            hidden = set()
            for row_idx, row_dim in ws.row_dimensions.items():
                if row_dim.hidden: hidden.add(row_idx - 1) 
            wb.close()
            return hidden
        except: return set()

    def process_folder(self):
        files = self.get_excel_files()
        self.configure()
        
        for idx_file, file_path in enumerate(files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            print(f"\n>>> PROCESANDO ({idx_file + 1}/{len(files)}): {file_name}")
            
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except Exception as e:
                print(f"  [ERROR] Archivo corrupto: {e}")
                continue

            for sheet in sheet_names:
                s_upper = sheet.upper()
                if "CONSOLIDADO" in s_upper:
                    print(f"  -> Omitiendo hoja: {sheet} (Consolidado ignorado)")
                    continue
                if not any(k in s_upper for k in self.valid_sheet_keywords):
                    print(f"  -> Omitiendo hoja: {sheet} (No es Proyección ni SIG)")
                    continue

                hidden_rows = self.get_hidden_rows(file_path, sheet)
                ignored_rows = set()
                if hidden_rows and self.opt_hidden_strategy == 'visible':
                    ignored_rows = hidden_rows

                try: df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except: continue

                h_idx = None
                for idx in range(min(15, len(df))):
                    if idx in ignored_rows: continue
                    row_vals = [str(x).upper().strip() for x in df.iloc[idx].values if pd.notna(x)]
                    if "NUMERO" in row_vals or "NÚMERO" in row_vals:
                        h_idx = idx
                        break
                
                if h_idx is None:
                    print(f"  [!] No se encontró cabecera 'NUMERO' en {sheet}. Saltando.")
                    continue 

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                def find_col(keywords):
                    for i, h in enumerate(headers):
                        if any(k in str(h).upper() for k in keywords): return i
                    return None

                col_num = find_col(["NUMERO", "NÚMERO"])     
                col_ind = find_col(["INDICADOR"])            
                col_form = find_col(["FORMULA"])             
                col_pond = find_col(["PONDERACIÓN", "PONDERACIÖN", "PONDERACION"]) 
                col_meta = find_col(["META"])                
                col_operandos = find_col(["OPERANDOS"])      

                month_cols = {}
                start_month_idx = 7 if len(headers) > 7 else (col_operandos + 2 if col_operandos else 0)
                
                for c_idx in range(start_month_idx, len(headers)):
                    h_name = headers[c_idx]
                    h_str = str(h_name).upper().strip()
                    if h_str in ["NAN", "", "NONE"] or "ACUM" in h_str or "CUMPLIMIENTO" in h_str or "EFECTIVO" in h_str or "META" in h_str: 
                        continue
                    nice_name = self.format_month(h_name)
                    month_cols[nice_name] = c_idx

                sheet_rows = []
                count_rows = 0
                
                current_numero, current_indicador, current_formula, current_pond, current_meta = None, None, None, None, None
                temp_indicator_data = {}

                for i in range(h_idx + 1, len(df)):
                    if i in ignored_rows: continue
                    
                    def get_v(col_idx, default_val):
                        if col_idx is None: return default_val
                        val = df.iloc[i, col_idx]
                        return val if pd.notna(val) else default_val

                    raw_num = get_v(col_num, None)
                    str_num = str(raw_num).strip() if raw_num is not None else ""
                    
                    if str_num != "" and str_num.upper() not in ["NAN", "NONE"]:
                        current_numero = str_num
                        current_indicador = get_v(col_ind, "")
                        current_formula = get_v(col_form, "")
                        current_pond = get_v(col_pond, "")
                        current_meta = get_v(col_meta, "")
                        
                        if temp_indicator_data:
                            sheet_rows.append(temp_indicator_data)
                            self.flat_data.append(temp_indicator_data)
                            count_rows += 1
                        
                        clean_formula, type_formula = self.analyze_formula(current_formula)
                        
                        temp_indicator_data = {
                            "ARCHIVO": file_name, 
                            "HOJA": sheet, 
                            "NÚMERO": current_numero,
                            "INDICADOR": current_indicador,
                            "FORMULA": clean_formula,   
                            "TIPO FORMULA": type_formula,
                            "PONDERACIÓN": self.transform_percentage(current_pond),
                            "META": current_meta,
                        }
                    
                    if not current_numero: continue

                    op_text = str(get_v(col_operandos, "")).upper().strip()
                    is_main_val = "VALOR INDICADOR=" in op_text or op_text == ""
                    is_op1 = "OPERANDO 1" in op_text
                    is_op2 = "OPERANDO 2" in op_text
                    
                    if is_op1: temp_indicator_data["Descripción Operando 1"] = str(get_v(col_operandos, ""))
                    if is_op2: temp_indicator_data["Descripción Operando 2"] = str(get_v(col_operandos, ""))

                    for m_name, m_idx in month_cols.items():
                        val_mes = get_v(m_idx, "")
                        if pd.isna(val_mes) or str(val_mes).lower() == "nan": val_mes = ""
                        
                        if is_main_val:
                            temp_indicator_data[f"{m_name} Valor_Ind"] = val_mes
                        elif is_op1:
                            temp_indicator_data[f"{m_name} Op 1"] = val_mes
                            if str(val_mes).strip() != "":
                                self.variable_data.append({
                                    "PERIODO (Mes)": m_name,
                                    "VARIABLE_COD": f"{current_numero}_A",
                                    "VALOR_TOTAL": val_mes,
                                    "ARCHIVO": file_name,
                                    "HOJA": sheet
                                })
                        elif is_op2:
                            temp_indicator_data[f"{m_name} Op 2"] = val_mes
                            if str(val_mes).strip() != "":
                                self.variable_data.append({
                                    "PERIODO (Mes)": m_name,
                                    "VARIABLE_COD": f"{current_numero}_B",
                                    "VALOR_TOTAL": val_mes,
                                    "ARCHIVO": file_name,
                                    "HOJA": sheet
                                })

                if temp_indicator_data:
                    sheet_rows.append(temp_indicator_data)
                    self.flat_data.append(temp_indicator_data)
                    count_rows += 1

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  -> {count_rows} indicadores procesados [Hoja: {sheet}]")

        self.print_summary_and_exit()

    def export_excel(self):
        print(f"\n{'='*60}\nGUARDANDO ARCHIVO MAESTRO ADP...\n{'='*60}")
        wb = Workbook()
        
        # 1. CARGA BRUTA
        ws_bruta = wb.active; ws_bruta.title = "Carga Bruta ADP"
        all_keys = []
        if self.flat_data:
            for d in self.flat_data:
                for k in d.keys():
                    if k not in all_keys: all_keys.append(k)
            ws_bruta.append(all_keys)
            for r in self.flat_data: 
                ws_bruta.append([r.get(k, "") for k in all_keys])
        
        # 2. PLANILLA ESTILIZADA
        ws_style = wb.create_sheet("Planilla Estilizada")
        styles = {
            'file': PatternFill("solid", fgColor="000000"),
            'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"),
            'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True),
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }

        row_idx = 1
        FULL_WIDTH = len(all_keys) if all_keys else 10

        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            
            c = ws_style.cell(row=row_idx, column=1, value=f"ARCHIVO: {fname}")
            c.fill = styles['file']; c.font = styles['w_font']; c.alignment = Alignment('center')
            ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
            row_idx += 1
            
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_style.cell(row=row_idx, column=1, value=f"HOJA: {sname}")
                c.fill = styles['sheet']; c.font = styles['w_font']
                ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
                row_idx += 1
                
                keys = [k for k in all_keys if k not in ["ARCHIVO", "HOJA"]]
                for c_i, k in enumerate(keys, 1):
                    c = ws_style.cell(row=row_idx, column=c_i, value=k)
                    c.fill = styles['head']; c.font = styles['b_font']; c.border = styles['border']
                row_idx += 1
                
                for r in rows:
                    for c_i, k in enumerate(keys, 1):
                        c = ws_style.cell(row=row_idx, column=c_i, value=r.get(k, ""))
                        c.border = styles['border']; c.alignment = Alignment(wrapText=True, vertical='top')
                    row_idx += 1
                row_idx += 1

        for i in range(1, FULL_WIDTH + 2):
            ws_style.column_dimensions[get_column_letter(i)].width = 22

        # 3. DATOS_VARIABLE
        if self.variable_data:
            ws_vars = wb.create_sheet("Transaccional Mensual ADP")
            headers_vars = ["PERIODO (Mes)", "VARIABLE_COD", "VALOR_TOTAL", "ARCHIVO", "HOJA"]
            ws_vars.append(headers_vars)
            
            for row in self.variable_data:
                ws_vars.append([row.get(k, "") for k in headers_vars])

        while True:
            try:
                wb.save(self.output_file)
                print(f"[ÉXITO] Archivo generado: {self.output_file}")
                break
            except PermissionError:
                print(f"\n[ERROR DE PERMISO] El archivo '{self.output_file}' está ABIERTO.")
                input("  >> Por favor CIERRA el Excel y presiona Enter para reintentar...")
            except Exception as e:
                print(f"\n[ERROR AL GUARDAR] {e}")
                input("  >> Presiona Enter para reintentar...")

if __name__ == "__main__":
    try:
        print("INICIANDO PROCESADOR MASIVO ADP v1.0.1")
        path = input("Ruta de la carpeta con las planillas (Enter para actual): ").strip() or os.getcwd()
        if os.path.isdir(path):
            parser = IPS_ADP_Parser(path)
            parser.process_folder()
        else: print("Ruta inválida.")
    except Exception as e:
        print(f"Error fatal: {e}")
        input("Enter para salir.")