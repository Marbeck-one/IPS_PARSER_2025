import pandas as pd
import os
import sys
import glob
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_ADP_PARSER_v1.0.4 - PROCESAMIENTO MASIVO CONVENIOS ADP (Ene-25 a Dic-26)
# =============================================================================

class IPS_ADP_Parser:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "CONSOLIDADO_ADP_v1.0.4.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.variable_data = [] 
        
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible'
        
        self.valid_sheet_keywords = ["PROYECCI", "SIG"]
        
        # Generar matriz fija de 24 meses (Ene-25 hasta Dic-26)
        meses_base = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        self.meses_fijos = [f"{m}-25" for m in meses_base] + [f"{m}-26" for m in meses_base]

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA ADP v1.0.4")
        print("="*60)
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        print(f"\n[OK] Configuración guardada.")
        print("-" * 60)

    def print_summary_and_exit(self):
        print("\n" + "="*60)
        print("   RESUMEN FINAL")
        print("="*60)
        print(f"  * Registros 'BRUTA ADP' (Indicadores): {len(self.flat_data)}")
        print(f"  * Registros 'DATOS VARIABLES ADP':     {len(self.variable_data)}")
        print("-" * 60)
        if self.flat_data: self.export_excel()
        else: print("[AVISO] No se generó archivo de salida.")
        sys.exit()

    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "": return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix, core, f_type = "", f_clean, "CUOCIENTE"
        if match:
            suffix, core, f_type = match.group(1), f_clean[:match.start()].strip(), "PORCENTAJE"
        if core.startswith("(") and core.endswith(")"): core = core[1:-1].strip()
        return core + suffix, f_type

    def transform_percentage(self, val):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or str(val).strip() == "": return ""
        try:
            num = float(val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    def extract_month_name(self, val):
        """ Parsea fechas o textos a formato estándar (Ej: 'Ago-25', 'Ene-26') """
        if pd.isna(val): return ""
        val_str = str(val).strip().lower()
        if not val_str: return ""
        
        try:
            # Intento de parsear como datetime nativo
            dt = pd.to_datetime(val, errors='raise')
            mes_idx = dt.month - 1
            year = str(dt.year)[-2:]
            meses_base = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
            return f"{meses_base[mes_idx]}-{year}"
        except: pass
        
        # Mapeo por texto manual (ej: "ene-25" o "25-ene")
        mapa_meses = {
            "ene": "Ene", "feb": "Feb", "mar": "Mar", "abr": "Abr", 
            "may": "May", "jun": "Jun", "jul": "Jul", "ago": "Ago", 
            "sep": "Sep", "oct": "Oct", "nov": "Nov", "dic": "Dic"
        }
        if "-" in val_str:
            parts = val_str.split("-")
            if len(parts) == 2:
                m, y = parts[0].strip(), parts[1].strip()
                if m in mapa_meses and y.isdigit():
                    return f"{mapa_meses[m]}-{y[-2:]}"
                elif y in mapa_meses and m.isdigit():
                    return f"{mapa_meses[y]}-{m[-2:]}"
        return val_str

    def get_excel_files(self):
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$") and "CONSOLIDADO_ADP" not in f]
        return valid_files

    def process_folder(self):
        files = self.get_excel_files()
        if not files:
            print("[ERROR] No hay archivos válidos.")
            return
            
        self.configure()
        
        for idx_file, file_path in enumerate(files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            print(f"\n>>> PROCESANDO ({idx_file + 1}/{len(files)}): {file_name}")
            
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except: continue

            for sheet in sheet_names:
                s_upper = sheet.upper()
                if "CONSOLIDADO" in s_upper: continue
                if not any(k in s_upper for k in self.valid_sheet_keywords): continue

                df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                
                h_idx = None
                for idx in range(min(15, len(df))):
                    row_vals = [str(x).upper().strip() for x in df.iloc[idx].values if pd.notna(x)]
                    if "NUMERO" in row_vals or "NÚMERO" in row_vals:
                        h_idx = idx
                        break
                
                if h_idx is None: continue 

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                def find_col(keywords):
                    for i, h in enumerate(headers):
                        h_clean = str(h).replace('\n', ' ').upper()
                        if any(k in h_clean for k in keywords): return i
                    return None
                
                def find_cols_containing(keyword):
                    return [i for i, h in enumerate(headers) if keyword in str(h).replace('\n', ' ').upper()]

                col_num = find_col(["NUMERO", "NÚMERO"])     
                col_ind = find_col(["INDICADOR"])            
                col_form = find_col(["FORMULA"])             
                col_pond = find_col(["PONDER"]) # Atrapa PONDERACION, PONDERADOR, PONDERACIÖN
                col_meta = find_col(["META"])                
                col_operandos = find_col(["OPERANDOS"])
                col_meta_est = find_col(["ESTIMADOS META"])
                col_efectivo = find_col(["EFECTIVO"]) # Busca la columna de cumplimiento efectivo (Op1 y Op2)
                col_porc_cump = find_col(["% CUMPLIMIENTO", "CUMPLIMIENTO DE META"])

                # Lógica de anclaje: Detectar Columnas de Meses y a quién le pertenece cada Acumulado
                month_cols = {} 
                month_layout = {m: [] for m in self.meses_fijos} 
                current_m = None
                
                start_month_idx = 7 if len(headers) > 7 else (col_operandos + 2 if col_operandos else 0)
                for c_idx in range(start_month_idx, len(headers)):
                    h_name = headers[c_idx]
                    h_str = str(h_name).replace('\n', ' ').upper().strip()
                    
                    if h_str in ["NAN", "", "NONE"] or "META" in h_str or "EFECTIVO" in h_str or "CUMPLIMIENTO" in h_str: 
                        continue
                        
                    if "ACUM" in h_str:
                        if current_m: month_layout[current_m].append(c_idx)
                    else:
                        generic_month = self.extract_month_name(h_name)
                        if generic_month in self.meses_fijos:
                            month_cols[generic_month] = c_idx
                            current_m = generic_month

                sheet_rows = []
                count_rows = 0
                
                def get_v(r_idx, c_idx, default=""):
                    if c_idx is None or r_idx >= len(df): return default
                    val = df.iloc[r_idx, c_idx]
                    return val if pd.notna(val) else default

                # Procesamiento por Bloques de 6 Filas
                i = h_idx + 1
                while i < len(df):
                    raw_num = get_v(i, col_num)
                    str_num = str(raw_num).strip()
                    
                    if str_num != "" and str_num.upper() not in ["NAN", "NONE"]:
                        clean_formula, type_formula = self.analyze_formula(get_v(i, col_form))
                        
                        temp_data = {
                            "ARCHIVO": file_name, 
                            "HOJA": sheet, 
                            "NÚMERO": str_num,
                            "INDICADOR": get_v(i, col_ind),
                            "FORMULA": clean_formula,   
                            "TIPO FORMULA": type_formula,
                            "PONDERACIÓN": self.transform_percentage(get_v(i, col_pond)),
                            "META": self.transform_percentage(get_v(i, col_meta)), # Meta ahora se multiplica
                            "Descripción Operando 1": get_v(i + 0, col_operandos),
                            "Descripción Operando 2": get_v(i + 3, col_operandos),
                            "Meta Operando 1": get_v(i + 3, col_meta_est),
                            "Meta Operando 2": get_v(i + 5, col_meta_est)
                        }

                        # Matriz Fija de 24 Meses (Entrelazando Op1, Op2 y sus Acumulados)
                        for mes in self.meses_fijos:
                            if mes in month_cols:
                                col_idx = month_cols[mes]
                                val_op1 = get_v(i + 3, col_idx)
                                val_op2 = get_v(i + 5, col_idx)
                                
                                if str(val_op1).strip() != "":
                                    self.variable_data.append({"PERIODO (Mes)": mes, "VARIABLE_COD": f"{str_num}_A", "VALOR_TOTAL": val_op1, "ARCHIVO": file_name, "HOJA": sheet})
                                if str(val_op2).strip() != "":
                                    self.variable_data.append({"PERIODO (Mes)": mes, "VARIABLE_COD": f"{str_num}_B", "VALOR_TOTAL": val_op2, "ARCHIVO": file_name, "HOJA": sheet})
                            else:
                                val_op1, val_op2 = "", ""
                                
                            temp_data[f"{mes} Op 1"] = val_op1
                            temp_data[f"{mes} Op 2"] = val_op2
                            
                            # Extraer su acumulado (si lo tiene) e intercalarlo inmediatamente después
                            acums = month_layout.get(mes, [])
                            if acums:
                                val_acum_op1 = get_v(i + 3, acums[0])
                                val_acum_op2 = get_v(i + 5, acums[0])
                            else:
                                val_acum_op1, val_acum_op2 = "", ""
                                
                            temp_data[f"{mes} Acum Op 1"] = val_acum_op1
                            temp_data[f"{mes} Acum Op 2"] = val_acum_op2

                        # Efectivo a la Fecha al final
                        temp_data["Efectivo Op 1"] = get_v(i + 3, col_efectivo)
                        temp_data["Efectivo Op 2"] = get_v(i + 5, col_efectivo)

                        cump_val = ""
                        for offset in range(6):
                            val = get_v(i + offset, col_porc_cump)
                            if str(val).strip() != "":
                                cump_val = val
                                break
                        temp_data["% Cumplimiento de Meta"] = self.transform_percentage(cump_val)

                        sheet_rows.append(temp_data)
                        self.flat_data.append(temp_data)
                        count_rows += 1
                        
                        i += 6 
                    else:
                        i += 1 

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  -> {count_rows} indicadores procesados [Hoja: {sheet}]")

        self.print_summary_and_exit()

    def export_excel(self):
        print(f"\n{'='*60}\nGUARDANDO ARCHIVO: {self.output_file}\n{'='*60}")
        wb = Workbook()
        
        # 1. BRUTA ADP
        ws_bruta = wb.active; ws_bruta.title = "BRUTA ADP"
        all_keys = []
        for d in self.flat_data:
            for k in d.keys():
                if k not in all_keys: all_keys.append(k)
        ws_bruta.append(all_keys)
        for r in self.flat_data: ws_bruta.append([r.get(k, "") for k in all_keys])
        
        # 2. ESTILIZADA ADP
        ws_style = wb.create_sheet("ESTILIZADA ADP")
        styles = {
            'file': PatternFill("solid", fgColor="000000"),
            'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"),
            'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True),
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }

        row_idx = 1
        FULL_WIDTH = len(all_keys) if all_keys else 10

        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            c = ws_style.cell(row=row_idx, column=1, value=f"ARCHIVO: {fname}")
            c.fill = styles['file']; c.font = styles['w_font']; c.alignment = Alignment('center')
            ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
            row_idx += 1
            
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_style.cell(row=row_idx, column=1, value=f"HOJA: {sname}")
                c.fill = styles['sheet']; c.font = styles['w_font']
                ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
                row_idx += 1
                
                keys = [k for k in all_keys if k not in ["ARCHIVO", "HOJA"]]
                for c_i, k in enumerate(keys, 1):
                    c = ws_style.cell(row=row_idx, column=c_i, value=k)
                    c.fill = styles['head']; c.font = styles['b_font']; c.border = styles['border']
                row_idx += 1
                
                for r in rows:
                    for c_i, k in enumerate(keys, 1):
                        c = ws_style.cell(row=row_idx, column=c_i, value=r.get(k, ""))
                        c.border = styles['border']; c.alignment = Alignment(wrapText=True, vertical='top')
                    row_idx += 1
                row_idx += 1

        for i in range(1, FULL_WIDTH + 2): ws_style.column_dimensions[get_column_letter(i)].width = 22

        # 3. DATOS VARIABLES ADP
        if self.variable_data:
            ws_vars = wb.create_sheet("DATOS VARIABLES ADP")
            headers_vars = ["PERIODO (Mes)", "VARIABLE_COD", "VALOR_TOTAL", "ARCHIVO", "HOJA"]
            ws_vars.append(headers_vars)
            for row in self.variable_data: ws_vars.append([row.get(k, "") for k in headers_vars])

        while True:
            try:
                wb.save(self.output_file)
                print(f"[ÉXITO] Archivo generado: {self.output_file}")
                break
            except Exception as e:
                print(f"\n[ERROR AL GUARDAR] {e}")
                input("  >> Cierra el archivo si está abierto y presiona Enter para reintentar...")

if __name__ == "__main__":
    try:
        print("INICIANDO PROCESADOR MASIVO ADP v1.0.4")
        path = input("Ruta de la carpeta (Enter para actual): ").strip() or os.getcwd()
        if os.path.isdir(path):
            parser = IPS_ADP_Parser(path)
            parser.process_folder()
        else: print("Ruta inválida.")
    except Exception as e:
        print(f"Error fatal: {e}")
        input("Enter para salir.")