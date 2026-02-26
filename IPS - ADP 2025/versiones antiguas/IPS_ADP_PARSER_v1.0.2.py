import pandas as pd
import os
import sys
import glob
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_ADP_PARSER_v1.0.2 - PROCESAMIENTO MASIVO CONVENIOS ADP (Bloques 6 Filas)
# =============================================================================

class IPS_ADP_Parser:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "CONSOLIDADO_ADP_v1.0.2.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.variable_data = [] 
        
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible'
        
        self.valid_sheet_keywords = ["PROYECCI", "SIG"]
        
        # Nombres genéricos fijos de meses para la salida
        self.meses_fijos = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA ADP v1.0.2")
        print("="*60)
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        print(f"\n[OK] Configuración guardada.")
        print("-" * 60)

    def print_summary_and_exit(self):
        print("\n" + "="*60)
        print("   RESUMEN FINAL")
        print("="*60)
        print(f"  * Registros 'Carga Bruta' (Indicadores): {len(self.flat_data)}")
        print(f"  * Registros 'Transaccional' (Meses):     {len(self.variable_data)}")
        print("-" * 60)
        if self.flat_data: self.export_excel()
        else: print("[AVISO] No se generó archivo de salida.")
        sys.exit()

    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "": return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix, core, f_type = "", f_clean, "CUOCIENTE"
        if match:
            suffix, core, f_type = match.group(1), f_clean[:match.start()].strip(), "PORCENTAJE"
        if core.startswith("(") and core.endswith(")"): core = core[1:-1].strip()
        return core + suffix, f_type

    def transform_percentage(self, val):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or str(val).strip() == "": return ""
        try:
            num = float(val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    def extract_month_name(self, val):
        """ Retorna solo el nombre del mes genérico (Ej: 'Agosto') detectando YYYY-MM o strings """
        if pd.isna(val): return ""
        val_str = str(val).strip().lower()
        try:
            dt = pd.to_datetime(val)
            return self.meses_fijos[dt.month - 1]
        except: pass
        
        mapa_meses = {
            "ene": "Enero", "feb": "Febrero", "mar": "Marzo", "abr": "Abril", 
            "may": "Mayo", "jun": "Junio", "jul": "Julio", "ago": "Agosto", 
            "sep": "Septiembre", "oct": "Octubre", "nov": "Noviembre", "dic": "Diciembre"
        }
        for k, v in mapa_meses.items():
            if k in val_str: return v
        return val_str

    def get_excel_files(self):
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$") and "CONSOLIDADO_ADP" not in f]
        return valid_files

    def process_folder(self):
        files = self.get_excel_files()
        if not files:
            print("[ERROR] No hay archivos válidos.")
            return
            
        self.configure()
        
        for idx_file, file_path in enumerate(files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            print(f"\n>>> PROCESANDO ({idx_file + 1}/{len(files)}): {file_name}")
            
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except: continue

            for sheet in sheet_names:
                s_upper = sheet.upper()
                if "CONSOLIDADO" in s_upper: continue
                if not any(k in s_upper for k in self.valid_sheet_keywords): continue

                df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                
                h_idx = None
                for idx in range(min(15, len(df))):
                    row_vals = [str(x).upper().strip() for x in df.iloc[idx].values if pd.notna(x)]
                    if "NUMERO" in row_vals or "NÚMERO" in row_vals:
                        h_idx = idx
                        break
                
                if h_idx is None: continue 

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                def find_col(keywords):
                    for i, h in enumerate(headers):
                        if any(k in str(h).upper() for k in keywords): return i
                    return None
                
                def find_cols_containing(keyword):
                    return [i for i, h in enumerate(headers) if keyword in str(h).upper()]

                col_num = find_col(["NUMERO", "NÚMERO"])     
                col_ind = find_col(["INDICADOR"])            
                col_form = find_col(["FORMULA"])             
                col_pond = find_col(["PONDERACIÓN", "PONDERACION"]) 
                col_meta = find_col(["META"])                
                col_operandos = find_col(["OPERANDOS"])
                col_meta_est = find_col(["ESTIMADOS META"])
                col_efectivo = find_col(["EFECTIVO A LA FECHA"])
                col_porc_cump = find_col(["% CUMPLIMIENTO"])

                # Detectar Columnas de Meses y Acumulados
                month_cols = {} # { "Agosto": idx_columna }
                acum_cols = find_cols_containing("ACUM")
                
                start_month_idx = 7 if len(headers) > 7 else (col_operandos + 2 if col_operandos else 0)
                for c_idx in range(start_month_idx, len(headers)):
                    h_name = headers[c_idx]
                    h_str = str(h_name).upper().strip()
                    if h_str in ["NAN", "", "NONE"] or "ACUM" in h_str or "CUMPLIMIENTO" in h_str or "EFECTIVO" in h_str or "META" in h_str: 
                        continue
                    generic_month = self.extract_month_name(h_name)
                    if generic_month in self.meses_fijos:
                        month_cols[generic_month] = c_idx

                sheet_rows = []
                count_rows = 0
                
                def get_v(r_idx, c_idx, default=""):
                    if c_idx is None or r_idx >= len(df): return default
                    val = df.iloc[r_idx, c_idx]
                    return val if pd.notna(val) else default

                # Procesamiento por Bloques de 6 Filas
                i = h_idx + 1
                while i < len(df):
                    raw_num = get_v(i, col_num)
                    str_num = str(raw_num).strip()
                    
                    if str_num != "" and str_num.upper() not in ["NAN", "NONE"]:
                        clean_formula, type_formula = self.analyze_formula(get_v(i, col_form))
                        
                        # Datos Base
                        temp_data = {
                            "ARCHIVO": file_name, 
                            "HOJA": sheet, 
                            "NÚMERO": str_num,
                            "INDICADOR": get_v(i, col_ind),
                            "FORMULA": clean_formula,   
                            "TIPO FORMULA": type_formula,
                            "PONDERACIÓN": self.transform_percentage(get_v(i, col_pond)),
                            "META": get_v(i, col_meta),
                            "Descripción Operando 1": get_v(i + 0, col_operandos),
                            "Descripción Operando 2": get_v(i + 3, col_operandos),
                            "Meta Operando 1": get_v(i + 3, col_meta_est),
                            "Meta Operando 2": get_v(i + 5, col_meta_est)
                        }

                        # Matriz Fija de 12 Meses
                        for mes in self.meses_fijos:
                            if mes in month_cols:
                                col_idx = month_cols[mes]
                                val_op1 = get_v(i + 3, col_idx)
                                val_op2 = get_v(i + 5, col_idx)
                                temp_data[f"{mes} Op 1"] = val_op1
                                temp_data[f"{mes} Op 2"] = val_op2
                                
                                # Generar Transaccional si hay valor Op1
                                if str(val_op1).strip() != "":
                                    self.variable_data.append({"PERIODO (Mes)": mes, "VARIABLE_COD": f"{str_num}_A", "VALOR_TOTAL": val_op1, "ARCHIVO": file_name, "HOJA": sheet})
                                # Generar Transaccional si hay valor Op2
                                if str(val_op2).strip() != "":
                                    self.variable_data.append({"PERIODO (Mes)": mes, "VARIABLE_COD": f"{str_num}_B", "VALOR_TOTAL": val_op2, "ARCHIVO": file_name, "HOJA": sheet})
                            else:
                                temp_data[f"{mes} Op 1"] = ""
                                temp_data[f"{mes} Op 2"] = ""

                        # Acumulados Dinámicos
                        for idx_acum, col_idx in enumerate(acum_cols, 1):
                            temp_data[f"Acumulado {idx_acum} Op 1"] = get_v(i + 3, col_idx)
                            temp_data[f"Acumulado {idx_acum} Op 2"] = get_v(i + 5, col_idx)

                        # Efectivo a la Fecha
                        temp_data["Efectivo Op 1"] = get_v(i + 3, col_efectivo)
                        temp_data["Efectivo Op 2"] = get_v(i + 5, col_efectivo)

                        # % Cumplimiento (Buscar en el rango de 6 filas)
                        cump_val = ""
                        for offset in range(6):
                            val = get_v(i + offset, col_porc_cump)
                            if str(val).strip() != "":
                                cump_val = val
                                break
                        temp_data["% Cumplimiento de Meta"] = self.transform_percentage(cump_val)

                        sheet_rows.append(temp_data)
                        self.flat_data.append(temp_data)
                        count_rows += 1
                        
                        i += 6 # Saltar directamente al siguiente bloque
                    else:
                        i += 1 # Si está vacío, avanzar de 1 en 1 buscando el próximo indicador

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  -> {count_rows} indicadores procesados [Hoja: {sheet}]")

        self.print_summary_and_exit()

    def export_excel(self):
        print(f"\n{'='*60}\nGUARDANDO ARCHIVO: {self.output_file}\n{'='*60}")
        wb = Workbook()
        
        # 1. PLANTILLA SIG ADP (Carga Bruta)
        ws_bruta = wb.active; ws_bruta.title = "PLANTILLA SIG ADP"
        all_keys = []
        for d in self.flat_data:
            for k in d.keys():
                if k not in all_keys: all_keys.append(k)
        ws_bruta.append(all_keys)
        for r in self.flat_data: ws_bruta.append([r.get(k, "") for k in all_keys])
        
        # 2. PROYECCIONES ADP (Estilizada)
        ws_style = wb.create_sheet("PROYECCIONES ADP")
        styles = {
            'file': PatternFill("solid", fgColor="000000"),
            'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"),
            'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True),
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }

        row_idx = 1
        FULL_WIDTH = len(all_keys) if all_keys else 10

        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            c = ws_style.cell(row=row_idx, column=1, value=f"ARCHIVO: {fname}")
            c.fill = styles['file']; c.font = styles['w_font']; c.alignment = Alignment('center')
            ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
            row_idx += 1
            
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_style.cell(row=row_idx, column=1, value=f"HOJA: {sname}")
                c.fill = styles['sheet']; c.font = styles['w_font']
                ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
                row_idx += 1
                
                keys = [k for k in all_keys if k not in ["ARCHIVO", "HOJA"]]
                for c_i, k in enumerate(keys, 1):
                    c = ws_style.cell(row=row_idx, column=c_i, value=k)
                    c.fill = styles['head']; c.font = styles['b_font']; c.border = styles['border']
                row_idx += 1
                
                for r in rows:
                    for c_i, k in enumerate(keys, 1):
                        c = ws_style.cell(row=row_idx, column=c_i, value=r.get(k, ""))
                        c.border = styles['border']; c.alignment = Alignment(wrapText=True, vertical='top')
                    row_idx += 1
                row_idx += 1

        for i in range(1, FULL_WIDTH + 2): ws_style.column_dimensions[get_column_letter(i)].width = 22

        # 3. TRANSACCIONAL MENSUAL ADP
        if self.variable_data:
            ws_vars = wb.create_sheet("Transaccional Mensual ADP")
            headers_vars = ["PERIODO (Mes)", "VARIABLE_COD", "VALOR_TOTAL", "ARCHIVO", "HOJA"]
            ws_vars.append(headers_vars)
            for row in self.variable_data: ws_vars.append([row.get(k, "") for k in headers_vars])

        while True:
            try:
                wb.save(self.output_file)
                print(f"[ÉXITO] Archivo generado: {self.output_file}")
                break
            except Exception as e:
                print(f"\n[ERROR AL GUARDAR] {e}")
                input("  >> Cierra el archivo si está abierto y presiona Enter para reintentar...")

if __name__ == "__main__":
    try:
        print("INICIANDO PROCESADOR MASIVO ADP v1.0.2")
        path = input("Ruta de la carpeta (Enter para actual): ").strip() or os.getcwd()
        if os.path.isdir(path):
            parser = IPS_ADP_Parser(path)
            parser.process_folder()
        else: print("Ruta inválida.")
    except Exception as e:
        print(f"Error fatal: {e}")
        input("Enter para salir.")