import pandas as pd
import os
import re
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# =============================================================================
# BLOQUE 0: CONFIGURACIÓN Y MAPAS INTERNOS
# =============================================================================

ARCHIVOS = {
    "MAESTRO": "Proyecciones Indicadores 2025 - División Planificación (1).xlsx",
    "ESTILIZADA": "Planilla_Estilizada_2025.xlsx",
    "BRUTA": "Planilla_Bruta_2025.xlsx",
    "VARS_IPS": "VARIABLES_IPS_2026.xlsx",
    "VARS_APP": "VARIABLES_APLICADAS_IPS_2026.xlsx",
    "INDS_IPS": "INDICADORES_IPS_2026.xlsx",
    "INDS_APP": "INDICADORES_APLICADOS_IPS_2026.xlsx"
}

HOJAS_CONFIG = {"CDC 2025": "CDC", "PMG 2025": "PMG", "Riesgos 2025": "Riesgos"}

# --- MAPA INTERNO DE PONDERADOS (REEMPLAZA AL CSV) ---
# Claves normalizadas (Sin CDC, sin tildes, mayúsculas, sin espacios)
MAPA_PONDERADOS_INTERNO = {
    'FORMULARIOH': 'IP25_711',
    'DIVISIONBENEFICIOS': 'IP25_712',
    'SUBDIRECCIONSERVICIOSALCLIENTE': 'IP25_713',
    'DIVISIONINFORMATICA': 'IP25_714',
    'DIVISIONJURIDICA': 'IP25_715',
    'DIVISIONPLANIFICACIONYDESARROLLO': 'IP25_716',
    'DEPARTAMENTODECOMUNICACIONES': 'IP25_717',
    'CONTRALORIAINTERNA': 'IP25_718',
    'REGIONARICAYPARINACOTA': 'IP25_719',
    'REGIONTARAPACA': 'IP25_720',
    'REGIONANTOFAGASTA': 'IP25_721',
    'REGIONATACAMA': 'IP25_722',
    'REGIONCOQUIMBO': 'IP25_723',
    'REGIONDEVALPARAISO': 'IP25_724',
    'REGIONDELIBERTADORBERNARDOOHIGGINS': 'IP25_725',
    'REGIONDELMAULE': 'IP25_726',
    'REGIONDELBIOBIO': 'IP25_727',
    'REGIONDELAARAUCANIA': 'IP25_728',
    'REGIONDELOSRIOS': 'IP25_729',
    'REGIONDELOSLAGOS': 'IP25_730',
    'REGIONDEAISEN': 'IP25_731',
    'REGIONDEMAGALLANESYLAANTARTICACHILENAR': 'IP25_732',
    'REGIONMETROPOLITANA': 'IP25_733',
    'AUDITORIAINTERNA': 'IP25_738',
    'SUBDIRECCIONDESISTEMASDEINFORMACIONYADMINISTRACION': 'IP25_739',
    'PMGSMDIOBJETIVO1GESTIONEFICAZ': 'IP25_740',
    'PMGSMDIOBJETIVO2EFICIENCIAINSTITUCIONAL': 'IP25_741',
    'PMGSMDIOBJETIVO3CALIDADDELOSSERVICIOS': 'IP25_742',
    'REGIONDENUBLE': 'IP25_748',
    'DEPARTAMENTODEGESTIONYDESARROLLODEPERSONAS': 'IP25_750',
    'GESTIONINTERNACALIDADDESERVICIOYEXPERIENCIAUSUARIA': 'IP25_752'
}

# --- FUNCIONES DE LIMPIEZA GENERAL ---
def limpiar_porcentaje_real(val):
    if pd.isna(val) or val == "": return 0
    if isinstance(val, str):
        limpio = val.replace('%', '').replace(',', '.').strip()
        try: return float(limpio)
        except: return 0
    if isinstance(val, (int, float)): return val * 100
    return 0

def limpiar_op1_inicio(val):
    if pd.isna(val) or val == "": return ""
    txt = str(val).strip()
    if txt.startswith("("): return txt[1:].strip()
    return txt

def limpiar_op2_final(val):
    if pd.isna(val) or val == "": return ""
    txt = str(val).strip()
    return re.sub(r'\)\s*\*100$', '', txt).strip()

def detectar_encabezados(df):
    for i in range(25):
        try:
            fila = df.iloc[i].astype(str).tolist()
            if "NÚMERO" in fila and any("INDICADOR" in s for s in fila):
                return i, fila
        except: continue
    raise ValueError("ERROR: No se encontró la fila de encabezados.")

def transformar_codigo_para_var_auto(cod_variable):
    cod_str = str(cod_variable).strip()
    if cod_str.startswith("---"): return None
    if "INDICADOR_NUEVO" in cod_str:
        parts = cod_str.split('_')
        if len(parts) >= 5: return f"{parts[-2]}_{'_'.join(parts[:-2])}_{parts[-1]}"
        elif len(parts) == 4: return f"{parts[-1]}_{'_'.join(parts[:-1])}"
    elif '_' in cod_str:
        partes = cod_str.rsplit('_', 1) 
        if len(partes) == 2: return f"{partes[1]}_{partes[0]}"
    return cod_str

# --- FUNCIONES FASE 4 (INDICADORES) ---
def parsear_nombre_indicador(texto_bruto):
    if pd.isna(texto_bruto): return "EFICACIA", "PROCESO", ""
    texto = str(texto_bruto).strip()
    patron = r"^(?:[\d\)\(\s]+)?([a-zA-ZáéíóúñÁÉÍÓÚÑ]+)/([a-zA-ZáéíóúñÁÉÍÓÚÑ]+)\s+(.*)"
    match = re.search(patron, texto, re.DOTALL) 
    if match:
        return match.group(1).upper(), match.group(2).upper(), match.group(3).replace('\n', ' ').strip()
    return "EFICACIA", "PROCESO", texto.replace('\n', ' ').strip()

def determinar_unidad(nombre_limpio):
    nombre_lower = str(nombre_limpio).lower()
    if "porcentaje" in nombre_lower or "%" in nombre_lower: return "%"
    if any(x in nombre_lower for x in ["tiempo", "medidas", "numero", "número", "cantidad", "tasa"]): return "n"
    return "?"

def limpiar_codigo_indicador_fase5(raw_num, contador, etiqueta):
    cod_str = str(raw_num).strip()
    es_invalido = (not raw_num) or (pd.isna(raw_num)) or (cod_str == "") or (cod_str.lower() == "nan")
    es_nuevo = "NUEVO" in cod_str.upper() or "INDICADOR" in cod_str.upper()
    if es_invalido or es_nuevo: return f"INDICADOR_NUEVO_{contador}_{etiqueta}", True
    else: return cod_str, False

# --- NORMALIZACIÓN PARA CRUCE (FASE 5) ---
def normalizar_clave_responsable(texto):
    if pd.isna(texto): return ""
    txt = str(texto).upper().strip()
    txt = txt.replace("CDC", "") # Quitar prefijo
    txt = ''.join(c for c in unicodedata.normalize('NFD', txt) if unicodedata.category(c) != 'Mn') # Quitar tildes
    txt = re.sub(r'[^\w\s]', '', txt) # Quitar puntuación
    txt = re.sub(r'\s+', '', txt) # Quitar espacios (SUPER NORMALIZACIÓN)
    return txt

# =============================================================================
# BLOQUE 1: FASE 1 (EXTRACCIÓN)
# =============================================================================

def obtener_dataframe_hoja(ruta_archivo, nombre_hoja_excel, etiqueta_log):
    print(f"   -> Extrayendo datos de: {etiqueta_log}...")
    try:
        df_raw = pd.read_excel(ruta_archivo, sheet_name=nombre_hoja_excel, header=None)
    except Exception as e:
        print(f"      [ERROR] No se pudo leer la hoja '{nombre_hoja_excel}': {e}")
        return None

    try:
        idx_header, encabezados = detectar_encabezados(df_raw)
    except ValueError as e:
        print(f"      [ERROR] {e}")
        return None

    mapa_cols = {str(nombre).strip().replace("\n", " "): i for i, nombre in enumerate(encabezados)}
    def buscar_col(claves):
        if isinstance(claves, str): claves = [claves]
        for c in claves:
            for nombre_real, idx in mapa_cols.items():
                if c.lower() in nombre_real.lower(): return idx
        return None

    idx_num = buscar_col("NÚMERO")
    idx_meta = buscar_col("Meta 2025")
    idx_ponderador = buscar_col("Ponderador")
    idx_op_desc = buscar_col("Operandos")
    idx_op_est = buscar_col(["Operandos Estimados", "Operandos  Estimados"])
    idx_medios = buscar_col("Medios de Verificación")
    idx_control = buscar_col("Control de Cambios")
    idx_inst = buscar_col(["Instrumentos de Gestión", "Instrumentos Gestión"])
    idx_cump_meta = buscar_col("% Cumplimiento de Meta")

    indices_inicio = []
    for i in range(idx_header + 1, len(df_raw)):
        val = df_raw.iloc[i, idx_num]
        if pd.notna(val) and str(val).strip() != "" and str(val) != "NÚMERO":
            indices_inicio.append(i)
            
    lista_datos = []
    meses_std = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sept", "Oct", "Nov", "Dic"]
    mapa_meses_indices = {}
    for mes in meses_std:
        idx = buscar_col(f"{mes}.") 
        if idx: mapa_meses_indices[mes] = idx
    idx_proy = buscar_col(["Cumplimiento Proyectado", "Cumplimiento\nProyectado"])
    if idx_proy: mapa_meses_indices["Cump. Proy."] = idx_proy

    for idx in indices_inicio:
        fila = {}
        fila["NÚMERO"] = df_raw.iloc[idx, idx_num]
        fila["INDICADOR"] = df_raw.iloc[idx, buscar_col("INDICADOR")] if buscar_col("INDICADOR") else ""
        
        col_prod = buscar_col("PRODUCTO"); fila["PRODUCTO O PROCESO ESPECÍFICO"] = df_raw.iloc[idx, col_prod] if col_prod else ""
        col_form = buscar_col("FORMULA"); fila["FORMULA"] = df_raw.iloc[idx, col_form] if col_form else ""
        col_unidad = buscar_col("UNIDAD"); fila["UNIDAD"] = df_raw.iloc[idx, col_unidad] if col_unidad else ""
        col_resp = buscar_col("RESPONSABLE CENTRO"); fila["RESPONSABLE CENTRO DE RESPONSABILIDAD"] = df_raw.iloc[idx, col_resp] if col_resp else ""
        col_gestor = buscar_col("GESTOR"); fila["GESTOR"] = df_raw.iloc[idx, col_gestor] if col_gestor else ""
        col_sup = buscar_col("SUPERVISORES"); fila["SUPERVISORES"] = df_raw.iloc[idx, col_sup] if col_sup else ""
        
        if idx_meta: fila["Meta 2025 (%)"] = df_raw.iloc[idx, idx_meta]
        if idx_ponderador: fila["Ponderador (%)"] = df_raw.iloc[idx, idx_ponderador]
        else: fila["Ponderador (%)"] = 0 
        
        if idx_op_desc:
            fila["Desc. Op1"] = df_raw.iloc[idx, idx_op_desc]
            fila["Desc. Op2"] = df_raw.iloc[idx+3, idx_op_desc]
            
        for nombre_mes, col_idx in mapa_meses_indices.items():
            fila[f"{nombre_mes} Ind (%)"] = df_raw.iloc[idx+1, col_idx]
            
        if idx_medios: fila["Medios Verificación"] = df_raw.iloc[idx, idx_medios]
        
        lista_datos.append(fila)

    df_final = pd.DataFrame(lista_datos).fillna(0)
    for col in df_final.columns:
        if "(%)" in col:
            df_final[col] = df_final[col].apply(limpiar_porcentaje_real)
            
    return df_final

def ejecutar_fase_1(opcion_formato="2"):
    print("\n--- EJECUTANDO FASE 1: EXTRACCIÓN ---")
    if not os.path.exists(ARCHIVOS["MAESTRO"]):
        print(f"[ERROR] Falta {ARCHIVOS['MAESTRO']}")
        return None

    dataframes_listos = {}
    for hoja_excel, etiqueta in HOJAS_CONFIG.items():
        df = obtener_dataframe_hoja(ARCHIVOS["MAESTRO"], hoja_excel, etiqueta)
        if df is not None: dataframes_listos[etiqueta] = df

    if not dataframes_listos: return None

    archivos_gen = []
    if opcion_formato in ["1", "3"]: archivos_gen.append((ARCHIVOS["BRUTA"], False))
    if opcion_formato in ["2", "3"]: archivos_gen.append((ARCHIVOS["ESTILIZADA"], True))

    for nombre, estilo in archivos_gen:
        print(f"   -> Guardando: {nombre}")
        with pd.ExcelWriter(nombre, engine='openpyxl') as writer:
            for etiq, df in dataframes_listos.items(): df.to_excel(writer, sheet_name=etiq, index=False)
        if estilo: aplicar_estilos_planillas(nombre)
    
    return dataframes_listos

# =============================================================================
# BLOQUE 2: FASE 2 (VARIABLES IPS)
# =============================================================================

def transformar_a_variables(df_origen, etiqueta_origen, contador_nuevos_global):
    filas_variables = []
    for index, row in df_origen.iterrows():
        raw_num = row.get('NÚMERO', ''); codigo_str = str(raw_num).strip()
        es_invalido = (not raw_num) or (pd.isna(raw_num)) or (codigo_str == "") or (codigo_str.lower() == "nan")
        es_nuevo = "NUEVO" in codigo_str.upper() or "INDICADOR" in codigo_str.upper()
        
        if es_invalido or es_nuevo:
            cod_A = f"INDICADOR_NUEVO_{contador_nuevos_global}_A_{etiqueta_origen}"
            cod_B = f"INDICADOR_NUEVO_{contador_nuevos_global}_B_{etiqueta_origen}"
            contador_nuevos_global += 1
        else:
            cod_A = f"{codigo_str}_A"; cod_B = f"{codigo_str}_B"

        raw_op1 = row.get('Desc. Op1', ''); nombre_A = limpiar_op1_inicio(raw_op1)
        raw_op2 = row.get('Desc. Op2', ''); nombre_B = limpiar_op2_final(raw_op2)
        raw_medios = row.get('Medios Verificación', ''); medios_verif = str(raw_medios).strip() if pd.notna(raw_medios) else ""

        base_fila = {
            'APLICA_DIST_GENERO': 0, 'APLICA_DESP_TERRITORIAL': 0, 'APLICA_SIN_INFORMACION': 1,
            'APLICA_VAL_PERS_JUR': 0, 'requiere_medio': 0, 'texto_ayuda': None, 'unidad': None,
            'valor_obligatorio': 1, 'permite_medio_escrito': 1, 'usa_ultimo_valor_ano': 1
        }
        fila_A = base_fila.copy(); fila_A.update({'cod_interno': cod_A, 'nombre_variable': nombre_A, 'descripcion': nombre_A, 'medio_verificacion': medios_verif})
        fila_B = base_fila.copy(); fila_B.update({'cod_interno': cod_B, 'nombre_variable': nombre_B, 'descripcion': nombre_B, 'medio_verificacion': medios_verif})
        filas_variables.extend([fila_A, fila_B])
        
    return pd.DataFrame(filas_variables), contador_nuevos_global

def ejecutar_fase_2(dataframes_input=None):
    print("\n--- EJECUTANDO FASE 2: VARIABLES IPS ---")
    if dataframes_input is None:
        dataframes_input = cargar_datos_intermedios_fase1()
        if not dataframes_input: return None

    dfs_consolidados = []; contador_global = 1
    cols_orden = ['cod_interno', 'nombre_variable', 'descripcion', 'medio_verificacion', 'APLICA_DIST_GENERO', 'APLICA_DESP_TERRITORIAL', 'APLICA_SIN_INFORMACION', 'APLICA_VAL_PERS_JUR', 'requiere_medio', 'texto_ayuda', 'unidad', 'valor_obligatorio', 'permite_medio_escrito', 'usa_ultimo_valor_ano']

    for etiqueta in ["CDC", "Riesgos", "PMG"]:
        if etiqueta in dataframes_input:
            df_vars, nuevo_cont = transformar_a_variables(dataframes_input[etiqueta], etiqueta, contador_global)
            contador_global = nuevo_cont
            if not df_vars.empty:
                fila_titulo = {col: None for col in cols_orden}; fila_titulo['cod_interno'] = f"--- {etiqueta.upper()} VARIABLES ---"
                dfs_consolidados.append(pd.DataFrame([fila_titulo])); dfs_consolidados.append(df_vars)

    if dfs_consolidados:
        df_final = pd.concat(dfs_consolidados, ignore_index=True).reindex(columns=cols_orden)
        with pd.ExcelWriter(ARCHIVOS["VARS_IPS"], engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name="VARIABLES_BRUTA", index=False)
            df_final.to_excel(writer, sheet_name="VARIABLES_ESTILIZADA", index=False)
        aplicar_estilos_variables(ARCHIVOS["VARS_IPS"], "VARIABLES_ESTILIZADA")
        print(f"   -> [OK] Archivo generado: {ARCHIVOS['VARS_IPS']}")
        return df_final
    return None

# =============================================================================
# BLOQUE 3: FASE 3 (VARIABLES APLICADAS)
# =============================================================================

def transformar_a_variables_aplicadas(df_variables_fase2):
    filas_aplicadas = []
    for index, row in df_variables_fase2.iterrows():
        cod_variable = str(row.get('cod_interno', '')).strip()
        if cod_variable.startswith("---"):
            filas_aplicadas.append({'cod_variable': cod_variable}); continue
        if not cod_variable or cod_variable.lower() == "nan": continue

        fila = {
            'cod_variable': cod_variable,
            'nombre_variable': row.get('nombre_variable', ''),
            'ano_mes_ini': 202501, 'ano_mes_fin': 202512,
            'ENE': 1, 'FEB': 1, 'MAR': 1, 'ABR': 1, 'MAY': 1, 'JUN': 1,
            'JUL': 1, 'AGO': 1, 'SEP': 1, 'OCT': 1, 'NOV': 1, 'DIC': 1,
            'cod_centro_resp_lugar_medicion': None, 'cod_region': None,
            'EMAIL_RESPONSABLE_INGRESO_DATO': 'prueba@arbol-logika.com',
            'EMAIL_PRIMER_REVISOR': None, 'EMAIL_SEGUNDO_REVISOR': None,
            'PERMITE_ADJUNTAR_MEDIO': 1, 'MOSTRAR_TABLA_ANOS': 1,
            'FORMULA_VAR_AUTO': 'SUMA_ANUAL',
            'codigo_var_auto': transformar_codigo_para_var_auto(cod_variable)
        }
        filas_aplicadas.append(fila)
    return pd.DataFrame(filas_aplicadas)

def ejecutar_fase_3(df_vars_input=None):
    print("\n--- EJECUTANDO FASE 3: VARIABLES APLICADAS ---")
    if df_vars_input is None:
        if os.path.exists(ARCHIVOS["VARS_IPS"]):
            try: df_vars_input = pd.read_excel(ARCHIVOS["VARS_IPS"], sheet_name="VARIABLES_BRUTA")
            except: return
        else: return

    df_aplicadas = transformar_a_variables_aplicadas(df_vars_input)
    cols_orden = ['cod_variable', 'nombre_variable', 'ano_mes_ini', 'ano_mes_fin', 'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC', 'cod_centro_resp_lugar_medicion', 'cod_region', 'EMAIL_RESPONSABLE_INGRESO_DATO', 'EMAIL_PRIMER_REVISOR', 'EMAIL_SEGUNDO_REVISOR', 'PERMITE_ADJUNTAR_MEDIO', 'MOSTRAR_TABLA_ANOS', 'FORMULA_VAR_AUTO', 'codigo_var_auto']
    df_aplicadas = df_aplicadas.reindex(columns=cols_orden)
    
    df_aplicadas.to_excel(ARCHIVOS["VARS_APP"], index=False)
    aplicar_estilos_aplicadas(ARCHIVOS["VARS_APP"])
    print(f"   -> [OK] Archivo generado: {ARCHIVOS['VARS_APP']}")

# =============================================================================
# BLOQUE 4: FASE 4 (INDICADORES IPS)
# =============================================================================

def transformar_a_indicadores_ips(df_origen, etiqueta_origen, contador_nuevos_global):
    filas_indicadores = []
    for index, row in df_origen.iterrows():
        raw_num = row.get('NÚMERO', '')
        codigo_indicador, _ = limpiar_codigo_indicador_fase5(raw_num, contador_nuevos_global, etiqueta_origen)
        if "INDICADOR_NUEVO" in codigo_indicador: contador_nuevos_global += 1

        raw_indicador = row.get('INDICADOR', '')
        ambito, dimension, nombre_limpio = parsear_nombre_indicador(raw_indicador)
        unidad = determinar_unidad(nombre_limpio)
        
        flags = {'IND_BGI': 0, 'IND_CDC': 0, 'IND_PROP': 0, 'IND_DISC': 0, 'IND_H': 0, 'IND_INT': 0, 'IND_H_NO_PMG': 0, 'IND_PRIO': 0, 'IND_PLAC': 0, 'IND_PMG': 0, 'IND_RIESGO': 0, 'IND_TRANS': 0}
        if etiqueta_origen == "CDC": flags['IND_CDC'] = 1
        elif etiqueta_origen == "Riesgos": flags['IND_RIESGO'] = 1
        elif etiqueta_origen == "PMG": flags['IND_PMG'] = 1
        
        fila = {
            'CODIGO': codigo_indicador, 'NOMBRE': nombre_limpio, 'DESCRIPCION': nombre_limpio,
            'ACTIVO': 1, 'UNIDAD': unidad, 'RANGO_MINIMO': 0, 'RANGO_MAXIMO': 100,
            'APLICA_DIST_GENERO': 0, 'APLICA_SIN_INFORMACION': 0, 'APLICA_VAL_PERS_JUR': 0,
            'APLICA_DESP_TERRITORIAL': 0, 'VALOR_DEFECTO': 0,
            'AMBITO_COD': ambito, 'DIMENSION_COD': dimension,
            'PERSPECTIVA_COD': None, 'FORMULA_COD': "PORCENTAJE",
            'PROD_ESTRATEGICO_COD': None, 'OBJ_SERVICIO_COD': None,
            'SENTIDO_META': 1, 'TIPO_META': "TOLERANCIA",
            'FACTOR_CUMPLIMIENTO': 10, 'FACTOR_NOCUMPLIMIENTO': 20, 'FACTOR_SOBRECUMPLIMIENTO': 0,
            **flags, 'ANO_ASOCIADO': 2025
        }
        filas_indicadores.append(fila)
    return pd.DataFrame(filas_indicadores), contador_nuevos_global

def ejecutar_fase_4(dataframes_input=None):
    print("\n--- EJECUTANDO FASE 4: INDICADORES IPS ---")
    if dataframes_input is None:
        dataframes_input = cargar_datos_intermedios_fase1()
        if not dataframes_input: return

    dfs_ind_consolidados = []; contador_ind_global = 1
    cols_orden = ['CODIGO', 'NOMBRE', 'DESCRIPCION', 'ACTIVO', 'UNIDAD', 'RANGO_MINIMO', 'RANGO_MAXIMO', 'APLICA_DIST_GENERO', 'APLICA_SIN_INFORMACION', 'APLICA_VAL_PERS_JUR', 'APLICA_DESP_TERRITORIAL', 'VALOR_DEFECTO', 'AMBITO_COD', 'DIMENSION_COD', 'PERSPECTIVA_COD', 'FORMULA_COD', 'PROD_ESTRATEGICO_COD', 'OBJ_SERVICIO_COD', 'SENTIDO_META', 'TIPO_META', 'FACTOR_CUMPLIMIENTO', 'FACTOR_NOCUMPLIMIENTO', 'FACTOR_SOBRECUMPLIMIENTO', 'IND_BGI', 'IND_CDC', 'IND_PROP', 'IND_DISC', 'IND_H', 'IND_INT', 'IND_H_NO_PMG', 'IND_PRIO', 'IND_PLAC', 'IND_PMG', 'IND_RIESGO', 'IND_TRANS', 'ANO_ASOCIADO']

    for etiqueta in ["CDC", "Riesgos", "PMG"]:
        if etiqueta in dataframes_input:
            df_inds, nuevo_cont = transformar_a_indicadores_ips(dataframes_input[etiqueta], etiqueta, contador_ind_global)
            contador_ind_global = nuevo_cont
            if not df_inds.empty:
                fila_titulo = {col: None for col in cols_orden}; fila_titulo['CODIGO'] = f"--- {etiqueta.upper()} INDICADORES ---"
                dfs_ind_consolidados.append(pd.DataFrame([fila_titulo])); dfs_ind_consolidados.append(df_inds)

    if dfs_ind_consolidados:
        df_final = pd.concat(dfs_ind_consolidados, ignore_index=True).reindex(columns=cols_orden)
        df_final.to_excel(ARCHIVOS["INDS_IPS"], index=False)
        aplicar_estilos_indicadores(ARCHIVOS["INDS_IPS"])
        print(f"   -> [OK] Archivo generado: {ARCHIVOS['INDS_IPS']}")

# =============================================================================
# BLOQUE 5: FASE 5 (INDICADORES APLICADOS)
# =============================================================================

def transformar_a_indicadores_aplicados(df_origen, etiqueta_origen, contador_nuevos_global):
    filas_app = []
    for index, row in df_origen.iterrows():
        raw_num = row.get('NÚMERO', '')
        codigo_indicador, _ = limpiar_codigo_indicador_fase5(raw_num, contador_nuevos_global, etiqueta_origen)
        if "INDICADOR_NUEVO" in codigo_indicador: contador_nuevos_global += 1

        _, _, nombre_limpio = parsear_nombre_indicador(row.get('INDICADOR', ''))
        meta_val = row.get('Meta 2025 (%)', 0)
        ponderacion_val = row.get('Ponderador (%)', 0) if etiqueta_origen == "CDC" else None

        # --- CRUCE CON MAPA INTERNO ---
        responsable_raw = row.get('RESPONSABLE CENTRO DE RESPONSABILIDAD', '')
        responsable_clave = normalizar_clave_responsable(responsable_raw)
        cod_ponderado = MAPA_PONDERADOS_INTERNO.get(responsable_clave, None)
        cod_var_auto = f"A_{cod_ponderado}" if cod_ponderado else None

        fila = {
            'INDICADOR_COD': codigo_indicador,
            'NOMBRE_INDICADOR': nombre_limpio,
            'JER_TIPO_COD': 1,
            'CENTRO_RESP_COD': None, 'COD_REGION': None,
            'EMAIL_RESPONSABLE': 'prueba@arbol-logika.com',
            'COD_GENERO': None, 'ANO_MES_INI': 202501, 'ANO_MES_FIN': 202512,
            'COD_ANALISIS_CAUSA': 'RESP_INDICADOR',
            'EMAIL_RESP_ANALISIS_CAUSA': None, 'CREAR_COMENTARIO_FORM': None,
            'ENE': 1, 'FEB': 1, 'MAR': 1, 'ABR': 1, 'MAY': 1, 'JUN': 1,
            'JUL': 1, 'AGO': 1, 'SEP': 1, 'OCT': 1, 'NOV': 1, 'DIC': 1,
            'ORIGEN': None, 'NOTAS_SUPUESTOS': None, 'MOSTRAR_PANEL': None,
            'APLICA_RIESGO': None, 'OBJ_ESTRATEGICO_COD': None, 'OBJ_ESPECIFICO_COD': None,
            'PROD_ESTRATEGICO_COD': None, 'COD_PROGRAMA': None, 'COD_COMPONENTE_PROG': None,
            'TIPO_META_ANUAL': 'PERIODO_ANUAL',
            'COMP_A': f"{codigo_indicador}_A", 'COMP_A_CR': None, 'COMP_A_GEN': None, 'COMP_A_REGION': None,
            'COMP_B': f"{codigo_indicador}_B", 'COMP_B_CR': None, 'COMP_B_GEN': None, 'COMP_B_REGION': None,
            'CONST_A': None, 'META_202512': meta_val,
            'Ponderacion': ponderacion_val, 'COD_PONDERADO': cod_ponderado,
            'FORMULA_VAR_AUTO': 'SUMA_ANUAL', 'COD_VAR_AUTO': cod_var_auto
        }
        filas_app.append(fila)
        
    return pd.DataFrame(filas_app), contador_nuevos_global

def ejecutar_fase_5(dataframes_input=None):
    print("\n--- EJECUTANDO FASE 5: INDICADORES APLICADOS ---")
    if dataframes_input is None:
        dataframes_input = cargar_datos_intermedios_fase1()
        if not dataframes_input: return

    dfs_consolidados = []; contador_ind_global = 1
    
    cols_orden = ['INDICADOR_COD', 'NOMBRE_INDICADOR', 'JER_TIPO_COD', 'CENTRO_RESP_COD', 'COD_REGION', 'EMAIL_RESPONSABLE', 'COD_GENERO', 'ANO_MES_INI', 'ANO_MES_FIN', 'COD_ANALISIS_CAUSA', 'EMAIL_RESP_ANALISIS_CAUSA', 'CREAR_COMENTARIO_FORM', 'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC', 'ORIGEN', 'NOTAS_SUPUESTOS', 'MOSTRAR_PANEL', 'APLICA_RIESGO', 'OBJ_ESTRATEGICO_COD', 'OBJ_ESPECIFICO_COD', 'PROD_ESTRATEGICO_COD', 'COD_PROGRAMA', 'COD_COMPONENTE_PROG', 'TIPO_META_ANUAL', 'COMP_A', 'COMP_A_CR', 'COMP_A_GEN', 'COMP_A_REGION', 'COMP_B', 'COMP_B_CR', 'COMP_B_GEN', 'COMP_B_REGION', 'CONST_A', 'META_202512', 'Ponderacion', 'COD_PONDERADO', 'FORMULA_VAR_AUTO', 'COD_VAR_AUTO']

    for etiqueta in ["CDC", "Riesgos", "PMG"]:
        if etiqueta in dataframes_input:
            df_app, nuevo_cont = transformar_a_indicadores_aplicados(dataframes_input[etiqueta], etiqueta, contador_ind_global)
            contador_ind_global = nuevo_cont
            if not df_app.empty:
                fila_titulo = {col: None for col in cols_orden}; fila_titulo['INDICADOR_COD'] = f"--- {etiqueta.upper()} APLICADOS ---"
                dfs_consolidados.append(pd.DataFrame([fila_titulo])); dfs_consolidados.append(df_app)

    if dfs_consolidados:
        df_final = pd.concat(dfs_consolidados, ignore_index=True).reindex(columns=cols_orden)
        with pd.ExcelWriter(ARCHIVOS["INDS_APP"], engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name="INDICADORES_BRUTA", index=False)
            df_final.to_excel(writer, sheet_name="INDICADORES_ESTILIZADA", index=False)
        aplicar_estilos_variables(ARCHIVOS["INDS_APP"], "INDICADORES_ESTILIZADA")
        print(f"   -> [OK] Archivo generado: {ARCHIVOS['INDS_APP']}")
    else:
        print("   -> [AVISO] No se generaron datos.")

# =============================================================================
# BLOQUE 6: UTILS CARGA
# =============================================================================

def cargar_datos_intermedios_fase1():
    archivo_cargar = ARCHIVOS["ESTILIZADA"]
    if not os.path.exists(archivo_cargar): archivo_cargar = ARCHIVOS["BRUTA"]
    if not os.path.exists(archivo_cargar):
        print(f"[ERROR] No se encontraron planillas procesadas. Ejecuta Fase 1."); return None
    print(f"   -> Cargando datos de: {archivo_cargar}...")
    try:
        xls = pd.ExcelFile(archivo_cargar)
        dataframes = {}
        for hoja in xls.sheet_names:
            if "CDC" in hoja: dataframes["CDC"] = pd.read_excel(xls, sheet_name=hoja)
            elif "PMG" in hoja: dataframes["PMG"] = pd.read_excel(xls, sheet_name=hoja)
            elif "Riesgos" in hoja: dataframes["Riesgos"] = pd.read_excel(xls, sheet_name=hoja)
        return dataframes
    except Exception as e: print(f"[ERROR] Carga fallida: {e}"); return None

def aplicar_estilos_planillas(n): pass 
def aplicar_estilos_variables(n, h): 
    try:
        wb = load_workbook(n); ws = wb[h]; bold = Font(b=True)
        for row in ws.iter_rows():
            if str(row[0].value).startswith("---"): 
                for c in row: c.font = bold
        ws.column_dimensions['B'].width = 50
        wb.save(n)
    except: pass
def aplicar_estilos_aplicadas(n): 
    try:
        wb = load_workbook(n); ws = wb.active; bold = Font(b=True)
        for row in ws.iter_rows():
            if str(row[0].value).startswith("---"): 
                for c in row: c.font = bold
        ws.column_dimensions['B'].width = 50
        wb.save(n)
    except: pass
def aplicar_estilos_indicadores(n):
    try:
        wb = load_workbook(n); ws = wb.active; bold = Font(b=True)
        for row in ws.iter_rows():
            if str(row[0].value).startswith("---"): 
                for c in row: c.font = bold
        ws.column_dimensions['B'].width = 60
        wb.save(n)
    except: pass

# =============================================================================
# BLOQUE 7: MENÚ
# =============================================================================

def menu_principal():
    while True:
        print("\n" + "="*60)
        print("   SISTEMA INTEGRAL IPS 2026 (PORTABLE / NO CSV)")
        print("="*60)
        print("   1. Ejecutar TODO (Fases 1 -> 5)")
        print("   2. Fase 1: Extracción")
        print("   3. Fase 2: Variables IPS")
        print("   4. Fase 3: Variables Aplicadas")
        print("   5. Fase 4: Indicadores IPS")
        print("   6. Fase 5: Indicadores Aplicados")
        print("   0. Salir")
        
        op = input("\n   >> Opción: ").strip()
        if op == "0": break
        
        if op == "1":
            dfs = ejecutar_fase_1()
            if dfs:
                df_v = ejecutar_fase_2(dfs)
                if df_v is not None: ejecutar_fase_3(df_v)
                ejecutar_fase_4(dfs)
                ejecutar_fase_5(dfs)
        elif op == "2": ejecutar_fase_1()
        elif op == "3": ejecutar_fase_2()
        elif op == "4": ejecutar_fase_3()
        elif op == "5": ejecutar_fase_4()
        elif op == "6": ejecutar_fase_5()

if __name__ == "__main__":
    menu_principal()