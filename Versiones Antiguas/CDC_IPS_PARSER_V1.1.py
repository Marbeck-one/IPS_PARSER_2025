import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# --- 1. Configuración ---
nombre_archivo_entrada = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
nombre_hoja = "CDC 2025"
nombre_archivo_salida = "Planilla_Indicadores_Final_Estilizada.xlsx"

# Función de limpieza (Lógica V8)
def limpiar_porcentaje_real(val):
    if pd.isna(val) or val == "": return 0
    if isinstance(val, str):
        limpio = val.replace('%', '').replace(',', '.').strip()
        try: return float(limpio)
        except: return 0
    if isinstance(val, (int, float)): return val * 100
    return 0

try:
    print(f"1. Leyendo archivo y procesando datos...")
    df_raw = pd.read_excel(nombre_archivo_entrada, sheet_name=nombre_hoja, header=None)

    # --- 2. Procesamiento de Datos (Lógica V8) ---
    indices_inicio = []
    for i in range(11, len(df_raw)):
        val = df_raw.iloc[i, 0]
        if pd.notna(val) and str(val).strip() != "" and str(val) != "NÚMERO":
            indices_inicio.append(i)
    
    lista_filas_procesadas = []
    encabezados_raw = df_raw.iloc[10, 0:10].tolist()
    encabezados_finales = ["Meta 2025 (%)" if x == "Meta 2025" else "Ponderador (%)" if x == "Ponderador" else x for x in encabezados_raw]

    for idx in indices_inicio:
        fila_base = df_raw.iloc[idx, 0:10].tolist()
        datos_dict = dict(zip(encabezados_finales, fila_base))
        
        datos_dict["Desc. Op1"] = df_raw.iloc[idx, 10]
        datos_dict["Desc. Op2"] = df_raw.iloc[idx+3, 10]
        datos_dict["Est. Meta Op1"] = df_raw.iloc[idx+3, 11]
        datos_dict["Est. Meta Op2"] = df_raw.iloc[idx+5, 11]
        
        mapa_cols = {
            "Ene": 12, "Feb": 13, "Mar": 15, "Abr": 17, "May": 19, "Jun": 21,
            "Jul": 23, "Ago": 25, "Sept": 27, "Oct": 29, "Nov": 31, "Dic": 33,
            "Cump. Proy.": 34
        }
        for mes, col_idx in mapa_cols.items():
            datos_dict[f"{mes} Ind (%)"] = df_raw.iloc[idx+1, col_idx]
            datos_dict[f"{mes} Op1"] = df_raw.iloc[idx+3, col_idx]
            datos_dict[f"{mes} Op2"] = df_raw.iloc[idx+5, col_idx]
            
        datos_dict["Cumplimiento Meta (%)"] = df_raw.iloc[idx+3, 35]
        datos_dict["Medios Verificación"] = df_raw.iloc[idx, 36]
        datos_dict["Control Cambios"] = df_raw.iloc[idx, 37]
        datos_dict["Instrumentos Gestión"] = df_raw.iloc[idx, 38]
        
        lista_filas_procesadas.append(datos_dict)

    df_final = pd.DataFrame(lista_filas_procesadas)
    df_final = df_final.fillna(0)
    
    cols_porcentaje = [c for c in df_final.columns if "(%)" in c]
    for col in cols_porcentaje:
        df_final[col] = df_final[col].apply(limpiar_porcentaje_real)

    # Exportar datos crudos primero
    df_final.to_excel(nombre_archivo_salida, index=False)
    print("   Datos exportados. Iniciando embellecimiento...")

    # --- 3. Embellecimiento (Estilos) ---
    wb = load_workbook(nombre_archivo_salida)
    ws = wb.active

    # Definir Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
    header_font = Font(color="FFFFFF", bold=True) # Letra blanca negrita
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Columnas que necesitan mucho espacio (Texto largo)
    cols_texto_largo = ["PRODUCTO O PROCESO ESPECÍFICO", "INDICADOR ", "FORMULA", "Desc. Op1", "Desc. Op2", "Medios Verificación", "Control Cambios", "Instrumentos Gestión"]
    
    # Aplicar estilos
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        col_letter = column_cells[0].column_letter
        col_name = column_cells[0].value
        
        # Ancho de columnas
        if col_name in cols_texto_largo:
            ws.column_dimensions[col_letter].width = 40 # Ancho grande
        elif len(str(col_name)) < 5: # Columnas cortas (Meses ej: Ene)
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 20 # Ancho estándar

        for cell in column_cells:
            # Bordes para todos
            cell.border = thin_border
            
            # Estilo Encabezado (Fila 1)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Estilo Cuerpo
                if col_name in cols_texto_largo:
                    # Texto largo: Alineado arriba-izquierda con ajuste de texto
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    # Datos numéricos/cortos: Centrados vertical y horizontalmente
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(nombre_archivo_salida)
    print(f"¡Listo! Archivo 'guapo' generado: {nombre_archivo_salida}")

except Exception as e:
    print(f"Error: {e}")