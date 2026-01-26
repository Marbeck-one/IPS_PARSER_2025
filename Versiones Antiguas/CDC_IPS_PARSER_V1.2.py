import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# --- 1. Configuración ---
nombre_archivo_entrada = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
nombre_hoja = "CDC 2025"
nombre_archivo_salida = "CDC_Extraccion_estilizada_2025.xlsx"

# Función de limpieza (Escala 0-100)
def limpiar_porcentaje_real(val):
    if pd.isna(val) or val == "": return 0
    if isinstance(val, str):
        limpio = val.replace('%', '').replace(',', '.').strip()
        try: return float(limpio)
        except: return 0
    if isinstance(val, (int, float)): return val * 100
    return 0

try:
    print(f"1. Leyendo archivo: {nombre_archivo_entrada}...")
    # Leemos el Excel
    df_raw = pd.read_excel(nombre_archivo_entrada, sheet_name=nombre_hoja, header=None)

    # --- 2. Procesamiento de Datos (Lógica Automática) ---
    print("2. Detectando indicadores y extrayendo datos...")
    
    indices_inicio = []
    # Escaneo columna A
    for i in range(11, len(df_raw)):
        val = df_raw.iloc[i, 0]
        if pd.notna(val) and str(val).strip() != "" and str(val) != "NÚMERO":
            indices_inicio.append(i)
    
    lista_filas_procesadas = []
    
    # Encabezados base
    encabezados_raw = df_raw.iloc[10, 0:10].tolist()
    encabezados_finales = ["Meta 2025 (%)" if x == "Meta 2025" else "Ponderador (%)" if x == "Ponderador" else x for x in encabezados_raw]

    for idx in indices_inicio:
        # A) Base
        fila_base = df_raw.iloc[idx, 0:10].tolist()
        datos_dict = dict(zip(encabezados_finales, fila_base))
        
        # B) Operandos
        datos_dict["Desc. Op1"] = df_raw.iloc[idx, 10]
        datos_dict["Desc. Op2"] = df_raw.iloc[idx+3, 10]
        datos_dict["Est. Meta Op1"] = df_raw.iloc[idx+3, 11]
        datos_dict["Est. Meta Op2"] = df_raw.iloc[idx+5, 11]
        
        # C) Meses
        mapa_cols = {
            "Ene": 12, "Feb": 13, "Mar": 15, "Abr": 17, "May": 19, "Jun": 21,
            "Jul": 23, "Ago": 25, "Sept": 27, "Oct": 29, "Nov": 31, "Dic": 33,
            "Cump. Proy.": 34
        }
        for mes, col_idx in mapa_cols.items():
            datos_dict[f"{mes} Ind (%)"] = df_raw.iloc[idx+1, col_idx]
            datos_dict[f"{mes} Op1"] = df_raw.iloc[idx+3, col_idx]
            datos_dict[f"{mes} Op2"] = df_raw.iloc[idx+5, col_idx]
            
        # D) Finales
        datos_dict["Cumplimiento Meta (%)"] = df_raw.iloc[idx+3, 35]
        datos_dict["Medios Verificación"] = df_raw.iloc[idx, 36]
        datos_dict["Control Cambios"] = df_raw.iloc[idx, 37]
        datos_dict["Instrumentos Gestión"] = df_raw.iloc[idx, 38]
        
        lista_filas_procesadas.append(datos_dict)

    # Crear DF y limpiar
    df_final = pd.DataFrame(lista_filas_procesadas)
    df_final = df_final.fillna(0)
    
    # Convertir porcentajes
    cols_porcentaje = [c for c in df_final.columns if "(%)" in c]
    for col in cols_porcentaje:
        df_final[col] = df_final[col].apply(limpiar_porcentaje_real)

    # Exportar datos crudos
    df_final.to_excel(nombre_archivo_salida, index=False)
    print("   Datos exportados. Iniciando diseño visual...")

    # --- 3. Embellecimiento (Estilos Personalizados) ---
    wb = load_workbook(nombre_archivo_salida)
    ws = wb.active

    # Definir Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Lógica de anchos
    cols_texto_muy_largo = ["PRODUCTO O PROCESO ESPECÍFICO", "INDICADOR ", "FORMULA", "Desc. Op1", "Desc. Op2", "Medios Verificación", "Control Cambios", "Instrumentos Gestión"]
    
    # Nuevas columnas anchas (E, F, G, H)
    cols_ancho_medio = ["UNIDAD", "RESPONSABLE CENTRO DE RESPONSABILIDAD", "GESTOR", "SUPERVISORES"]

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        col_letter = column_cells[0].column_letter
        col_name = column_cells[0].value
        
        # --- APLICAR ANCHOS ---
        if col_name in cols_texto_muy_largo:
            ws.column_dimensions[col_letter].width = 45 # Texto muy largo
        elif col_name in cols_ancho_medio:
            ws.column_dimensions[col_letter].width = 30 # E, F, G, H (Ancho medio para nombres/unidades)
        elif len(str(col_name)) < 5: 
            ws.column_dimensions[col_letter].width = 12 # Columnas cortas (Meses)
        else:
            ws.column_dimensions[col_letter].width = 18 # Estándar

        for cell in column_cells:
            cell.border = thin_border
            
            # Encabezado
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Cuerpos
                if col_name in cols_texto_muy_largo or col_name in cols_ancho_medio:
                    # Alineación Izquierda para texto
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    # Centrado para números
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(nombre_archivo_salida)
    print(f"¡Listo! Archivo final guardado como: {nombre_archivo_salida}")

except Exception as e:
    print(f"ERROR: {e}")