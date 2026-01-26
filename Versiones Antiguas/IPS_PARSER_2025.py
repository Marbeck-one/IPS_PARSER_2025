import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# ==========================================
# 1. MOTOR DE PROCESAMIENTO (LÓGICA CENTRAL)
# ==========================================

def limpiar_porcentaje_real(val):
    """Convierte texto '20%' o número 0.2 a valor limpio 20."""
    if pd.isna(val) or val == "": return 0
    if isinstance(val, str):
        limpio = val.replace('%', '').replace(',', '.').strip()
        try: return float(limpio)
        except: return 0
    if isinstance(val, (int, float)): return val * 100
    return 0

def detectar_encabezados(df):
    """Escanea las primeras 25 filas buscando dónde empieza la tabla."""
    for i in range(25):
        fila = df.iloc[i].astype(str).tolist()
        # La clave es encontrar "NÚMERO" y alguna columna de "INDICADOR"
        if "NÚMERO" in fila and any("INDICADOR" in s for s in fila):
            return i, fila
    raise ValueError("ERROR: No se encontró la fila de encabezados (NÚMERO, INDICADOR...)")

def procesar_hoja_universal(ruta_archivo, nombre_hoja, prefijo_salida, opciones_salida):
    """
    Procesa una hoja (CDC o Riesgos) adaptándose a sus columnas.
    opciones_salida: 1=Bruta, 2=Estilizada, 3=Ambas
    """
    print(f"\n--- Iniciando proceso para: {prefijo_salida} ---")
    print(f"1. Leyendo archivo: {ruta_archivo} [{nombre_hoja}]...")
    
    try:
        df_raw = pd.read_excel(ruta_archivo, sheet_name=nombre_hoja, header=None)
    except Exception as e:
        print(f"ERROR CRÍTICO leyendo el archivo: {e}")
        return

    # A. Detección Dinámica de Estructura
    try:
        idx_header, encabezados = detectar_encabezados(df_raw)
        print(f"   -> Encabezados detectados en la fila {idx_header + 1}")
    except ValueError as e:
        print(e)
        return

    # Mapeo de columnas (Nombre -> Índice)
    # Normalizamos nombres (quitamos saltos de línea y espacios extra)
    mapa_cols = {str(nombre).strip().replace("\n", " "): i for i, nombre in enumerate(encabezados)}
    
    # Función interna para buscar columnas de forma flexible
    def buscar_col(claves):
        if isinstance(claves, str): claves = [claves]
        for c in claves:
            for nombre_real, idx in mapa_cols.items():
                if c.lower() in nombre_real.lower():
                    return idx
        return None

    # Identificación de Índices Clave
    idx_num = buscar_col("NÚMERO")
    idx_meta = buscar_col("Meta 2025")
    idx_ponderador = buscar_col("Ponderador") # Puede no existir en Riesgos
    
    idx_op_desc = buscar_col("Operandos") # Columna K usualmente
    idx_op_est = buscar_col(["Operandos Estimados", "Operandos  Estimados"]) # Columna L usualmente
    
    # Columnas finales de texto
    idx_medios = buscar_col("Medios de Verificación")
    idx_control = buscar_col("Control de Cambios")
    idx_inst = buscar_col(["Instrumentos de Gestión", "Instrumentos Gestión"])
    idx_cump_meta = buscar_col("% Cumplimiento de Meta")

    # B. Extracción de Datos
    indices_inicio = []
    # Escaneo columna A desde la fila siguiente al encabezado
    for i in range(idx_header + 1, len(df_raw)):
        val = df_raw.iloc[i, idx_num]
        if pd.notna(val) and str(val).strip() != "" and str(val) != "NÚMERO":
            indices_inicio.append(i)
            
    print(f"   -> Se encontraron {len(indices_inicio)} indicadores.")
    
    lista_datos = []
    
    # Preparamos el ciclo de meses (Ene a Dic + Proyección)
    meses_std = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sept", "Oct", "Nov", "Dic"]
    mapa_meses_indices = {}
    
    # Buscamos dónde está cada mes en ESTA hoja específica
    for mes in meses_std:
        idx = buscar_col(f"{mes}.") # Busca "Ene.", "Feb.", etc.
        if idx: mapa_meses_indices[mes] = idx
        
    # Agregamos "Cumplimiento Proyectado" como un mes más
    idx_proy = buscar_col(["Cumplimiento Proyectado", "Cumplimiento\nProyectado"])
    if idx_proy: mapa_meses_indices["Cump. Proy."] = idx_proy

    # --- BUCLE DE EXTRACCIÓN ---
    for idx in indices_inicio:
        fila = {}
        
        # 1. Datos Identificatorios
        fila["NÚMERO"] = df_raw.iloc[idx, idx_num]
        
        col_prod = buscar_col("PRODUCTO")
        if col_prod: fila["PRODUCTO O PROCESO ESPECÍFICO"] = df_raw.iloc[idx, col_prod]
        
        col_ind = buscar_col("INDICADOR")
        if col_ind: fila["INDICADOR"] = df_raw.iloc[idx, col_ind]
        
        col_form = buscar_col("FORMULA")
        if col_form: fila["FORMULA"] = df_raw.iloc[idx, col_form]
        
        col_unidad = buscar_col("UNIDAD")
        if col_unidad: fila["UNIDAD"] = df_raw.iloc[idx, col_unidad]
        
        # 2. Responsables
        col_resp = buscar_col("RESPONSABLE CENTRO")
        if col_resp: fila["RESPONSABLE CENTRO DE RESPONSABILIDAD"] = df_raw.iloc[idx, col_resp]
        
        col_gestor = buscar_col("GESTOR")
        if col_gestor: fila["GESTOR"] = df_raw.iloc[idx, col_gestor]
        
        col_sup = buscar_col("SUPERVISORES")
        if col_sup: fila["SUPERVISORES"] = df_raw.iloc[idx, col_sup]
        
        # 3. Metas y Ponderadores
        if idx_meta: fila["Meta 2025 (%)"] = df_raw.iloc[idx, idx_meta]
        
        # Manejo especial: Si no hay columna ponderador (Riesgos), ponemos 0
        if idx_ponderador: fila["Ponderador (%)"] = df_raw.iloc[idx, idx_ponderador]
        else: fila["Ponderador (%)"] = 0 
        
        # 4. Operandos (Descripciones y Metas)
        if idx_op_desc:
            fila["Desc. Op1"] = df_raw.iloc[idx, idx_op_desc]
            fila["Desc. Op2"] = df_raw.iloc[idx+3, idx_op_desc]
        if idx_op_est:
            fila["Est. Meta Op1"] = df_raw.iloc[idx+3, idx_op_est]
            fila["Est. Meta Op2"] = df_raw.iloc[idx+5, idx_op_est]
            
        # 5. Ciclo Mensual (Meses + Proyección)
        # La lógica de intervalos es la misma para CDC y Riesgos: +1, +3, +5
        for nombre_mes, col_idx in mapa_meses_indices.items():
            fila[f"{nombre_mes} Ind (%)"] = df_raw.iloc[idx+1, col_idx]
            fila[f"{nombre_mes} Op1"] = df_raw.iloc[idx+3, col_idx]
            fila[f"{nombre_mes} Op2"] = df_raw.iloc[idx+5, col_idx]
            
        # 6. Columnas Finales
        if idx_cump_meta: fila["Cumplimiento Meta (%)"] = df_raw.iloc[idx+3, idx_cump_meta]
        if idx_medios: fila["Medios Verificación"] = df_raw.iloc[idx, idx_medios]
        if idx_control: fila["Control Cambios"] = df_raw.iloc[idx, idx_control]
        if idx_inst: fila["Instrumentos Gestión"] = df_raw.iloc[idx, idx_inst]
        
        lista_datos.append(fila)

    # C. Limpieza y Creación del DataFrame
    df_final = pd.DataFrame(lista_datos).fillna(0)
    
    # Limpiar porcentajes
    for col in df_final.columns:
        if "(%)" in col:
            df_final[col] = df_final[col].apply(limpiar_porcentaje_real)

    # D. Exportación según Opciones
    nombre_bruta = f"{prefijo_salida}_Extraccion_bruta_2025.xlsx"
    nombre_estilizada = f"{prefijo_salida}_Extraccion_estilizada_2025.xlsx"
    
    # Opción 1 o 3: Bruta
    if opciones_salida in ["1", "3"]:
        print(f"2. Guardando archivo Bruto: {nombre_bruta}")
        df_final.to_excel(nombre_bruta, index=False)
        
    # Opción 2 o 3: Estilizada
    if opciones_salida in ["2", "3"]:
        print(f"3. Guardando archivo Estilizado: {nombre_estilizada}")
        df_final.to_excel(nombre_estilizada, index=False)
        aplicar_estilos(nombre_estilizada)
        
    print("   -> ¡Proceso finalizado para esta hoja!")

def aplicar_estilos(nombre_archivo):
    """Aplica el maquillaje visual: Anchos, Colores y Bordes."""
    print("   -> Aplicando estilos visuales...")
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Configuración de anchos (Funciona igual para CDC y Riesgos)
    cols_texto_largo = ["PRODUCTO O PROCESO ESPECÍFICO", "INDICADOR", "FORMULA", "Desc. Op1", "Desc. Op2", "Medios Verificación", "Control Cambios", "Instrumentos Gestión"]
    cols_responsables = ["UNIDAD", "RESPONSABLE CENTRO DE RESPONSABILIDAD", "GESTOR", "SUPERVISORES"]

    for col_cells in ws.columns:
        col_name = col_cells[0].value
        col_letter = col_cells[0].column_letter
        
        # --- APLICAR ANCHOS ---
        if col_name in cols_texto_largo:
            ws.column_dimensions[col_letter].width = 40 
        elif col_name in cols_responsables:
            ws.column_dimensions[col_letter].width = 30
        elif col_name and len(str(col_name)) < 6: 
            ws.column_dimensions[col_letter].width = 10 
        else:
            ws.column_dimensions[col_letter].width = 18

        for cell in col_cells:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                if col_name in cols_texto_largo or col_name in cols_responsables:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(nombre_archivo)

# ==========================================
# 2. INTERFAZ DE USUARIO (MENÚ)
# ==========================================

def menu_principal():
    print("="*50)
    print("   SISTEMA DE EXTRACCIÓN DE INDICADORES 2025")
    print("="*50)
    
    # Archivo de entrada (Debe estar en la misma carpeta)
    archivo_input = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
    
    if not os.path.exists(archivo_input):
        print(f"ERROR: No se encuentra el archivo '{archivo_input}'")
        return

    # --- PREGUNTA 1: CDC ---
    procesar_cdc = input("\n¿Deseas procesar la hoja 'CDC 2025'? (s/n): ").lower().strip()
    
    if procesar_cdc == 's':
        print("\nOpciones de formato para CDC:")
        print("1. Solo Bruta")
        print("2. Solo Estilizada")
        print("3. Ambas")
        opcion_cdc = input("Elige una opción (1-3): ").strip()
        
        if opcion_cdc in ["1", "2", "3"]:
            procesar_hoja_universal(archivo_input, "CDC 2025", "CDC", opcion_cdc)
        else:
            print("Opción inválida. Saltando CDC.")

    # --- PREGUNTA 2: RIESGOS ---
    procesar_riesgos = input("\n¿Deseas procesar la hoja 'Riesgos 2025'? (s/n): ").lower().strip()
    
    if procesar_riesgos == 's':
        print("\nOpciones de formato para Riesgos:")
        print("1. Solo Bruta")
        print("2. Solo Estilizada")
        print("3. Ambas")
        opcion_riesgos = input("Elige una opción (1-3): ").strip()
        
        if opcion_riesgos in ["1", "2", "3"]:
            # Nota: Asegúrate que el nombre de la hoja en el Excel sea exactamente este
            # Si tienes dudas, abre el Excel y revisa la pestaña. 
            # Asumiré que se llama "Riesgos 2025" o similar.
            procesar_hoja_universal(archivo_input, "Riesgos 2025", "Riesgos", opcion_riesgos)
        else:
            print("Opción inválida. Saltando Riesgos.")

    print("\n" + "="*50)
    print("PROCESO TERMINADO. ¡Gracias!")
    print("="*50)

if __name__ == "__main__":
    menu_principal()