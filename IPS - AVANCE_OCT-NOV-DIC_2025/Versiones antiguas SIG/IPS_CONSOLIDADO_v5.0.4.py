import pandas as pd
import os
import sys
import glob
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v5.0.4 - "B9 TARGETING" EDITION
# =============================================================================

class IPSParserV504:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_v5.0.4.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.variable_data = [] 
        
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible'
        self.blacklist_auto = ["NÚMERO", "NUMERO", "N°", "NO", "Nº", "ITEM"]
        
        self.decisions = {"use_segment": None, "use_col_a_as_num": None}
        self.file_teams = {} # Cache de equipos por archivo

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA v5.0.4")
        print("="*60)
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        print("\n2. ¿Cómo manejar filas OCULTAS?")
        print("   [v] Automático: Procesar SOLO VISIBLES (Recomendado).")
        print("   [t] Automático: Procesar TODO.")
        resp_h = input("   >> Elija opción (Enter=v): ").lower().strip()
        self.opt_hidden_strategy = 'all' if resp_h == 't' else 'visible'
        print("-" * 60)

    def ask_user_decision(self, key, prompt_text):
        if self.decisions.get(key) is not None: return self.decisions[key]
        print(f"\n[ATENCIÓN] {prompt_text}")
        while True:
            resp = input("   >> ¿Confirmar? (S/N): ").strip().upper()
            if resp in ["S", "N"]:
                self.decisions[key] = (resp == "S")
                return self.decisions[key]

    # --- UTILIDADES ---
    def is_fully_enclosed_by_parens(self, text):
        if not text.startswith("(") or not text.endswith(")"): return False
        balance = 0
        for i, char in enumerate(text):
            if char == '(': balance += 1
            elif char == ')': balance -= 1
            if balance == 0 and i < len(text) - 1: return False
        return balance == 0

    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "": return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix = ""
        core = f_clean
        f_type = "CUOCIENTE"
        if match:
            suffix = match.group(1)
            core = f_clean[:match.start()].strip()
            f_type = "PORCENTAJE"
        while self.is_fully_enclosed_by_parens(core):
            core = core[1:-1].strip()
        return core + suffix, f_type

    def parse_indicator_text(self, text):
        if pd.isna(text) or str(text).strip() == "": return "", "No identificado", "No identificado"
        text_str = str(text).strip()
        lines = text_str.split('\n')
        first_line = lines[0].strip()
        match = re.search(r'^[\d\)\.\-\s]*([^/]+)/(.+)', first_line)
        clean_text = text_str
        dim = "No identificado"
        amb = "No identificado"
        if match:
            dim = match.group(1).strip()
            amb = match.group(2).strip()
            if len(lines) > 1: clean_text = "\n".join(lines[1:]).strip()
            else: clean_text = ""
        while self.is_fully_enclosed_by_parens(clean_text):
            clean_text = clean_text[1:-1].strip()
        return clean_text, dim, amb

    def transform_percentage(self, val, col_name):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or val == "No aplica": return val
        if "Valor" in str(val) or "Operando" in str(val): return "" 
        try:
            clean_val = str(val).replace(",", ".")
            num = float(clean_val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    # --- LÓGICA DE EQUIPO MEJORADA (TARGET B9) ---
    def determine_team(self, df, file_name, sheet_name):
        """
        1. Busca etiquetas 'Centro de Responsabilidad'.
        2. Si falla, busca DIRECTAMENTE en celdas B9, B8, B12.
        3. Si falla, pregunta al usuario.
        """
        # 1. Cache Check
        if file_name in self.file_teams: return self.file_teams[file_name]

        candidate = None
        
        # 2. Búsqueda por Etiqueta (Standard)
        limit_row = 20
        limit_col = min(15, len(df.columns))
        for r in range(min(limit_row, len(df))):
            for c in range(limit_col):
                val = str(df.iloc[r, c]).strip().upper()
                if "CENTRO DE RESPONSABILIDAD" in val or "DIRECCIÓN REGIONAL" in val or "DIRECCION REGIONAL" in val:
                    parts = val.split(":")
                    if len(parts) > 1 and parts[1].strip(): 
                        candidate = parts[1].strip()
                    elif c + 1 < len(df.columns): 
                        next_val = str(df.iloc[r, c+1]).strip()
                        if next_val and next_val.upper() != "NAN": candidate = next_val
                    if not candidate:
                        candidate = val.replace("CENTRO DE RESPONSABILIDAD", "").replace(":", "").strip()
                    break
            if candidate: break

        # 3. ESTRATEGIA "MIRA FRANCOTIRADOR" (B9, B8, B12)
        # Si no encontramos etiqueta, miramos las coordenadas donde suelen estar los nombres "sueltos"
        if not candidate:
            # Lista de coordenadas prioritarias (Fila index 0-based, Columna index 0-based)
            # B9 -> Fila 8, Col 1
            # B8 -> Fila 7, Col 1
            # B12 -> Fila 11, Col 1
            targets = [(8, 1), (7, 1), (11, 1), (8, 0), (7, 0)] 
            
            for r_idx, c_idx in targets:
                if r_idx < len(df) and c_idx < len(df.columns):
                    val = str(df.iloc[r_idx, c_idx]).strip()
                    val_up = val.upper()
                    
                    # Validar que NO sea basura ni cabecera
                    if (len(val) > 3 and 
                        val_up != "NAN" and 
                        "INDICADOR" not in val_up and 
                        "FECHA" not in val_up and 
                        "ELABORADO" not in val_up and
                        "NUMERO" not in val_up):
                        
                        candidate = val
                        print(f"   [AUTO-DETECT] Equipo encontrado por posición ({r_idx+1},{get_column_letter(c_idx+1)}): {candidate}")
                        break

        # 4. Fallback Manual
        if not candidate or candidate.upper() in ["NO APLICA", "NAN"]:
            print(f"\n{'!'*60}")
            print(f"[DECISIÓN MANUAL] No detecté el EQUIPO en: {file_name}")
            print(f"Hoja: {sheet_name}. Busqué etiquetas y en celda B9.")
            print(f"{'!'*60}")
            user_input = input(">> Escribe el nombre del EQUIPO: ").strip()
            candidate = user_input if user_input else "No aplica"
        
        self.file_teams[file_name] = candidate
        return candidate

    # --- LÓGICA DE BÚSQUEDA DE FILAS (OPERANDOS) ---
    def get_real_data_row_index(self, df, start_idx, col_check_idx):
        if start_idx >= len(df): return start_idx
        row_str = "".join([str(x).upper() for x in df.iloc[start_idx].values])
        if "VALOR INDICADOR" in row_str or "OPERANDO" in row_str:
            return start_idx + 1
        return start_idx

    def find_operand_offsets(self, df, start_row, c_map, ignored_rows):
        offsets = []
        max_look = 12
        cols_to_check = []
        if c_map.get("op_est") is not None: cols_to_check.append(c_map["op_est"])
        if c_map.get("Oct.") is not None: cols_to_check.append(c_map["Oct."])
        
        for offset in range(1, max_look + 1):
            target = start_row + offset
            if target >= len(df) or target in ignored_rows: continue
            row_str = "".join([str(x).upper() for x in df.iloc[target].values if pd.notna(x)])
            
            # Filtros agresivos de basura
            if "VALOR INDICADOR" in row_str: continue
            if "OPERANDO 1=" in row_str or "OPERANDO 2=" in row_str: continue
            if "OPERANDO 1 =" in row_str or "OPERANDO 2 =" in row_str: continue
            
            is_data = False
            if c_map["op_desc"] is not None:
                desc = str(df.iloc[target, c_map["op_desc"]]).strip()
                if desc and desc.lower() != "nan" and len(desc) > 3: is_data = True

            if not is_data:
                for col_idx in cols_to_check:
                    val = str(df.iloc[target, col_idx]).strip()
                    if any(c.isdigit() for c in val):
                        is_data = True; break
            
            if is_data:
                offsets.append(offset)
                if len(offsets) == 2: break
        
        return (offsets[0] if len(offsets) > 0 else None, 
                offsets[1] if len(offsets) > 1 else None)

    # --- PROCESAMIENTO ---
    def process_folder(self):
        files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in files if "IPS_CONSOLIDADO" not in f and not os.path.basename(f).startswith("~$")]
        if not valid_files: print("[ERROR] No hay archivos."); sys.exit()
        self.configure()
        
        for idx_file, file_path in enumerate(valid_files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            print(f"\n>>> PROCESANDO: {file_name}")
            
            try: xls = pd.ExcelFile(file_path)
            except: continue

            for sheet in xls.sheet_names:
                try: df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except: continue
                ignored_rows = set()
                
                # 1. DETECTAR EQUIPO (Mejorado con B9)
                global_center = self.determine_team(df, file_name, sheet)
                
                # 2. DETECTAR BLOQUES
                header_indices = []
                for idx, row in df.iterrows():
                    vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    if "INDICADOR" in vals and ("FORMULA" in vals or "FÓRMULA" in vals):
                        header_indices.append(idx)
                    elif "NÚMERO" in vals and "INDICADOR" in vals:
                        if idx not in header_indices: header_indices.append(idx)
                
                if not header_indices: continue

                sheet_rows = []
                for loop_idx, h_idx in enumerate(header_indices):
                    end_idx = header_indices[loop_idx + 1] if loop_idx + 1 < len(header_indices) else len(df)
                    
                    # Segmento
                    current_segment = "GENERAL"
                    if h_idx > 0:
                        prev = df.iloc[h_idx - 1].dropna()
                        if len(prev) == 1:
                            cand = str(prev.iloc[0]).strip()
                            if len(cand) < 60:
                                msg = f"Hoja '{sheet}': Detectado título '{cand}'. ¿Es SEGMENTO?"
                                if self.ask_user_decision("use_segment", msg): current_segment = cand

                    # Mapeo
                    headers = [str(h).strip() for h in df.iloc[h_idx]]
                    def fc(names):
                        for i, h in enumerate(headers):
                            if any(n.lower() in str(h).lower() for n in names): return i
                        return None

                    col_num = fc(["NÚMERO", "NUMERO", "N°"])
                    if col_num is None:
                        if self.ask_user_decision("use_col_a_as_num", f"Hoja '{sheet}': Sin col NÚMERO. ¿Usar Col A?"): col_num = 0
                        else: continue

                    c_map = {
                        "num": col_num, "prod": fc(["PRODUCTO"]), "ind": fc(["INDICADOR"]), "form": fc(["FORMULA"]),
                        "uni": fc(["UNIDAD"]), "resp": fc(["RESPONSABLE"]), "gest": fc(["GESTOR"]),
                        "sup": fc(["SUPERVISORES"]), "meta": fc(["Meta 2025", "Meta 2026", "Meta"]),
                        "pond": fc(["Ponderador"]), "op_desc": fc(["Operandos"]), 
                        "op_est": fc(["Operandos Estimados", "Estimados Meta"]),
                        "proy": fc(["Cumplimiento Proyectado", "Proyectado"]), "cump_meta": fc(["% Cumplimiento"]),
                        "medios": fc(["Medios"]), "control": fc(["Control"]), "inst": fc(["Instrumentos"])
                    }
                    
                    months_list = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                                   "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                                   "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                    m_map = {m: fc([m]) for m in months_list}
                    c_map.update(m_map)

                    for i in range(h_idx + 1, end_idx):
                        if i in ignored_rows: continue
                        raw_num = str(df.iloc[i, c_map["num"]]).strip() if c_map["num"] is not None else ""
                        if not raw_num or raw_num.upper() in self.blacklist_auto or "VALOR" in raw_num.upper(): continue
                        if "NUEVO" not in raw_num.upper() and not re.match(r'^\d', raw_num): continue

                        check_c = c_map["meta"] if c_map["meta"] else (m_map["Oct."] if m_map["Oct."] else None)
                        idx_ind_data = self.get_real_data_row_index(df, i, check_c)
                        off1, off2 = self.find_operand_offsets(df, i, c_map, ignored_rows)
                        idx_op1 = (i + off1) if off1 else idx_ind_data 
                        idx_op2 = (i + off2) if off2 else idx_ind_data

                        def gd(r_idx, c_idx):
                            if c_idx is None or r_idx >= len(df): return ""
                            val = df.iloc[r_idx, c_idx]
                            return val if pd.notna(val) else ""

                        row_data = {
                            "ARCHIVO": file_name, "HOJA": sheet, "EQUIPO": global_center, "SEGMENTO": current_segment,
                            "TIPO INDICADOR": "CDC" if "CDC" in sheet.upper() else "PMG",
                            "NÚMERO": raw_num,
                            "PRODUCTO O PROCESO ESPECÍFICO": gd(i, c_map["prod"]),
                            "INDICADOR": self.parse_indicator_text(gd(i, c_map["ind"]))[0],
                            "DIMENSIÓN": self.parse_indicator_text(gd(i, c_map["ind"]))[1],
                            "ÁMBITO": self.parse_indicator_text(gd(i, c_map["ind"]))[2],
                            "FORMULA": self.analyze_formula(gd(i, c_map["form"]))[0],
                            "TIPO FORMULA": self.analyze_formula(gd(i, c_map["form"]))[1],
                            "UNIDAD": gd(i, c_map["uni"]), "RESPONSABLE": gd(i, c_map["resp"]),
                            "GESTOR": gd(i, c_map["gest"]), "SUPERVISORES": gd(i, c_map["sup"]),
                            "Meta 2026": self.transform_percentage(gd(idx_ind_data, c_map["meta"]), "Meta"),
                            "Ponderador": self.transform_percentage(gd(idx_ind_data, c_map["pond"]), "Pond"),
                            "Descripción Operando 1": gd(idx_op1, c_map["op_desc"]),
                            "Meta Operando 1 (Valor)": gd(idx_op1, c_map["op_est"]),
                            "Descripción Operando 2": gd(idx_op2, c_map["op_desc"]),
                            "Meta Operando 2 (Valor)": gd(idx_op2, c_map["op_est"]),
                            "Cumplimiento Proyectado Op 1": gd(idx_op1, c_map["proy"]),
                            "Cumplimiento Proyectado Op 2": gd(idx_op2, c_map["proy"]),
                            "% Cumplimiento de Meta": self.transform_percentage(gd(idx_op1, c_map["cump_meta"]), "%"),
                            "Medios de Verificación": gd(idx_ind_data, c_map["medios"]),
                            "Control de Cambios": gd(idx_ind_data, c_map["control"]),
                            "Instrumentos de Gestión": gd(idx_ind_data, c_map["inst"])
                        }

                        for m_key, m_col in m_map.items():
                            row_data[f"{m_key} Op 1"] = gd(idx_op1, m_col)
                            row_data[f"{m_key} Op 2"] = gd(idx_op2, m_col)

                        self.flat_data.append(row_data)
                        sheet_rows.append(row_data)

                        # DATOS VARIABLE
                        meses_v = [(10, "Oct."), (11, "Nov."), (12, "Dic.")]
                        for m_num, m_txt in meses_v:
                            col_m = m_map.get(m_txt)
                            if col_m is not None:
                                v1 = gd(idx_op1, col_m)
                                if v1 and "Valor" not in str(v1):
                                    self.variable_data.append({
                                        "ANO": 2025, "MES": m_num, "VARIABLE_COD": f"{raw_num}_A",
                                        "CENTRO_RESP_COD": global_center, "COD_REGION": 0,
                                        "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "",
                                        "VALOR_TOTAL": v1, "ARCHIVO": file_name, "HOJA": sheet
                                    })
                                v2 = gd(idx_op2, col_m)
                                if v2 and "Valor" not in str(v2):
                                    self.variable_data.append({
                                        "ANO": 2025, "MES": m_num, "VARIABLE_COD": f"{raw_num}_B",
                                        "CENTRO_RESP_COD": global_center, "COD_REGION": 0,
                                        "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "",
                                        "VALOR_TOTAL": v2, "ARCHIVO": file_name, "HOJA": sheet
                                    })

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"   -> {len(sheet_rows)} indicadores extraídos.")

        self.export_excel()

    def export_excel(self):
        print(f"\nGenerando Excel Maestro...")
        wb = Workbook()
        ws = wb.active; ws.title = "Carga Bruta"
        if self.flat_data:
            keys = list(self.flat_data[0].keys())
            if "SEGMENTO" in keys:
                keys.remove("SEGMENTO"); idx_eq = keys.index("EQUIPO") + 1 if "EQUIPO" in keys else 0
                keys.insert(idx_eq, "SEGMENTO")
            ws.append(keys)
            fill = PatternFill("solid", fgColor="002060"); font = Font(color="FFFFFF", bold=True)
            for c in ws[1]: c.fill = fill; c.font = font
            for d in self.flat_data: ws.append([d.get(k, "") for k in keys])

        ws_s = wb.create_sheet("Planilla Estilizada")
        styles = {
            'file': PatternFill("solid", fgColor="000000"), 'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"), 'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True), 'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }
        r_idx = 1
        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            c = ws_s.cell(r_idx, 1, f"ARCHIVO: {fname}"); c.fill=styles['file']; c.font=styles['w_font']; ws_s.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=60); r_idx+=1
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_s.cell(r_idx, 1, f"HOJA: {sname}"); c.fill=styles['sheet']; c.font=styles['w_font']; ws_s.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=60); r_idx+=1
                disp_keys = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                for i, k in enumerate(disp_keys, 1):
                    c = ws_s.cell(r_idx, i, k); c.fill=styles['head']; c.font=styles['b_font']; c.border=styles['border']
                r_idx+=1
                for row in rows:
                    for i, k in enumerate(disp_keys, 1):
                        c = ws_s.cell(r_idx, i, row[k]); c.border=styles['border']; c.alignment=Alignment(wrapText=True, vertical='top')
                    r_idx+=1
                r_idx+=1

        ws_v = wb.create_sheet("DATOS_VARIABLE")
        h_vars = ["ANO", "MES", "VARIABLE_COD", "CENTRO_RESP_COD", "COD_REGION", "VALOR_M", "VALOR_F", "VALOR_S", "VALOR_J", "VALOR_TOTAL", "ARCHIVO", "HOJA"]
        ws_v.append(h_vars)
        for c in ws_v[1]: c.fill = fill; c.font = font
        for item in self.variable_data: ws_v.append([item.get(k, "") for k in h_vars])

        try: wb.save(self.output_file); print(f"[EXITO] Guardado en: {self.output_file}")
        except Exception as e: print(f"[ERROR] {e}")

if __name__ == "__main__":
    path = input("Ruta: ").strip() or os.getcwd()
    IPSParserV504(path).process_folder()