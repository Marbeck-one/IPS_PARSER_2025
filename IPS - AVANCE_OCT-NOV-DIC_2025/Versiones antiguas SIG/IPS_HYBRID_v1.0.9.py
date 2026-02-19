import pandas as pd
import os
import sys
import glob
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_HYBRID_v1.0.9 - OPERAND FIX & ROBUSTNESS
# =============================================================================

class IPSParserHybridV109:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_v1.0.9_OCT-NOV-DIC_2025.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.variable_data = [] 
        self.new_indicator_count = 1
        
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible'
        self.blacklist_auto = ["NÚMERO", "NUMERO", "N°", "NO", "Nº", "ITEM", "INDICADOR", "PRODUCTO"]
        
        self.memory_skip = set()      
        self.memory_generate = False
        self.memory_skip_empty = False
        self.known_segments = set(["HOMBRE", "MUJER", "HOMBRES", "MUJERES", "TOTAL PAÍS", "TOTAL PAIS"])
        self.decisions = {
            "use_segment": None, 
            "use_col_a_as_num": None,
            "use_embedded_id": None,
            "missing_id_strategy": None
        }
        self.file_teams = {} 
        self.file_auto = {}

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA v1.0.9")
        print("="*60)
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        print("\n2. ¿Cómo manejar filas OCULTAS?")
        print("   [v] Automático: Procesar SOLO VISIBLES (Recomendado).")
        print("   [t] Automático: Procesar TODO.")
        print("   [i] Interactivo: Preguntar caso a caso.")
        resp_h = input("   >> Elija opción (Enter=v): ").lower().strip()
        if resp_h == 't': self.opt_hidden_strategy = 'all'
        elif resp_h == 'i': self.opt_hidden_strategy = 'interactive'
        else: self.opt_hidden_strategy = 'visible'
        print("-" * 60)

    # --- INTERACCIÓN ---
    def ask_segment_confirmation(self, text, context):
        if self.file_auto.get("segment_always_yes"): return True
        print(f"\n[ATENCIÓN] {context}")
        print(f"   Texto '{text}'. ¿Es un SEGMENTO?")
        print("   [s] Si")
        print("   [n] No")
        print("   [a] Si a todo en este archivo (Automático)")
        while True:
            resp = input("   >> Elija: ").lower().strip()
            if resp == 's': return True
            if resp == 'n': return False
            if resp == 'a':
                self.file_auto["segment_always_yes"] = True
                return True

    def ask_id_extraction(self, extracted_id, context):
        if self.decisions.get("use_embedded_id") is not None: return self.decisions["use_embedded_id"]
        print(f"\n[DECISIÓN DE ID] {context}")
        print(f"   No hay columna Número, pero encontré '{extracted_id}' en el texto/columna A.")
        print("   [s]  Usar este código.")
        print("   [a]  Usar SIEMPRE códigos detectados (Global).")
        print("   [n]  No usar.")
        while True:
            choice = input("   >> Elija: ").lower().strip()
            if choice == 's': return True
            if choice == 'a': 
                self.decisions["use_embedded_id"] = True
                return True
            if choice == 'n': return False

    def ask_missing_id_strategy(self, prev_id, context):
        if self.file_auto.get("missing_id_strat"): return self.file_auto["missing_id_strat"]
        print(f"\n[DECISIÓN ID FALTANTE] {context}")
        print(f"   No encontré un número válido.")
        print(f"   [s]   Saltar esta fila (Probablemente basura/vacía).")
        print(f"   [sa]  Saltar SIEMPRE filas sin ID en este archivo.")
        print(f"   [p]   Usar el ID anterior: '{prev_id}'.")
        print(f"   [pa]  Usar SIEMPRE el anterior en este archivo.")
        print(f"   [n]   Generar código NUEVO único (GEN_X).")
        while True:
            choice = input("   >> Elija: ").lower().strip()
            if choice == 's': return 'skip'
            if choice == 'sa': self.file_auto["missing_id_strat"] = 'skip'; return 'skip'
            if choice == 'p': return 'prev'
            if choice == 'pa': self.file_auto["missing_id_strat"] = 'prev'; return 'prev'
            if choice == 'n': return 'new'

    def ask_column_action(self, missing_cols, context):
        if self.file_auto.get("missing_col_continue"): return 'continue'
        print(f"\n[ALERTA] {context}")
        print(f"   Falta columna crítica: {missing_cols}")
        print("   [c]  Continuar (Intentar inferir)")
        print("   [ca] Continuar SIEMPRE en este archivo")
        print("   [s]  Saltar hoja")
        print("   [d]  Detener")
        while True:
            choice = input("   >> Elija: ").lower().strip()
            if choice == 'c': return 'continue'
            if choice == 'ca': self.file_auto["missing_col_continue"] = True; return 'continue'
            if choice == 's': return 'skip_sheet'
            if choice == 'd': self.export_excel(); sys.exit()

    def ask_weird_row_action(self, content, context):
        clean = str(content).strip().upper()
        if clean in self.memory_skip: return 'skip'
        if "NUEVO" in clean and self.memory_generate: return 'auto'
        if content == "[VACÍO]" and self.memory_skip_empty: return 'skip'
        if "VALOR INDICADOR" in clean or "OPERANDO" in clean: return 'skip'
        print(f"\n[FILA RARA] {context}")
        print(f"   Contenido NÚMERO: '{content}'")
        print("   [c] Procesar  [s] Saltar  [x] Saltar Siempre  [d] Detener")
        while True:
            choice = input("   >> Elija: ").lower().strip()
            if choice == 'c': return 'auto'
            if choice == 's': return 'skip'
            if choice == 'x':
                if content == "[VACÍO]": self.memory_skip_empty = True
                else: self.memory_skip.add(clean)
                return 'skip'
            if choice == 'd': self.export_excel(); sys.exit()

    def get_hidden_rows(self, file_path, sheet_name):
        try:
            wb = load_workbook(file_path, read_only=False, data_only=True)
            if sheet_name not in wb.sheetnames: return set()
            ws = wb[sheet_name]
            hidden = set()
            for row_idx, row_dim in ws.row_dimensions.items():
                if row_dim.hidden: hidden.add(row_idx - 1)
            wb.close()
            return hidden
        except: return set()

    # --- UTILIDADES ---
    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "": return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix, core, f_type = "", f_clean, "CUOCIENTE"
        if match:
            suffix = match.group(1); core = f_clean[:match.start()].strip(); f_type = "PORCENTAJE"
        if core.startswith("(") and core.endswith(")"): core = core[1:-1].strip()
        return core + suffix, f_type

    def parse_indicator_text(self, text):
        if pd.isna(text) or str(text).strip() == "": return "", "No identificado", "No identificado"
        text_str = str(text).strip(); lines = text_str.split('\n'); first_line = lines[0].strip()
        match = re.search(r'^[\d\)\.\-\s]*([^/]+)/(.+)', first_line)
        clean_text, dim, amb = text_str, "No identificado", "No identificado"
        if match:
            dim = match.group(1).strip(); amb = match.group(2).strip()
            clean_text = "\n".join(lines[1:]).strip() if len(lines) > 1 else ""
        if clean_text.startswith("(") and clean_text.endswith(")"): clean_text = clean_text[1:-1].strip()
        return clean_text, dim, amb

    def transform_percentage(self, val):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or val == "No aplica": return val
        if "Valor" in str(val) or "Operando" in str(val): return "" 
        try:
            clean = str(val).replace(",", ".")
            num = float(clean)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    def determine_team(self, df, file_name, sheet_name):
        if file_name in self.file_teams: return self.file_teams[file_name]
        candidate = None
        limit_row = 20; limit_col = min(15, len(df.columns))
        for r in range(min(limit_row, len(df))):
            for c in range(limit_col):
                val = str(df.iloc[r, c]).strip().upper()
                if "CENTRO DE RESPONSABILIDAD" in val or "DIRECCIÓN REGIONAL" in val:
                    parts = val.split(":")
                    if len(parts) > 1 and parts[1].strip(): candidate = parts[1].strip()
                    elif c + 1 < len(df.columns): 
                        next_val = str(df.iloc[r, c+1]).strip()
                        if next_val and next_val.upper() != "NAN": candidate = next_val
                    if not candidate: candidate = val.replace("CENTRO DE RESPONSABILIDAD", "").replace(":", "").strip()
                    break
            if candidate: break
        if not candidate:
            targets = [(8, 1), (7, 1), (11, 1), (8, 0), (7, 0)] 
            for r_idx, c_idx in targets:
                if r_idx < len(df) and c_idx < len(df.columns):
                    val = str(df.iloc[r_idx, c_idx]).strip()
                    if len(val) > 3 and val.upper() != "NAN" and "INDICADOR" not in val.upper():
                        candidate = val; break
        if not candidate or candidate.upper() in ["NO APLICA", "NAN"]:
            print(f"\n[DECISIÓN MANUAL] Equipo no detectado en: {file_name}")
            print(f"   Hoja: {sheet_name}"); print("   [m] Manual  [n] No aplica")
            while True:
                c = input("   >> ").lower().strip()
                if c == 'm': candidate = input("   >> Nombre: ").strip(); break
                if c == 'n': candidate = "No aplica"; break
        self.file_teams[file_name] = candidate
        return candidate

    def get_real_data_row_index(self, df, start_idx):
        if start_idx >= len(df): return start_idx
        row_str = "".join([str(x).upper() for x in df.iloc[start_idx].values])
        if "VALOR INDICADOR" in row_str or "OPERANDO" in row_str: return start_idx + 1
        return start_idx

    # --- CORRECCIÓN EN OPERANDOS ---
    def find_operand_offsets(self, df, start_row, ignored_rows):
        off1 = None; off2 = None
        for offset in range(1, 12):
            t = start_row + offset
            if t >= len(df): continue 
            
            # Verificar si es una fila de indicador principal para ignorarla
            if t in ignored_rows: continue 

            row_str = "".join([str(x).upper() for x in df.iloc[t].values if pd.notna(x)])
            
            # Búsqueda laxa de etiquetas (espacios opcionales)
            # Regex: OPERANDO \s* 1 \s* =
            if re.search(r'OPERANDO\s*1\s*=', row_str): off1 = offset + 1
            if re.search(r'OPERANDO\s*2\s*=', row_str): off2 = offset + 1
            
            if off1 and off2: break
        
        if not off1 and not off2: # Fallback Legacy
            offsets = []
            for offset in range(1, 10):
                t = start_row + offset
                if t >= len(df): continue
                if t in ignored_rows: continue # CRÍTICO: No tomar la fila del indicador como operando

                row_s = "".join([str(x).upper() for x in df.iloc[t].values])
                # Buscar filas con números que NO sean etiquetas de valor
                if any(c.isdigit() for c in row_s) and "VALOR" not in row_s: 
                    offsets.append(offset)
                if len(offsets) == 2: break
            return (offsets[0] if len(offsets)>0 else None, offsets[1] if len(offsets)>1 else None)
        return off1, off2

    def process_folder(self):
        files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in files if "IPS_CONSOLIDADO" not in f and not os.path.basename(f).startswith("~$")]
        if not valid_files: print("[ERROR] No hay archivos."); sys.exit()
        self.configure()
        
        for idx_file, file_path in enumerate(valid_files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            self.file_auto = {} 
            print(f"\n>>> PROCESANDO ({idx_file + 1}/{len(valid_files)}): {file_name}")
            
            try: xls = pd.ExcelFile(file_path)
            except Exception as e: print(f" [ERROR] Corrupto: {e}"); continue

            for sheet in xls.sheet_names:
                hidden_rows = self.get_hidden_rows(file_path, sheet)
                ignored_rows = set()
                if hidden_rows:
                    if self.opt_hidden_strategy == 'interactive':
                        action = self.ask_hidden_interactive(len(hidden_rows), sheet)
                        if action == 'skip': continue
                        if action == 'visible': ignored_rows = hidden_rows
                    elif self.opt_hidden_strategy == 'visible': ignored_rows = hidden_rows

                try: df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except: continue
                
                global_center = self.determine_team(df, file_name, sheet)
                if global_center is None: continue

                header_indices = []
                for idx, row in df.iterrows():
                    if idx in ignored_rows: continue
                    vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    if "INDICADOR" in vals and ("FORMULA" in vals or "FÓRMULA" in vals): header_indices.append(idx)
                    elif "NÚMERO" in vals and "INDICADOR" in vals: 
                        if idx not in header_indices: header_indices.append(idx)
                
                sheet_rows = []
                last_valid_id = "N/A"

                for loop_idx, h_idx in enumerate(header_indices):
                    end_idx = header_indices[loop_idx + 1] if loop_idx + 1 < len(header_indices) else len(df)
                    current_segment = "GENERAL"
                    if h_idx > 0:
                        prev = df.iloc[h_idx - 1].dropna()
                        if len(prev) == 1:
                            cand = str(prev.iloc[0]).strip()
                            if len(cand) < 60:
                                ctx = f"[{file_name}] > [{sheet}]"
                                if cand.upper() in self.known_segments: current_segment = cand
                                else:
                                    if self.ask_segment_confirmation(cand, ctx): 
                                        current_segment = cand
                                        self.known_segments.add(cand.upper())

                    headers = [str(h).strip() for h in df.iloc[h_idx]]
                    def fc(names):
                        for i, h in enumerate(headers):
                            if any(n.lower() in str(h).lower() for n in names): return i
                        return None

                    col_num = fc(["NÚMERO", "NUMERO", "N°"])
                    col_ind = fc(["INDICADOR"])
                    if col_num is None:
                        ctx = f"[{file_name}] > [{sheet}]"
                        action = self.ask_column_action("NÚMERO", ctx)
                        if action == 'skip_sheet': break
                        if action == 'continue': pass 

                    c_map = {
                        "num": col_num, "prod": fc(["PRODUCTO"]), "ind": col_ind, "form": fc(["FORMULA"]),
                        "uni": fc(["UNIDAD"]), "resp": fc(["RESPONSABLE"]), "gest": fc(["GESTOR"]),
                        "sup": fc(["SUPERVISORES"]), "meta": fc(["Meta 2025", "Meta 2026", "Meta"]),
                        "pond": fc(["Ponderador"]), "op_desc": fc(["Operandos"]), 
                        "op_est": fc(["Operandos Estimados", "Estimados Meta"]),
                        "proy": fc(["Cumplimiento Proyectado", "Proyectado"]), "cump_meta": fc(["% Cumplimiento"]),
                        "medios": fc(["Medios"]), "control": fc(["Control"]), "inst": fc(["Instrumentos"])
                    }
                    months_list = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                                   "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                                   "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                    m_map = {m: fc([m]) for m in months_list}
                    c_map.update(m_map)

                    for i in range(h_idx + 1, end_idx):
                        ctx = f"[{file_name}] > [{sheet}] > Fila {i+1}"
                        # 1. SEGMENTO
                        if c_map["num"] is None or c_map["num"] == 0:
                            possible_seg = str(df.iloc[i, 0]).strip()
                            if len(possible_seg) > 2 and len(possible_seg) < 30 and not any(c.isdigit() for c in possible_seg) and "INDICADOR" not in possible_seg.upper() and possible_seg != "" and possible_seg.upper() != "NAN":
                                if possible_seg.upper() in self.known_segments:
                                    current_segment = possible_seg; continue
                                if self.ask_segment_confirmation(possible_seg, ctx):
                                    current_segment = possible_seg
                                    self.known_segments.add(possible_seg.upper()); continue

                        raw_num = str(df.iloc[i, c_map["num"]]).strip() if c_map["num"] is not None else ""
                        
                        # 2. EXTRACCIÓN INTELIGENTE
                        if not raw_num or raw_num.lower() == "nan":
                            if c_map["num"] is None:
                                col0_val = str(df.iloc[i, 0]).strip()
                                if re.match(r'^\d+(\.\d+)+', col0_val): raw_num = col0_val
                            
                            if not raw_num or raw_num.lower() == "nan":
                                ind_content = str(df.iloc[i, c_map["ind"]]).strip() if c_map["ind"] is not None else ""
                                if len(ind_content) > 5:
                                    match = re.search(r'(?:^|[\s\n])(\d+\.\d+\.\d+(?:\.\d+)*)', ind_content)
                                    if match:
                                        found_id = match.group(1)
                                        if self.ask_id_extraction(found_id, ctx): raw_num = found_id
                            
                            if not raw_num or raw_num.lower() == "nan":
                                ind_content = str(df.iloc[i, c_map["ind"]]).strip() if c_map["ind"] is not None else ""
                                if len(ind_content) > 5:
                                    strat = self.ask_missing_id_strategy(last_valid_id, ctx)
                                    if strat == 'skip': continue 
                                    if strat == 'prev': raw_num = last_valid_id
                                    else: 
                                        raw_num = f"GEN_{self.new_indicator_count}"; self.new_indicator_count += 1
                                else: continue

                        if not raw_num or raw_num.lower() == "nan": continue
                        if raw_num.upper() in self.blacklist_auto or "VALOR" in raw_num.upper(): continue

                        if not (re.match(r'^\d', raw_num) or "NUEVO" in raw_num.upper() or "GEN" in raw_num.upper() or "S/N" in raw_num.upper()):
                            action = self.ask_weird_row_action(raw_num, ctx)
                            if action == 'skip': continue
                            if action == 'auto': 
                                raw_num = f"GEN_{self.new_indicator_count}"; self.new_indicator_count += 1

                        last_valid_id = raw_num

                        check_c = c_map["meta"] if c_map["meta"] else (m_map["Oct."] if m_map["Oct."] else None)
                        idx_ind_data = self.get_real_data_row_index(df, i)
                        
                        # --- FIX: Evitar que el indicador sea detectado como operando ---
                        rows_to_ignore_for_op = set(ignored_rows)
                        rows_to_ignore_for_op.add(idx_ind_data) # <--- CLAVE
                        
                        off1, off2 = self.find_operand_offsets(df, i, rows_to_ignore_for_op)
                        idx_op1 = (i + off1) if off1 else idx_ind_data 
                        idx_op2 = (i + off2) if off2 else idx_ind_data

                        def gd(r_idx, c_idx):
                            if c_idx is None or r_idx >= len(df): return ""
                            val = df.iloc[r_idx, c_idx]
                            return val if pd.notna(val) else ""

                        row_data = {
                            "ARCHIVO": file_name, "HOJA": sheet, "EQUIPO": global_center, "SEGMENTO": current_segment,
                            "TIPO INDICADOR": "CDC" if "CDC" in sheet.upper() else "PMG",
                            "NÚMERO": raw_num,
                            "PRODUCTO O PROCESO ESPECÍFICO": gd(i, c_map["prod"]),
                            "INDICADOR": self.parse_indicator_text(gd(i, c_map["ind"]))[0],
                            "DIMENSIÓN": self.parse_indicator_text(gd(i, c_map["ind"]))[1],
                            "ÁMBITO": self.parse_indicator_text(gd(i, c_map["ind"]))[2],
                            "FORMULA": self.analyze_formula(gd(i, c_map["form"]))[0],
                            "TIPO FORMULA": self.analyze_formula(gd(i, c_map["form"]))[1],
                            "UNIDAD": gd(i, c_map["uni"]), "RESPONSABLE": gd(i, c_map["resp"]),
                            "GESTOR": gd(i, c_map["gest"]), "SUPERVISORES": gd(i, c_map["sup"]),
                            "Meta 2025": self.transform_percentage(gd(idx_ind_data, c_map["meta"])),
                            "Ponderador": self.transform_percentage(gd(idx_ind_data, c_map["pond"])),
                            "Descripción Operando 1": gd(idx_op1, c_map["op_desc"]),
                            "Descripción Operando 2": gd(idx_op2, c_map["op_desc"]),
                            "Meta Operando 1 (Valor)": gd(idx_op1, c_map["op_est"]),
                            "Meta Operando 2 (Valor)": gd(idx_op2, c_map["op_est"]),
                            "Cumplimiento Proyectado 2026 Op 1": gd(idx_op1, c_map["proy"]),
                            "Cumplimiento Proyectado 2026 Op 2": gd(idx_op2, c_map["proy"]),
                            "% Cumplimiento de Meta": self.transform_percentage(gd(idx_op1, c_map["cump_meta"])),
                            "Medios de Verificación": gd(idx_ind_data, c_map["medios"]),
                            "Control de Cambios": gd(idx_ind_data, c_map["control"]),
                            "Instrumentos de Gestión Asociados": gd(idx_ind_data, c_map["inst"])
                        }
                        for m_key, m_col in m_map.items():
                            row_data[f"{m_key} Op 1"] = gd(idx_op1, m_col)
                            row_data[f"{m_key} Op 2"] = gd(idx_op2, m_col)
                        self.flat_data.append(row_data)
                        sheet_rows.append(row_data)

                        meses_v = [(10, "Oct."), (11, "Nov."), (12, "Dic.")]
                        for m_num, m_txt in meses_v:
                            col_m = m_map.get(m_txt)
                            v1_raw = gd(idx_op1, col_m) if col_m is not None else ""
                            if "VALOR" in str(v1_raw).upper() or "OPERANDO" in str(v1_raw).upper(): v1_raw = ""
                            self.variable_data.append({
                                "ANO": 2025, "MES": m_num, "VARIABLE_COD": f"{raw_num}_A",
                                "CENTRO_RESP_COD": global_center, "COD_REGION": 0, "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "", "VALOR_TOTAL": v1_raw, "ARCHIVO": file_name, "HOJA": sheet
                            })
                            v2_raw = gd(idx_op2, col_m) if col_m is not None else ""
                            if "VALOR" in str(v2_raw).upper() or "OPERANDO" in str(v2_raw).upper(): v2_raw = ""
                            self.variable_data.append({
                                "ANO": 2025, "MES": m_num, "VARIABLE_COD": f"{raw_num}_B",
                                "CENTRO_RESP_COD": global_center, "COD_REGION": 0, "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "", "VALOR_TOTAL": v2_raw, "ARCHIVO": file_name, "HOJA": sheet
                            })

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"   -> {len(sheet_rows)} indicadores extraídos.")

        self.export_excel()

    def get_ordered_headers(self):
        base = [
            "ARCHIVO", "HOJA", "EQUIPO", "SEGMENTO", "TIPO INDICADOR", "NÚMERO",
            "PRODUCTO O PROCESO ESPECÍFICO", "INDICADOR", "DIMENSIÓN", "ÁMBITO",
            "FORMULA", "TIPO FORMULA", "UNIDAD", "RESPONSABLE", "GESTOR", "SUPERVISORES",
            "Meta 2025", "Ponderador",
            "Descripción Operando 1", "Descripción Operando 2",
            "Meta Operando 1 (Valor)", "Meta Operando 2 (Valor)"
        ]
        months_order = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                        "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                        "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
        for m in months_order:
            base.append(f"{m} Op 1"); base.append(f"{m} Op 2")
        base.append("Dic. Op 1"); base.append("Dic. Op 2")
        seen = set(); final_base = []
        for x in base:
            if x not in seen: final_base.append(x); seen.add(x)
        final_base.extend([
            "Cumplimiento Proyectado 2026 Op 1", "Cumplimiento Proyectado 2026 Op 2",
            "% Cumplimiento de Meta", "Medios de Verificación",
            "Control de Cambios", "Instrumentos de Gestión Asociados"
        ])
        return final_base

    def export_excel(self):
        print(f"\nGenerando Excel Maestro...")
        wb = Workbook()
        ws = wb.active; ws.title = "Carga Bruta"
        keys = self.get_ordered_headers()
        ws.append(keys)
        fill = PatternFill("solid", fgColor="002060"); font = Font(color="FFFFFF", bold=True)
        for c in ws[1]: c.fill = fill; c.font = font
        for d in self.flat_data: ws.append([d.get(k, "") for k in keys])

        ws_s = wb.create_sheet("Planilla Estilizada")
        styles = {
            'file': PatternFill("solid", fgColor="000000"), 'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"), 'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True), 'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }
        r_idx = 1
        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            c = ws_s.cell(r_idx, 1, f"ARCHIVO: {fname}"); c.fill=styles['file']; c.font=styles['w_font']; ws_s.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(keys)); r_idx+=1
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_s.cell(r_idx, 1, f"HOJA: {sname}"); c.fill=styles['sheet']; c.font=styles['w_font']; ws_s.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(keys)); r_idx+=1
                for i, k in enumerate(keys, 1):
                    c = ws_s.cell(r_idx, i, k); c.fill=styles['head']; c.font=styles['b_font']; c.border=styles['border']
                r_idx+=1
                for row in rows:
                    for i, k in enumerate(keys, 1):
                        c = ws_s.cell(r_idx, i, row.get(k, "")); c.border=styles['border']; c.alignment=Alignment(wrapText=True, vertical='top')
                    r_idx+=1
                r_idx+=1

        ws_v = wb.create_sheet("DATOS_VARIABLE")
        h_vars = ["ANO", "MES", "VARIABLE_COD", "CENTRO_RESP_COD", "COD_REGION", "VALOR_M", "VALOR_F", "VALOR_S", "VALOR_J", "VALOR_TOTAL", "ARCHIVO", "HOJA"]
        ws_v.append(h_vars)
        for c in ws_v[1]: c.fill = fill; c.font = font
        for item in self.variable_data: ws_v.append([item.get(k, "") for k in h_vars])

        try: wb.save(self.output_file); print(f"[EXITO] Guardado en: {self.output_file}")
        except Exception as e: print(f"[ERROR] {e}")

if __name__ == "__main__":
    path = input("Ruta: ").strip() or os.getcwd()
    IPSParserHybridV109(path).process_folder()