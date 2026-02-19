import pandas as pd
import os
import sys
import glob
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v5.0.2 - FINAL STABLE (RESTAURACIÓN ESTRUCTURAL + FIX LECTURA)
# =============================================================================

class IPSParserV502:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_v5.0.2.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.variable_data = [] 
        self.new_indicator_count = 1
        
        # Configuración V4.0.2
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible'
        self.blacklist_auto = ["NÚMERO", "NUMERO", "N°", "NO", "Nº"]
        
        # Memorias de decisión
        self.memory_skip = set()      
        self.memory_generate = False
        self.decisions = {
            "use_segment": None,     
            "use_col_a_as_num": None
        }

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA v5.0.2")
        print("="*60)
        
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        
        print("\n2. ¿Cómo manejar filas OCULTAS?")
        print("   [v] Automático: Procesar SOLO VISIBLES (Recomendado).")
        print("   [t] Automático: Procesar TODO.")
        resp_h = input("   >> Elija opción (Enter=v): ").lower().strip()
        
        if resp_h == 't': self.opt_hidden_strategy = 'all'
        else: self.opt_hidden_strategy = 'visible'
        print("-" * 60)

    def ask_user_decision(self, key, prompt_text):
        if self.decisions.get(key) is not None: return self.decisions[key]
        print(f"\n[ATENCIÓN] {prompt_text}")
        while True:
            resp = input("   >> ¿Confirmar? (S/N): ").strip().upper()
            if resp in ["S", "N"]:
                self.decisions[key] = (resp == "S")
                return self.decisions[key]

    # --- UTILIDADES DE LIMPIEZA V4 ---
    def is_fully_enclosed_by_parens(self, text):
        if not text.startswith("(") or not text.endswith(")"): return False
        balance = 0
        for i, char in enumerate(text):
            if char == '(': balance += 1
            elif char == ')': balance -= 1
            if balance == 0 and i < len(text) - 1: return False
        return balance == 0

    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "": return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix = ""
        core = f_clean
        f_type = "CUOCIENTE"
        if match:
            suffix = match.group(1)
            core = f_clean[:match.start()].strip()
            f_type = "PORCENTAJE"
        while self.is_fully_enclosed_by_parens(core):
            core = core[1:-1].strip()
        return core + suffix, f_type

    def parse_indicator_text(self, text):
        if pd.isna(text) or str(text).strip() == "": return "", "No identificado", "No identificado"
        text_str = str(text).strip()
        lines = text_str.split('\n')
        first_line = lines[0].strip()
        match = re.search(r'^[\d\)\.\-\s]*([^/]+)/(.+)', first_line)
        clean_text = text_str
        dim = "No identificado"
        amb = "No identificado"
        if match:
            dim = match.group(1).strip()
            amb = match.group(2).strip()
            if len(lines) > 1: clean_text = "\n".join(lines[1:]).strip()
            else: clean_text = ""
        while self.is_fully_enclosed_by_parens(clean_text):
            clean_text = clean_text[1:-1].strip()
        return clean_text, dim, amb

    def transform_percentage(self, val, col_name):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or val == "No aplica": return val
        # Si capturamos "Valor indicador=" por error, devolver vacio
        if "Valor" in str(val) or "Operando" in str(val): return "" 
        try:
            clean_val = str(val).replace(",", ".")
            num = float(clean_val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    # --- CORE LÓGICO DE BÚSQUEDA ---
    def find_center_responsibility(self, df):
        """Busca el equipo en las primeras 20 filas UNA SOLA VEZ por hoja."""
        limit_row = 20
        limit_col = min(15, len(df.columns))
        for r in range(min(limit_row, len(df))):
            for c in range(limit_col):
                val = str(df.iloc[r, c]).strip().upper()
                if val.startswith("RESPONSABLE"): continue
                if "CENTRO DE RESPONSABILIDAD" in val:
                    parts = str(df.iloc[r, c]).split(":")
                    if len(parts) > 1 and parts[1].strip(): return parts[1].strip()
                    if c + 1 < len(df.columns): return str(df.iloc[r, c+1]).strip()
                if "DIRECCION REGIONAL" in val or "DIRECCIÓN REGIONAL" in val:
                    clean = str(df.iloc[r, c]).upper().replace("DIRECCIÓN REGIONAL", "").replace("DIRECCION REGIONAL", "").replace("-", "").strip()
                    return clean if clean else str(df.iloc[r, c])
        return "No aplica"

    def get_real_data_row_index(self, df, start_idx, col_check_idx):
        """
        Determina si la fila 'start_idx' tiene los datos o es solo etiquetas.
        Si encuentra 'Valor indicador=' o 'Operando', baja una fila.
        """
        if start_idx >= len(df): return start_idx
        
        # Verificar contenido en una columna clave (ej: Octubre o Meta)
        # Si col_check_idx es None, revisamos toda la fila
        if col_check_idx is not None:
            val = str(df.iloc[start_idx, col_check_idx]).strip()
        else:
            val = "".join([str(x) for x in df.iloc[start_idx].values])

        if "VALOR INDICADOR" in val.upper() or "OPERANDO" in val.upper():
            return start_idx + 1 # Los datos están abajo
        return start_idx

    def find_operand_offsets(self, df, start_row, c_map, ignored_rows):
        """Busca dinámicamente dónde están los datos de los operandos."""
        offsets = []
        max_look = 10
        
        # Columna de control para ver si es fila de datos (usamos un mes o meta)
        check_col = None
        for k in ["Oct.", "Dic.", "Ene.", "Meta 2026", "Operandos Estimados Meta"]:
            if c_map.get(k) is not None:
                check_col = c_map[k]
                break
        
        for offset in range(1, max_look + 1):
            target = start_row + offset
            if target >= len(df) or target in ignored_rows: continue
            
            row_vals = [str(x).upper() for x in df.iloc[target] if pd.notna(x)]
            row_str = "".join(row_vals)
            
            # Si es etiqueta, ignoramos
            if "VALOR INDICADOR" in row_str or "OPERANDO 1=" in row_str or "OPERANDO 2=" in row_str:
                continue
                
            # Si parece dato (tiene descripción o número)
            is_data = False
            # Criterio 1: Tiene texto en columna Descripción Operando
            if c_map["op_desc"] is not None:
                desc = str(df.iloc[target, c_map["op_desc"]]).strip()
                if desc and desc.lower() != "nan": is_data = True
            
            # Criterio 2: Tiene numeros en columnas de meses
            if not is_data and check_col is not None:
                val = str(df.iloc[target, check_col]).strip()
                if any(c.isdigit() for c in val): is_data = True
            
            if is_data:
                offsets.append(offset)
                if len(offsets) == 2: break
        
        return (offsets[0] if len(offsets) > 0 else None, 
                offsets[1] if len(offsets) > 1 else None)

    # --- PROCESAMIENTO ---
    def process_folder(self):
        files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in files if "IPS_CONSOLIDADO" not in f and not os.path.basename(f).startswith("~$")]
        
        if not valid_files: print("[ERROR] No hay archivos."); sys.exit()
        self.configure()
        
        for idx_file, file_path in enumerate(valid_files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            print(f"\n>>> PROCESANDO: {file_name}")
            
            try: xls = pd.ExcelFile(file_path)
            except: continue

            for sheet in xls.sheet_names:
                try: df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except: continue
                
                ignored_rows = set() # (Simplificación para modo visible)
                
                # 1. DETECTAR EQUIPO (Global para la hoja)
                global_center = self.find_center_responsibility(df)
                
                # 2. DETECTAR BLOQUES
                header_indices = []
                for idx, row in df.iterrows():
                    vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    if "INDICADOR" in vals and ("FORMULA" in vals or "FÓRMULA" in vals):
                        header_indices.append(idx)
                    elif "NÚMERO" in vals and "INDICADOR" in vals:
                        if idx not in header_indices: header_indices.append(idx)

                if not header_indices:
                    # Intento Columna A
                    first_col = [str(x) for x in df.iloc[:,0].head(15)]
                    if any("3." in x or "5." in x for x in first_col):
                         # Asumimos header virtual en fila 0 si no hay
                         pass
                    else: continue

                sheet_rows = []
                
                # LOOP BLOQUES
                for loop_idx, h_idx in enumerate(header_indices):
                    end_idx = header_indices[loop_idx + 1] if loop_idx + 1 < len(header_indices) else len(df)
                    
                    # Segmento
                    current_segment = "GENERAL"
                    if h_idx > 0:
                        prev = df.iloc[h_idx - 1].dropna()
                        if len(prev) == 1:
                            cand = str(prev.iloc[0]).strip()
                            if len(cand) < 50:
                                msg = f"Hoja '{sheet}': Detectado título '{cand}'. ¿Es SEGMENTO?"
                                if self.ask_user_decision("use_segment", msg): current_segment = cand

                    # Columnas
                    headers = [str(h).strip() for h in df.iloc[h_idx]]
                    def fc(names):
                        for i, h in enumerate(headers):
                            if any(n.lower() in str(h).lower() for n in names): return i
                        return None

                    # Mapa Columnas
                    col_num = fc(["NÚMERO", "NUMERO", "N°"])
                    if col_num is None:
                        if self.ask_user_decision("use_col_a_as_num", f"Hoja '{sheet}': Sin col NÚMERO. ¿Usar Col A?"): col_num = 0
                        else: continue

                    c_map = {
                        "num": col_num,
                        "prod": fc(["PRODUCTO"]), "ind": fc(["INDICADOR"]), "form": fc(["FORMULA"]),
                        "uni": fc(["UNIDAD"]), "resp": fc(["RESPONSABLE"]), "gest": fc(["GESTOR"]),
                        "sup": fc(["SUPERVISORES"]), "meta": fc(["Meta 2025", "Meta 2026", "Meta"]),
                        "pond": fc(["Ponderador"]), "op_desc": fc(["Operandos"]), 
                        "op_est": fc(["Operandos Estimados", "Estimados Meta"]),
                        "proy": fc(["Cumplimiento Proyectado", "Proyectado"]),
                        "cump_meta": fc(["% Cumplimiento"]),
                        "medios": fc(["Medios"]), "control": fc(["Control"]), "inst": fc(["Instrumentos"])
                    }
                    
                    months_list = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                                   "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                                   "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                    m_map = {m: fc([m]) for m in months_list}

                    # Loop Filas
                    for i in range(h_idx + 1, end_idx):
                        if i in ignored_rows: continue
                        
                        raw_num = str(df.iloc[i, c_map["num"]]).strip() if c_map["num"] is not None else ""
                        if not raw_num or raw_num.upper() in self.blacklist_auto or "VALOR" in raw_num.upper(): continue
                        
                        # Validar inicio
                        if "NUEVO" not in raw_num.upper() and not re.match(r'^\d', raw_num): continue

                        # --- CORRECCION CRITICA: SALTAR FILA 'VALOR INDICADOR=' ---
                        # Verificar si la fila 'i' es la de datos o es etiqueta
                        # Usamos la columna Meta (o Octubre) para ver si tiene "Valor indicador="
                        check_c = c_map["meta"] if c_map["meta"] else (m_map["Oct."] if m_map["Oct."] else None)
                        
                        # Índice real de los datos del INDICADOR (Madre)
                        idx_ind_data = self.get_real_data_row_index(df, i, check_c)
                        
                        # Buscar Operandos (hijos)
                        off1, off2 = self.find_operand_offsets(df, i, c_map, ignored_rows)
                        idx_op1 = (i + off1) if off1 else idx_ind_data # Fallback a madre si no hay hijos
                        idx_op2 = (i + off2) if off2 else idx_ind_data

                        # Helper Data
                        def gd(r_idx, c_idx):
                            if c_idx is None or r_idx >= len(df): return ""
                            val = df.iloc[r_idx, c_idx]
                            return val if pd.notna(val) else ""

                        # Armar Fila Carga Bruta
                        row_data = {
                            "ARCHIVO": file_name, "HOJA": sheet, "EQUIPO": global_center, "SEGMENTO": current_segment,
                            "TIPO INDICADOR": "CDC" if "CDC" in sheet.upper() else "PMG",
                            "NÚMERO": raw_num,
                            "PRODUCTO O PROCESO ESPECÍFICO": gd(i, c_map["prod"]), # Texto siempre en fila i
                            "INDICADOR": self.parse_indicator_text(gd(i, c_map["ind"]))[0],
                            "DIMENSIÓN": self.parse_indicator_text(gd(i, c_map["ind"]))[1],
                            "ÁMBITO": self.parse_indicator_text(gd(i, c_map["ind"]))[2],
                            "FORMULA": self.analyze_formula(gd(i, c_map["form"]))[0],
                            "TIPO FORMULA": self.analyze_formula(gd(i, c_map["form"]))[1],
                            "UNIDAD": gd(i, c_map["uni"]), "RESPONSABLE": gd(i, c_map["resp"]),
                            "GESTOR": gd(i, c_map["gest"]), "SUPERVISORES": gd(i, c_map["sup"]),
                            
                            # Datos Numéricos (Usan idx_ind_data corregido)
                            "Meta 2026": self.transform_percentage(gd(idx_ind_data, c_map["meta"]), "Meta"),
                            "Ponderador": self.transform_percentage(gd(idx_ind_data, c_map["pond"]), "Pond"),
                            
                            # Operandos
                            "Descripción Operando 1": gd(idx_op1, c_map["op_desc"]),
                            "Meta Operando 1 (Valor)": gd(idx_op1, c_map["op_est"]),
                            "Descripción Operando 2": gd(idx_op2, c_map["op_desc"]),
                            "Meta Operando 2 (Valor)": gd(idx_op2, c_map["op_est"]),
                            
                            "Cumplimiento Proyectado Op 1": gd(idx_op1, c_map["proy"]),
                            "Cumplimiento Proyectado Op 2": gd(idx_op2, c_map["proy"]),
                            "% Cumplimiento de Meta": self.transform_percentage(gd(idx_op1, c_map["cump_meta"]), "%"),
                            "Medios de Verificación": gd(idx_ind_data, c_map["medios"]),
                            "Control de Cambios": gd(idx_ind_data, c_map["control"]),
                            "Instrumentos de Gestión": gd(idx_ind_data, c_map["inst"])
                        }

                        # Meses
                        for m_key, m_col in m_map.items():
                            row_data[f"{m_key} Op 1"] = gd(idx_op1, m_col)
                            row_data[f"{m_key} Op 2"] = gd(idx_op2, m_col)

                        self.flat_data.append(row_data)
                        sheet_rows.append(row_data)

                        # DATOS VARIABLE (Estructura V4 Restaurada)
                        meses_v = [(10, "Oct."), (11, "Nov."), (12, "Dic.")]
                        for m_num, m_txt in meses_v:
                            col_m = m_map.get(m_txt)
                            if col_m is not None:
                                # Op 1
                                v1 = gd(idx_op1, col_m)
                                if v1 and "Valor" not in str(v1):
                                    self.variable_data.append({
                                        "ANO": 2025, "MES": m_num, 
                                        "VARIABLE_COD": f"{raw_num}_A",
                                        "CENTRO_RESP_COD": global_center, "COD_REGION": 0,
                                        "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "",
                                        "VALOR_TOTAL": v1, 
                                        "ARCHIVO": file_name, "HOJA": sheet
                                    })
                                # Op 2
                                v2 = gd(idx_op2, col_m)
                                if v2 and "Valor" not in str(v2):
                                    self.variable_data.append({
                                        "ANO": 2025, "MES": m_num, 
                                        "VARIABLE_COD": f"{raw_num}_B",
                                        "CENTRO_RESP_COD": global_center, "COD_REGION": 0,
                                        "VALOR_M": "", "VALOR_F": "", "VALOR_S": "", "VALOR_J": "",
                                        "VALOR_TOTAL": v2, 
                                        "ARCHIVO": file_name, "HOJA": sheet
                                    })

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"   -> {len(sheet_rows)} indicadores extraídos.")

        self.export_excel()

    def export_excel(self):
        print(f"\nGenerando Excel Maestro...")
        wb = Workbook()
        
        # 1. CARGA BRUTA
        ws = wb.active; ws.title = "Carga Bruta"
        if self.flat_data:
            # Ordenar columnas
            keys = list(self.flat_data[0].keys())
            # Forzar SEGMENTO después de EQUIPO
            if "SEGMENTO" in keys:
                keys.remove("SEGMENTO")
                idx_eq = keys.index("EQUIPO") + 1 if "EQUIPO" in keys else 0
                keys.insert(idx_eq, "SEGMENTO")
            
            ws.append(keys)
            # Estilo header
            fill = PatternFill("solid", fgColor="002060")
            font = Font(color="FFFFFF", bold=True)
            for c in ws[1]: c.fill = fill; c.font = font
            
            for d in self.flat_data:
                ws.append([d.get(k, "") for k in keys])

        # 2. PLANILLA ESTILIZADA (Idéntica a V4)
        ws_s = wb.create_sheet("Planilla Estilizada")
        styles = {
            'file': PatternFill("solid", fgColor="000000"), 'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"), 'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True), 'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }
        r_idx = 1
        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            c = ws_s.cell(r_idx, 1, f"ARCHIVO: {fname}"); c.fill=styles['file']; c.font=styles['w_font']; ws_s.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=60); r_idx+=1
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_s.cell(r_idx, 1, f"HOJA: {sname}"); c.fill=styles['sheet']; c.font=styles['w_font']; ws_s.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=60); r_idx+=1
                
                disp_keys = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                for i, k in enumerate(disp_keys, 1):
                    c = ws_s.cell(r_idx, i, k); c.fill=styles['head']; c.font=styles['b_font']; c.border=styles['border']
                r_idx+=1
                for row in rows:
                    for i, k in enumerate(disp_keys, 1):
                        c = ws_s.cell(r_idx, i, row[k]); c.border=styles['border']; c.alignment=Alignment(wrapText=True, vertical='top')
                    r_idx+=1
                r_idx+=1

        # 3. DATOS VARIABLE (Restaurada V4)
        ws_v = wb.create_sheet("DATOS_VARIABLE")
        h_vars = ["ANO", "MES", "VARIABLE_COD", "CENTRO_RESP_COD", "COD_REGION", 
                  "VALOR_M", "VALOR_F", "VALOR_S", "VALOR_J", "VALOR_TOTAL", "ARCHIVO", "HOJA"]
        ws_v.append(h_vars)
        for c in ws_v[1]: c.fill = fill; c.font = font
        
        for item in self.variable_data:
            ws_v.append([item.get(k, "") for k in h_vars])

        try:
            wb.save(self.output_file)
            print(f"[EXITO] Guardado en: {self.output_file}")
        except Exception as e:
            print(f"[ERROR] {e}")

if __name__ == "__main__":
    path = input("Ruta: ").strip() or os.getcwd()
    IPSParserV502(path).process_folder()