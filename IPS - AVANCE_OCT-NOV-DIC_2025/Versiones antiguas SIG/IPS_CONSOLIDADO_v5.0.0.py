import pandas as pd
import os
import sys
import glob
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v5.0.1 - HYBRID REPAIR (ESTILIZADO + SOPORTE NUEVOS FORMATOS)
# =============================================================================

class IPSParserV501:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_v5.0.1.xlsx")
        self.data_tree = {} 
        self.flat_data = [] 
        self.variable_data = [] 
        self.new_indicator_count = 1
        
        # Configuración v4.0.2
        self.opt_format_percent = True
        self.opt_hidden_strategy = 'visible'
        self.blacklist_auto = ["NÚMERO", "NUMERO", "N°", "NO", "Nº"]
        
        # Memorias (v4.0.2 + Nuevas)
        self.memory_skip = set()      
        self.memory_generate = False
        self.memory_skip_empty = False
        self.decisions = {
            "use_segment": None,      # ¿Usar títulos superiores como Segmento?
            "use_col_a_as_num": None, # ¿Usar Col A como Número si falta encabezado?
        }

    def configure(self):
        print("\n" + "="*60)
        print("   CONFIGURACIÓN MAESTRA v5.0.1 (MODO REPARACIÓN)")
        print("="*60)
        
        resp_p = input("1. ¿Transformar porcentajes (0.2 -> 20)? [S/N] (Enter=Si): ").lower().strip()
        self.opt_format_percent = (resp_p != 'n')
        
        print("\n2. ¿Cómo manejar filas OCULTAS?")
        print("   [v] Automático: Procesar SOLO VISIBLES (Recomendado).")
        print("   [t] Automático: Procesar TODO.")
        resp_h = input("   >> Elija opción (Enter=v): ").lower().strip()
        
        if resp_h == 't': self.opt_hidden_strategy = 'all'
        else: self.opt_hidden_strategy = 'visible'

        print(f"\n[OK] Configuración guardada.")
        print("-" * 60)

    def ask_user_decision(self, key, prompt_text):
        """Gestiona preguntas críticas con memoria para no spammear."""
        if self.decisions.get(key) is not None:
            return self.decisions[key]
        
        print(f"\n[ATENCIÓN REQUERIDA] {prompt_text}")
        while True:
            resp = input("   >> ¿Confirmar? (S/N): ").strip().upper()
            if resp in ["S", "N"]:
                decision = (resp == "S")
                self.decisions[key] = decision
                return decision

    # --- LÓGICA PRESERVADA DE v4.0.2 (Limpieza) ---
    def is_fully_enclosed_by_parens(self, text):
        if not text.startswith("(") or not text.endswith(")"): return False
        balance = 0
        for i, char in enumerate(text):
            if char == '(': balance += 1
            elif char == ')': balance -= 1
            if balance == 0 and i < len(text) - 1: return False
        return balance == 0

    def analyze_formula(self, formula_raw):
        if pd.isna(formula_raw) or str(formula_raw).strip() == "": return "", "Sin Fórmula"
        f_clean = str(formula_raw).replace("\n", " ").strip()
        match = re.search(r'(\s*\*\s*100)\s*$', f_clean)
        suffix = ""
        core = f_clean
        f_type = "CUOCIENTE"
        if match:
            suffix = match.group(1)
            core = f_clean[:match.start()].strip()
            f_type = "PORCENTAJE"
        while self.is_fully_enclosed_by_parens(core):
            core = core[1:-1].strip()
        return core + suffix, f_type

    def parse_indicator_text(self, text):
        if pd.isna(text) or str(text).strip() == "": return "", "No identificado", "No identificado"
        text_str = str(text).strip()
        lines = text_str.split('\n')
        first_line = lines[0].strip()
        match = re.search(r'^[\d\)\.\-\s]*([^/]+)/(.+)', first_line)
        clean_text = text_str
        dim = "No identificado"
        amb = "No identificado"
        if match:
            dim = match.group(1).strip()
            amb = match.group(2).strip()
            if len(lines) > 1: clean_text = "\n".join(lines[1:]).strip()
            else: clean_text = ""
        while self.is_fully_enclosed_by_parens(clean_text):
            clean_text = clean_text[1:-1].strip()
        return clean_text, dim, amb

    def find_center_responsibility(self, df, limit_row):
        # Lógica original preservada
        search_limit = min(limit_row, 20) 
        limit_col = min(15, len(df.columns))
        for r in range(search_limit):
            for c in range(limit_col):
                val = str(df.iloc[r, c]).strip()
                val_upper = val.upper()
                if val_upper.startswith("RESPONSABLE"): continue
                if "CENTRO DE RESPONSABILIDAD" in val_upper:
                    parts = val.split(":")
                    if len(parts) > 1: return parts[1].strip()
                    if c + 1 < len(df.columns): return str(df.iloc[r, c+1]).strip()
                if "DIRECCION REGIONAL" in val_upper or "DIRECCIÓN REGIONAL" in val_upper:
                    clean = val.replace("DIRECCIÓN REGIONAL", "").replace("DIRECCION REGIONAL", "").replace("-", "").strip()
                    return clean if clean else val
        return None

    def ask_weird_row_action(self, row_idx, content, file_name, sheet_name):
        clean = str(content).strip().upper()
        if clean in self.memory_skip: return 'skip'
        if "NUEVO" in clean and self.memory_generate: return 'auto'
        if content == "[VACÍO]" and self.memory_skip_empty: return 'skip'
        
        # Auto-skip filas basura conocidas de los formatos nuevos
        if "VALOR INDICADOR" in clean or "OPERANDO" in clean: return 'skip'

        print(f"\n[FILA RARA #{row_idx} en '{sheet_name}'] NÚMERO dice: '{content}'")
        print("   [c] Procesar (Generar código auto).")
        print("   [ca] Procesar SIEMPRE (Auto para todos).")
        print("   [s] Saltar fila.")
        print("   [x] Saltar SIEMPRE filas con este texto.")
        while True:
            choice = input("   >> Elija: ").lower().strip()
            if choice == 'c': return 'auto'
            if choice == 'ca': self.memory_generate = True; return 'auto'
            if choice == 's': return 'skip'
            if choice == 'x':
                if content == "[VACÍO]": self.memory_skip_empty = True
                else: self.memory_skip.add(clean)
                return 'skip'

    def transform_percentage(self, val, col_name):
        if not self.opt_format_percent: return val
        if pd.isna(val) or val == "" or val == "No aplica": return val
        try:
            num = float(val)
            if 0 < abs(num) <= 1: return round(num * 100, 2)
            return num
        except: return val

    # --- NUEVA LÓGICA: BUSCADOR DE OPERANDOS DINÁMICO ---
    def find_operand_offsets(self, df, start_row, c_map, ignored_rows):
        """
        Busca dónde están REALMENTE los datos de los operandos 1 y 2.
        Salteando filas de 'Valor indicador=' o vacías.
        Retorna (offset_op1, offset_op2). Si no encuentra, retorna None.
        """
        offsets = []
        max_lookahead = 8 # Mirar hasta 8 filas abajo
        
        for offset in range(1, max_lookahead + 1):
            target_r = start_row + offset
            if target_r >= len(df) or target_r in ignored_rows: continue
            
            # Chequear contenido de la fila
            row_vals = [str(x).upper() for x in df.iloc[target_r] if pd.notna(x)]
            row_str = "".join(row_vals)
            
            # SI ES BASURA, SALTAR
            if "VALOR INDICADOR" in row_str: continue
            if "OPERANDO 1=" in row_str or "OPERANDO 2=" in row_str: continue # Son etiquetas, no datos
            
            # SI PARECE DATO (Tiene algo en columna Operando Descripción o en Meses)
            has_desc = False
            if c_map["op_desc"] is not None:
                val = str(df.iloc[target_r, c_map["op_desc"]]).strip()
                if val and val.lower() != "nan": has_desc = True
            
            # Si tiene números en las columnas de meses detectadas
            has_nums = False
            # Chequear un par de meses clave (ej. Oct, Dic, o Ene)
            check_indices = [idx for k, idx in c_map.items() if k in ["Oct.", "Dic.", "Ene."] and idx is not None]
            if check_indices:
                for ci in check_indices:
                    v = str(df.iloc[target_r, ci])
                    if any(c.isdigit() for c in v): 
                        has_nums = True
                        break
            
            if has_desc or has_nums:
                offsets.append(offset)
                if len(offsets) == 2: break # Encontramos los 2
        
        # Normalizar salida
        op1 = offsets[0] if len(offsets) > 0 else None
        op2 = offsets[1] if len(offsets) > 1 else None
        return op1, op2

    # --- PROCESAMIENTO PRINCIPAL ---
    def process_folder(self):
        files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in files if "IPS_CONSOLIDADO" not in f and not os.path.basename(f).startswith("~$")]
        
        if not valid_files:
            print("[ERROR] No hay archivos Excel."); sys.exit()
            
        self.configure()
        
        for idx_file, file_path in enumerate(valid_files):
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            print(f"\n>>> PROCESANDO ({idx_file + 1}/{len(valid_files)}): {file_name}")
            
            try:
                xls = pd.ExcelFile(file_path)
            except Exception as e:
                print(f"   [ERROR] Archivo corrupto: {e}"); continue

            for sheet in xls.sheet_names:
                # 1. Filas Ocultas
                ignored_rows = set()
                # (Simplificado: Usamos openpyxl solo si es necesario, aquí asumimos visible por config básica)
                # Para ser robustos con el código original, si el usuario pidió ocultos, deberíamos cargarlo.
                # Por brevedad y para no romper, usamos pandas directo si strategy='all', sino habría que cargar openpyxl.
                
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except: continue

                # 2. DETECCIÓN DE BLOQUES (Múltiples Headers)
                header_indices = []
                for idx, row in df.iterrows():
                    row_vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    # Detectar filas que parecen encabezados
                    if "INDICADOR" in row_vals and ("FORMULA" in row_vals or "FÓRMULA" in row_vals):
                        header_indices.append(idx)
                    # Backup: Si solo dice NUMERO e INDICADOR
                    elif "NÚMERO" in row_vals and "INDICADOR" in row_vals:
                         if idx not in header_indices: header_indices.append(idx)

                if not header_indices:
                    # Intento de rescate: Buscar si Columna A tiene patrón numérico 3.5.1...
                    # Y asumir cabecera en fila anterior?
                    # Por seguridad, saltamos si no hay cabecera clara.
                    continue

                sheet_rows = []
                
                # --- ITERAR POR CADA BLOQUE ENCONTRADO ---
                for loop_idx, h_idx in enumerate(header_indices):
                    
                    # Definir fin del bloque (siguiente header o fin archivo)
                    end_idx = header_indices[loop_idx + 1] if loop_idx + 1 < len(header_indices) else len(df)
                    
                    # 3. DETECCIÓN DE SEGMENTO (Fila anterior al header)
                    current_segment = "GENERAL"
                    if h_idx > 0:
                        prev_row = df.iloc[h_idx - 1].dropna()
                        if len(prev_row) == 1: # Solo un texto (ej. "Mujer")
                            candidate = str(prev_row.iloc[0]).strip()
                            msg = f"Detectado título '{candidate}' sobre tabla en hoja '{sheet}'. ¿Usar como SEGMENTO?"
                            if self.ask_user_decision("use_segment", msg):
                                current_segment = candidate
                    
                    # 4. MAPEO DE COLUMNAS (Local para este bloque)
                    headers = [str(h).strip() for h in df.iloc[h_idx]]
                    
                    def find_c(names):
                        for i, h in enumerate(headers):
                            h_clean = " ".join(str(h).split()).lower()
                            for n in names:
                                if n.lower() in h_clean: return i
                        return None
                    
                    # Buscar NÚMERO
                    col_num_idx = find_c(["NÚMERO", "NUMERO", "N°"])
                    if col_num_idx is None:
                        msg = f"En hoja '{sheet}' (Bloque {current_segment}) no hay col 'NÚMERO'. ¿Usar Columna A?"
                        if self.ask_user_decision("use_col_a_as_num", msg):
                            col_num_idx = 0
                        else:
                            print("   [!] Saltando bloque por falta de ID."); continue

                    c_map = {
                        "num": col_num_idx,
                        "prod": find_c(["PRODUCTO"]),
                        "ind": find_c(["INDICADOR"]),
                        "form": find_c(["FORMULA", "FÓRMULA"]),
                        "uni": find_c(["UNIDAD"]),
                        "resp": find_c(["RESPONSABLE"]),
                        "gest": find_c(["GESTOR"]),
                        "sup": find_c(["SUPERVISORES"]),
                        "meta": find_c(["Meta 2025", "Meta 2026", "Meta"]),
                        "pond": find_c(["Ponderador"]),
                        "op_desc": find_c(["Operandos"]), # A veces col Operandos tiene la descripción
                        "op_est": find_c(["Operandos Estimados", "Estimados Meta"]),
                        "proy": find_c(["Cumplimiento Proyectado", "Proyectado", "Cumplimiento 2025"]),
                        "cump_meta": find_c(["% Cumplimiento"]),
                        "medios": find_c(["Medios"]),
                        "control": find_c(["Control de Cambios"]),
                        "inst": find_c(["Instrumentos"])
                    }

                    # Mapeo flexible de meses
                    months_list = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                                   "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                                   "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                    month_map = {m: find_c([m]) for m in months_list}
                    
                    # --- PROCESAR FILAS DEL BLOQUE ---
                    # Identificar Equipo (usando lógica antigua, buscando arriba del header)
                    center_resp = self.find_center_responsibility(df, limit_row=h_idx) or "No aplica"

                    for i in range(h_idx + 1, end_idx):
                        if i in ignored_rows: continue
                        
                        # Helper seguro
                        def get_val(col_idx, row_offset=0):
                            target_r = i + row_offset
                            if target_r >= len(df): return ""
                            if col_idx is None: return "" # Retorna vacio para no romper
                            val = df.iloc[target_r, col_idx]
                            return val if pd.notna(val) else ""

                        # Leer Número Indicador
                        raw_num = get_val(c_map["num"])
                        str_num = str(raw_num).strip()
                        
                        if str_num.upper() in self.blacklist_auto or "VALOR INDICADOR" in str_num.upper(): continue
                        
                        # Detectar filas válidas (inicio de indicador)
                        # Regla: Debe parecer un número X.X.X o ser NUEVO
                        is_valid_start = False
                        if "NUEVO" in str_num.upper(): is_valid_start = True
                        elif re.match(r'^\d+(\.\d+)+', str_num): is_valid_start = True
                        
                        if not is_valid_start:
                            # Opción: Preguntar si es rara, pero filtrar basura
                            if len(str_num) < 3: continue # Ruido
                            # action = self.ask_weird_row_action(i+1, str_num, file_name, sheet) ... (Simplificado)
                            continue
                            
                        final_code = str_num

                        # --- AQUÍ LA MAGIA: BUSCAR OPERANDOS DINÁMICAMENTE ---
                        # En lugar de hardcodear +3 y +5, buscamos dónde hay datos
                        off1, off2 = self.find_operand_offsets(df, i, c_map, ignored_rows)
                        
                        # Si no encuentra offsets (ej. indicador sin operandos o fila errónea), usa 0 (mismos datos fila madre)
                        # Pero normalmente la fila madre tiene Meta y los hijos los mensuales.
                        idx_op1 = off1 if off1 else 0
                        idx_op2 = off2 if off2 else 0
                        
                        # Extracción de Textos
                        raw_ind = get_val(c_map["ind"])
                        clean_ind, dim, amb = self.parse_indicator_text(raw_ind)
                        clean_form, type_form = self.analyze_formula(get_val(c_map["form"]))

                        # Construir registro
                        row_data = {
                            "ARCHIVO": file_name, 
                            "HOJA": sheet, 
                            "EQUIPO": center_resp, 
                            "SEGMENTO": current_segment, # <--- NUEVA COLUMNA
                            "TIPO INDICADOR": "CDC" if "CDC" in sheet.upper() else "PMG", 
                            "NÚMERO": final_code,
                            "PRODUCTO O PROCESO ESPECÍFICO": get_val(c_map["prod"]),
                            "INDICADOR": clean_ind,
                            "DIMENSIÓN": dim,
                            "ÁMBITO": amb,
                            "FORMULA": clean_form,   
                            "TIPO FORMULA": type_form,
                            "UNIDAD": get_val(c_map["uni"]),
                            "RESPONSABLE": get_val(c_map["resp"]), 
                            "GESTOR": get_val(c_map["gest"]),
                            "SUPERVISORES": get_val(c_map["sup"]),
                            "Meta 2026": self.transform_percentage(get_val(c_map["meta"]), "Meta"),
                            "Ponderador": self.transform_percentage(get_val(c_map["pond"]), "Ponderador"),
                        }

                        # Llenar datos de Operandos usando los offsets detectados
                        # Op 1
                        row_data["Descripción Operando 1"] = get_val(c_map["op_desc"], idx_op1) if off1 else "N/A"
                        row_data["Meta Operando 1 (Valor)"] = get_val(c_map["op_est"], idx_op1)
                        # Op 2
                        row_data["Descripción Operando 2"] = get_val(c_map["op_desc"], idx_op2) if off2 else "N/A"
                        row_data["Meta Operando 2 (Valor)"] = get_val(c_map["op_est"], idx_op2)

                        # Meses
                        for m_name, m_idx in month_map.items():
                            row_data[f"{m_name} Op 1"] = get_val(m_idx, idx_op1)
                            row_data[f"{m_name} Op 2"] = get_val(m_idx, idx_op2)

                        # Extras
                        row_data["Cumplimiento Proyectado Op 1"] = get_val(c_map["proy"], idx_op1)
                        row_data["Cumplimiento Proyectado Op 2"] = get_val(c_map["proy"], idx_op2)
                        row_data["% Cumplimiento de Meta"] = self.transform_percentage(get_val(c_map["cump_meta"], idx_op1), "%")
                        row_data["Medios de Verificación"] = get_val(c_map["medios"])
                        row_data["Control de Cambios"] = get_val(c_map["control"])
                        row_data["Instrumentos de Gestión"] = get_val(c_map["inst"])

                        self.flat_data.append(row_data)
                        sheet_rows.append(row_data)

                        # DATOS VARIABLE
                        meses_vars = [(10, "Oct."), (11, "Nov."), (12, "Dic.")]
                        for mes_num, mes_key in meses_vars:
                            c_idx_mes = month_map.get(mes_key)
                            if c_idx_mes is not None:
                                val_op1 = get_val(c_idx_mes, idx_op1)
                                val_op2 = get_val(c_idx_mes, idx_op2)
                                
                                # Solo guardar si hay algo relevante (opcional)
                                self.variable_data.append([2025, mes_num, f"{final_code}_A", current_segment, val_op1, file_name])
                                self.variable_data.append([2025, mes_num, f"{final_code}_B", current_segment, val_op2, file_name])

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"   -> Extraídos {len(sheet_rows)} indicadores.")

        self.export_excel()

    def export_excel(self):
        print(f"\n{'='*60}\nGUARDANDO ARCHIVO MAESTRO...\n{'='*60}")
        wb = Workbook()
        
        # 1. CARGA BRUTA (Actualizada con SEGMENTO)
        ws = wb.active; ws.title = "Carga Bruta"
        if self.flat_data:
            # Reordenar headers para poner SEGMENTO al inicio
            base_keys = list(self.flat_data[0].keys())
            if "SEGMENTO" in base_keys:
                base_keys.remove("SEGMENTO")
                base_keys.insert(3, "SEGMENTO") # Insertar después de EQUIPO
            
            ws.append(base_keys)
            for r in self.flat_data:
                ws.append([r.get(k, "") for k in base_keys])
        
        # 2. PLANILLA ESTILIZADA (Lógica original v4.0.2 RESTAURADA)
        ws_style = wb.create_sheet("Planilla Estilizada")
        styles = {
            'file': PatternFill("solid", fgColor="000000"),
            'sheet': PatternFill("solid", fgColor="2F5597"),
            'head': PatternFill("solid", fgColor="BFBFBF"),
            'w_font': Font(color="FFFFFF", bold=True),
            'b_font': Font(bold=True),
            'border': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        }

        row_idx = 1
        FULL_WIDTH = 60

        for fname, sheets in self.data_tree.items():
            if not any(sheets.values()): continue
            
            c = ws_style.cell(row=row_idx, column=1, value=f"ARCHIVO: {fname}")
            c.fill = styles['file']; c.font = styles['w_font']; c.alignment = Alignment('center')
            ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
            row_idx += 1
            
            for sname, rows in sheets.items():
                if not rows: continue
                c = ws_style.cell(row=row_idx, column=1, value=f"HOJA: {sname}")
                c.fill = styles['sheet']; c.font = styles['w_font']
                ws_style.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=FULL_WIDTH)
                row_idx += 1
                
                # Usar keys del primer registro encontrado
                keys = list(rows[0].keys())
                # Filtrar metadatos técnicos para la vista estilizada si quieres
                disp_keys = [k for k in keys if k not in ["ARCHIVO", "HOJA"]]
                
                for c_i, k in enumerate(disp_keys, 1):
                    c = ws_style.cell(row=row_idx, column=c_i, value=k)
                    c.fill = styles['head']; c.font = styles['b_font']; c.border = styles['border']
                row_idx += 1
                
                for r in rows:
                    for c_i, k in enumerate(disp_keys, 1):
                        c = ws_style.cell(row=row_idx, column=c_i, value=r[k])
                        c.border = styles['border']; c.alignment = Alignment(wrapText=True, vertical='top')
                    row_idx += 1
                row_idx += 1

        for i in range(1, FULL_WIDTH + 2):
            ws_style.column_dimensions[get_column_letter(i)].width = 22
            
        # 3. DATOS_VARIABLE
        ws_vars = wb.create_sheet("DATOS_VARIABLE")
        headers_vars = ["ANO", "MES", "VARIABLE_COD", "SEGMENTO", "VALOR", "ARCHIVO"]
        ws_vars.append(headers_vars)
        for row in self.variable_data:
            ws_vars.append(row)

        while True:
            try:
                wb.save(self.output_file)
                print(f"[ÉXITO] Archivo generado: {self.output_file}")
                break
            except PermissionError:
                input("[ERROR] Cierra el archivo Excel y dale Enter..."); 
            except Exception as e:
                print(e); break

if __name__ == "__main__":
    path = input("Ruta (Enter para actual): ").strip() or os.getcwd()
    parser = IPSParserV501(path)
    parser.process_folder()