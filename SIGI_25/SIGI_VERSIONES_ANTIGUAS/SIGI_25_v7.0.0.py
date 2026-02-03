"""
=============================================================================
 PROGRAMA: SIGI 25 (Sistema Integral de Gestión de Indicadores)
 VERSIÓN:  v7.0.0 (Massive Surgical Extraction)
 FECHA:    Febrero 2026
 
 DESCRIPCIÓN:
 Motor ETL definitivo.
 1. Escanea todos los Excel (.xlsx) de la carpeta.
 2. Identifica el responsable por el nombre del archivo (Mapa Directo).
 3. Aplica la "Lógica de Francotirador" (+3 filas, +5 filas) para extraer
    metas y operandos ocultos en la estructura visual.
 4. Consolida todo en los 3 formatos de salida oficiales.
=============================================================================
"""

import pandas as pd
import os
import glob
import re
import warnings
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Silenciar alertas de compatibilidad
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)

# =============================================================================
# 1. CONFIGURACIÓN Y MAPAS
# =============================================================================

ANO_PROCESO = "2026"

ARCHIVOS_SALIDA = {
    "F1": f"1_PLANILLA_SIG_CONSOLIDADO_{ANO_PROCESO}.xlsx",
    "F2": f"2_CARGA_BRUTA_CONSOLIDADO_{ANO_PROCESO}.xlsx",
    "F3": f"3_REPORTE_VISUAL_CONSOLIDADO_{ANO_PROCESO}.xlsx"
}

# Mapa de Asignación Directa (Nombre de Archivo -> Código IP)
MAPA_DIRECTO = {
    # REGIONES
    "ARICA": "IP25_719", "PARINACOTA": "IP25_719",
    "TARAPACA": "IP25_720",
    "ANTOFAGASTA": "IP25_721",
    "ATACAMA": "IP25_722",
    "COQUIMBO": "IP25_723",
    "VALPARAISO": "IP25_724",
    "OHIGGINS": "IP25_725", "LIBERTADOR": "IP25_725",
    "MAULE": "IP25_726",
    "BIOBIO": "IP25_727",
    "ARAUCANIA": "IP25_728",
    "LOS RIOS": "IP25_729",
    "LOS LAGOS": "IP25_730",
    "AYSEN": "IP25_731", "AISEN": "IP25_731",
    "MAGALLANES": "IP25_732",
    "METROPOLITANA": "IP25_733",
    "ÑUBLE": "IP25_748", "NUBLE": "IP25_748",

    # CENTRAL
    "BENEFICIOS": "IP25_712",
    "CLIENTES": "IP25_713",
    "INFORMATICA": "IP25_714",
    "JURIDICA": "IP25_715",
    "PLANIFICACION": "IP25_716",
    "COMUNICACIONES": "IP25_717",
    "CONTRALORIA": "IP25_718",
    "AUDITORIA": "IP25_738",
    "SIST INFORM": "IP25_739", "SISTEMAS DE INFORMACION": "IP25_739",
    "GESTION PERSONAS": "IP25_750", "DESARROLLO DE PERSONAS": "IP25_750"
}

# =============================================================================
# 2. UTILIDADES DE LIMPIEZA
# =============================================================================

def obtener_codigo_por_archivo(nombre_archivo):
    nombre_upper = nombre_archivo.upper()
    for clave, codigo in MAPA_DIRECTO.items():
        if clave in nombre_upper:
            return codigo
    return "?" # Fallback si no encuentra

def limpiar_porcentaje(val):
    if pd.isna(val) or str(val).strip() == "": return 0
    s = str(val).replace('%', '').strip()
    # Manejo de decimales con coma o punto
    s = s.replace(',', '.')
    try: return float(s)
    except: return 0

def limpiar_numero(val):
    if pd.isna(val) or str(val).strip() == "": return 0
    s = str(val).strip()
    # Eliminar puntos de miles (ej: 25.500 -> 25500)
    # Pero cuidado con decimales. Asumimos formato Chileno: 1.000,00
    if "," in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        # Si no hay coma, asumimos que los puntos son miles si hay más de 1 o son enteros grandes
        s = s.replace('.', '')
    try: return float(s)
    except: return 0

def limpiar_texto(val):
    if pd.isna(val): return ""
    txt = str(val).strip()
    return txt.replace("\n", " ").replace("\r", " ")

def detectar_encabezados(df):
    """Busca la fila donde dice 'INDICADOR' para empezar a leer desde ahí."""
    for i in range(min(30, len(df))):
        fila = [str(x).upper() for x in df.iloc[i].tolist()]
        if any("INDICADOR" in x for x in fila):
            return i
    return None

def aplicar_estilo_profesional(ruta, hoja):
    try:
        wb = load_workbook(ruta)
        if hoja not in wb.sheetnames: return
        ws = wb[hoja]
        
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(b=True, color="FFFFFF", size=10)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font; cell.border = border
            
        for row in ws.iter_rows(min_row=2):
            es_sep = str(row[0].value).startswith("---")
            for c in row:
                c.border = border
                if es_sep: c.fill = PatternFill("solid", fgColor="D9D9D9"); c.font = Font(b=True)
        
        ws.column_dimensions["B"].width = 15 # Codigo
        ws.column_dimensions["C"].width = 50 # Nombre
        
        wb.save(ruta)
    except: pass

# =============================================================================
# 3. MOTOR DE EXTRACCIÓN (CORE)
# =============================================================================

def procesar_archivo(ruta_archivo):
    nombre_archivo = os.path.basename(ruta_archivo)
    codigo_resp = obtener_codigo_por_archivo(nombre_archivo)
    
    print(f"   -> Procesando: {nombre_archivo[:40]}... (Resp: {codigo_resp})")
    
    dfs_extraidos = []
    
    try:
        xls = pd.ExcelFile(ruta_archivo)
        
        # Buscar hojas relevantes
        hojas = [h for h in xls.sheet_names if any(x in h.upper() for x in ["CDC", "PMG", "RIESGO"])]
        if not hojas: 
            print("      [SKIP] No se encontraron hojas CDC/PMG.")
            return None

        for nombre_hoja in hojas:
            # Leer hoja completa
            df = pd.read_excel(ruta_archivo, sheet_name=nombre_hoja, header=None)
            
            # 1. Encontrar el inicio (Ancla)
            idx_header = detectar_encabezados(df)
            if idx_header is None: continue
            
            # 2. Definir Índices de Columnas (Basado en la estructura "Planilla SIG")
            # Asumimos que la estructura visual se mantiene en las columnas A, B, D, E...
            IDX_NUM = 0  # A: Número
            IDX_IND = 1  # B: Indicador
            IDX_OP_DESC = 3 # D: Operandos (Texto)
            IDX_OP_EST = 4  # E: Estimados (Valores)
            
            # Meses (Oct, Nov, Dic suelen estar fijos en estas versiones light)
            # Buscamos dinámicamente las columnas de meses en la fila de encabezado
            fila_head = df.iloc[idx_header].astype(str).tolist()
            mapa_cols = {txt.upper().strip(): i for i, txt in enumerate(fila_head)}
            
            idx_oct = next((i for txt, i in mapa_cols.items() if "OCT" in txt and "ACUM" not in txt), 5)
            idx_nov = next((i for txt, i in mapa_cols.items() if "NOV" in txt and "ACUM" not in txt), 7)
            idx_dic = next((i for txt, i in mapa_cols.items() if "DIC" in txt and "ACUM" not in txt), 9)

            extracted_rows = []
            
            # 3. Iterar buscando bloques de indicadores
            # Empezamos después del encabezado
            for i in range(idx_header + 1, len(df) - 5):
                val_num = str(df.iloc[i, IDX_NUM]).strip()
                
                # DETECTOR: Si la celda A tiene un punto y longitud > 3 (ej: "3.5.1.24.")
                if pd.notna(df.iloc[i, IDX_NUM]) and "." in val_num and len(val_num) >= 3:
                    
                    try:
                        # --- EXTRACCIÓN QUIRÚRGICA ---
                        # Datos Base
                        codigo = val_num
                        nombre = limpiar_texto(df.iloc[i, IDX_IND])
                        
                        # Meta (Fila +1)
                        meta_raw = df.iloc[i+1, IDX_OP_EST]
                        meta = limpiar_porcentaje(meta_raw) if isinstance(meta_raw, str) and "%" in meta_raw else limpiar_numero(meta_raw)
                        
                        # Operando 1 (Numerador)
                        op1_desc = limpiar_texto(df.iloc[i, IDX_OP_DESC])
                        # A veces tiene parentesis inicial
                        if op1_desc.startswith("("): op1_desc = op1_desc[1:]
                        op1_val = df.iloc[i+3, IDX_OP_EST] # Valor en fila +3
                        
                        # Operando 2 (Denominador)
                        op2_desc = limpiar_texto(df.iloc[i+3, IDX_OP_DESC])
                        # A veces tiene ")*100" al final
                        op2_desc = op2_desc.split(")")[0]
                        op2_val = df.iloc[i+5, IDX_OP_EST] # Valor en fila +5
                        
                        # Construir Fila
                        fila = {
                            "ORIGEN_ARCHIVO": nombre_archivo,
                            "NÚMERO": codigo,
                            "INDICADOR": nombre,
                            "RESPONSABLE CENTRO DE RESPONSABILIDAD": nombre_archivo.replace(".xlsx", ""),
                            "CODIGO_RESPONSABLE_ASIGNADO": codigo_resp, # CLAVE PARA EL CRUCE
                            "Meta 2025 (%)": meta,
                            "Ponderador (%)": 0,
                            "Desc. Op1": op1_desc,
                            "Desc. Op2": op2_desc,
                            "Est. Meta Op1": limpiar_numero(op1_val),
                            "Est. Meta Op2": limpiar_numero(op2_val),
                            
                            # Meses Rellenos (Ene-Sep)
                            "Ene Ind (%)": "No aplica", "Ene Op1": 0, "Ene Op2": 0,
                            "Feb Ind (%)": "No aplica", "Feb Op1": 0, "Feb Op2": 0,
                            "Mar Ind (%)": "No aplica", "Mar Op1": 0, "Mar Op2": 0,
                            "Abr Ind (%)": "No aplica", "Abr Op1": 0, "Abr Op2": 0,
                            "May Ind (%)": "No aplica", "May Op1": 0, "May Op2": 0,
                            "Jun Ind (%)": "No aplica", "Jun Op1": 0, "Jun Op2": 0,
                            "Jul Ind (%)": "No aplica", "Jul Op1": 0, "Jul Op2": 0,
                            "Ago Ind (%)": "No aplica", "Ago Op1": 0, "Ago Op2": 0,
                            "Sept Ind (%)": "No aplica", "Sept Op1": 0, "Sept Op2": 0,
                            
                            # Meses Reales (Oct-Dic)
                            "Oct Ind (%)": limpiar_numero(df.iloc[i+1, idx_oct]),
                            "Oct Op1": limpiar_numero(df.iloc[i+3, idx_oct]),
                            "Oct Op2": limpiar_numero(df.iloc[i+5, idx_oct]),
                            
                            "Nov Ind (%)": limpiar_numero(df.iloc[i+1, idx_nov]),
                            "Nov Op1": limpiar_numero(df.iloc[i+3, idx_nov]),
                            "Nov Op2": limpiar_numero(df.iloc[i+5, idx_nov]),
                            
                            "Dic Ind (%)": limpiar_numero(df.iloc[i+1, idx_dic]),
                            "Dic Op1": limpiar_numero(df.iloc[i+3, idx_dic]),
                            "Dic Op2": limpiar_numero(df.iloc[i+5, idx_dic]),
                            
                            "Medios Verificación": ""
                        }
                        extracted_rows.append(fila)
                    except Exception as e:
                        print(f"      [WARN] Error leyendo indicador en fila {i}: {e}")
                        continue

            if extracted_rows:
                df_res = pd.DataFrame(extracted_rows)
                # Agregar separador visual
                sep = {c: None for c in df_res.columns}; sep['NÚMERO'] = f"--- {nombre_hoja} ---"
                dfs_extraidos.append(pd.DataFrame([sep]))
                dfs_extraidos.append(df_res)

        return pd.concat(dfs_extraidos, ignore_index=True) if dfs_extraidos else None

    except Exception as e:
        print(f"   [ERROR FATAL] No se pudo leer {nombre_archivo}: {e}")
        return None

# =============================================================================
# 4. GENERADORES DE SALIDA (F2, F3, F4, F5)
# =============================================================================

def generar_f2(df_f1):
    rows = []
    for _, r in df_f1.iterrows():
        cod = str(r["NÚMERO"]).strip()
        if cod.startswith("---") or pd.isna(r["NÚMERO"]): continue
        
        # Componente A
        rows.append({
            "cod_interno": f"{cod}_A",
            "nombre_variable": r["Desc. Op1"],
            "descripcion": r["Desc. Op1"],
            "unidad": "Número",
            "valor_obligatorio": 1,
            "APLICA_SIN_INFORMACION": 1
        })
        # Componente B
        rows.append({
            "cod_interno": f"{cod}_B",
            "nombre_variable": r["Desc. Op2"],
            "descripcion": r["Desc. Op2"],
            "unidad": "Número",
            "valor_obligatorio": 1,
            "APLICA_SIN_INFORMACION": 1
        })
    return pd.DataFrame(rows)

def generar_f3(df_f2):
    rows = []
    seen = set()
    for _, r in df_f2.iterrows():
        cod = r["cod_interno"]
        if cod in seen: continue
        seen.add(cod)
        
        # Invertir codigo para var auto: 3.5.1_A -> A_3.5.1
        parts = cod.split('_')
        suffix = parts[-1]
        base = "_".join(parts[:-1])
        var_auto = f"{suffix}_{base}"
        
        rows.append({
            "cod_variable": cod,
            "nombre_variable": r["nombre_variable"],
            "ano_mes_ini": 202501, "ano_mes_fin": 202512,
            "ENE": 1, "FEB": 1, "MAR": 1, "ABR": 1, "MAY": 1, "JUN": 1,
            "JUL": 1, "AGO": 1, "SEP": 1, "OCT": 1, "NOV": 1, "DIC": 1,
            "EMAIL_RESPONSABLE_INGRESO_DATO": "prueba@arbol-logika.com",
            "FORMULA_VAR_AUTO": "SUMA_ANUAL",
            "codigo_var_auto": var_auto
        })
    return pd.DataFrame(rows)

def generar_f4(df_f1):
    rows = []
    for _, r in df_f1.iterrows():
        cod = str(r["NÚMERO"]).strip()
        if cod.startswith("---") or pd.isna(r["NÚMERO"]): continue
        
        rows.append({
            "CODIGO": cod,
            "NOMBRE": r["INDICADOR"],
            "DESCRIPCION": r["INDICADOR"],
            "ACTIVO": 1,
            "UNIDAD": "%",
            "RANGO_MINIMO": 0, "RANGO_MAXIMO": 100,
            "APLICA_SIN_INFORMACION": 0,
            "FORMULA_COD": "PORCENTAJE",
            "TIPO_META": "TOLERANCIA",
            "FACTOR_CUMPLIMIENTO": 10, "FACTOR_NOCUMPLIMIENTO": 20,
            "IND_CDC": 1, "ANO_ASOCIADO": 2025
        })
    return pd.DataFrame(rows).drop_duplicates(subset=["CODIGO"])

def generar_f5(df_f1):
    rows = []
    for _, r in df_f1.iterrows():
        cod = str(r["NÚMERO"]).strip()
        if cod.startswith("---") or pd.isna(r["NÚMERO"]): continue
        
        cod_pond = r["CODIGO_RESPONSABLE_ASIGNADO"]
        
        rows.append({
            "INDICADOR_COD": cod,
            "NOMBRE_INDICADOR": r["INDICADOR"],
            "JER_TIPO_COD": 1,
            "EMAIL_RESPONSABLE": "prueba@arbol-logika.com",
            "COD_ANALISIS_CAUSA": "RESP_INDICADOR",
            "ANO_MES_INI": 202501, "ANO_MES_FIN": 202512,
            "ENE": 1, "FEB": 1, "MAR": 1, "ABR": 1, "MAY": 1, "JUN": 1,
            "JUL": 1, "AGO": 1, "SEP": 1, "OCT": 1, "NOV": 1, "DIC": 1,
            "TIPO_META_ANUAL": "PERIODO_ANUAL",
            "COMP_A": f"{cod}_A",
            "COMP_B": f"{cod}_B",
            "META_202512": r["Meta 2025 (%)"],
            "Ponderacion": 0,
            "COD_PONDERADO": cod_pond,
            "FORMULA_VAR_AUTO": "SUMA_ANUAL",
            "COD_VAR_AUTO": f"A_{cod_pond}"
        })
    return pd.DataFrame(rows).drop_duplicates(subset=["INDICADOR_COD"])

# =============================================================================
# 5. ORQUESTACIÓN PRINCIPAL
# =============================================================================

def ejecutar_masivo():
    # Buscar archivos Excel (ignorando los generados por el script)
    archivos = [f for f in glob.glob("*.xlsx") if not f.startswith("1_") and not f.startswith("2_") and not f.startswith("3_") and not f.startswith("~$")]
    
    print(f"\n[SIGI 25 v7.0.0] INICIANDO PROCESO MASIVO ({len(archivos)} archivos)")
    
    if not archivos: print("[ERROR] Carpeta vacía."); return

    master_list = []
    
    for idx, archivo in enumerate(archivos):
        df_ind = procesar_archivo(archivo)
        if df_ind is not None and not df_ind.empty:
            master_list.append(df_ind)
    
    if not master_list: print("\n[ERROR] No se extrajeron datos."); return

    # Consolidación
    print("\n   -> Generando Archivos Maestros...")
    df_full = pd.concat(master_list, ignore_index=True)
    
    # Generar F1
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F1"], engine='openpyxl') as w:
        df_full.to_excel(w, index=False, sheet_name="DATOS_BRUTOS")
        df_full.to_excel(w, index=False, sheet_name="DATOS_ESTILIZADOS")
    aplicar_estilo_profesional(ARCHIVOS_SALIDA["F1"], "DATOS_ESTILIZADOS")
    
    # Generar Paquetes F2-F5
    f2 = generar_f2(df_full)
    f3 = generar_f3(f2)
    f4 = generar_f4(df_full)
    f5 = generar_f5(df_full)
    
    # Guardar Carga Bruta
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F2"], engine='openpyxl') as w:
        f2.to_excel(w, index=False, sheet_name="F2_VARIABLES")
        f3.to_excel(w, index=False, sheet_name="F3_VAR_APLICADAS")
        f4.to_excel(w, index=False, sheet_name="F4_INDICADORES")
        f5.to_excel(w, index=False, sheet_name="F5_IND_APLICADOS")
        
    # Guardar Reporte Visual
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F3"], engine='openpyxl') as w:
        f2.to_excel(w, index=False, sheet_name="VISUAL_VARIABLES")
        f3.to_excel(w, index=False, sheet_name="VISUAL_VAR_APP")
        f4.to_excel(w, index=False, sheet_name="VISUAL_INDICADORES")
        f5.to_excel(w, index=False, sheet_name="VISUAL_IND_APP")
    
    for h in ["VISUAL_VARIABLES", "VISUAL_VAR_APP", "VISUAL_INDICADORES", "VISUAL_IND_APP"]:
        aplicar_estilo_profesional(ARCHIVOS_SALIDA["F3"], h)

    print("\n   ¡MISIÓN CUMPLIDA! Archivos generados correctamente.")
    print(f"   1. {ARCHIVOS_SALIDA['F1']}")
    print(f"   2. {ARCHIVOS_SALIDA['F2']}")
    print(f"   3. {ARCHIVOS_SALIDA['F3']}")

if __name__ == "__main__":
    ejecutar_masivo()