"""
=============================================================================
 PROGRAMA: SIGI 25 (Sistema Integral de Gestión de Indicadores)
 VERSIÓN:  v7.3.0 (F3 Fix - Variables Aplicadas)
 FECHA:    Febrero 2026
 
 NOVEDADES:
 - Ajuste de Fase 3 (Variables Aplicadas) con columnas Q-Y según requerimiento.
 - Lógica para completar 'cod_centro_resp_lugar_medicion' con 'DIRECCION REGIONAL [Nombre]'.
 - Ajuste de columnas vacías y valores fijos en F3.
=============================================================================
"""

import pandas as pd
import os
import glob
import re
import warnings
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Silenciar alertas
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)

# =============================================================================
# 1. CONFIGURACIÓN
# =============================================================================

ANO_PROCESO = "2026"

ARCHIVOS_SALIDA = {
    "F1": f"1_PLANILLA_SIG_CONSOLIDADO_{ANO_PROCESO}.xlsx",
    "F2": f"2_CARGA_BRUTA_CONSOLIDADO_{ANO_PROCESO}.xlsx",
    "F3": f"3_REPORTE_VISUAL_CONSOLIDADO_{ANO_PROCESO}.xlsx"
}

# MAPA DIRECTO (Nombre de Archivo -> Código)
MAPA_DIRECTO = {
    # REGIONES
    "ARICA": "IP25_719", "PARINACOTA": "IP25_719",
    "TARAPACA": "IP25_720",
    "ANTOFAGASTA": "IP25_721",
    "ATACAMA": "IP25_722",
    "COQUIMBO": "IP25_723",
    "VALPARAISO": "IP25_724",
    "OHIGGINS": "IP25_725", "LIBERTADOR": "IP25_725",
    "MAULE": "IP25_726",
    "BIOBIO": "IP25_727",
    "ARAUCANIA": "IP25_728",
    "LOS RIOS": "IP25_729",
    "LOS LAGOS": "IP25_730",
    "AYSEN": "IP25_731", "AISEN": "IP25_731",
    "MAGALLANES": "IP25_732",
    "METROPOLITANA": "IP25_733",
    "ÑUBLE": "IP25_748", "NUBLE": "IP25_748",

    # CENTRAL
    "BENEFICIOS": "IP25_712",
    "CLIENTES": "IP25_713",
    "INFORMATICA": "IP25_714",
    "JURIDICA": "IP25_715",
    "PLANIFICACION": "IP25_716",
    "COMUNICACIONES": "IP25_717",
    "CONTRALORIA": "IP25_718",
    "AUDITORIA": "IP25_738",
    "SIST INFORM": "IP25_739", "SISTEMAS DE INFORMACION": "IP25_739",
    "GESTION PERSONAS": "IP25_750", "DESARROLLO DE PERSONAS": "IP25_750"
}

# =============================================================================
# 2. UTILS
# =============================================================================

def obtener_codigo_por_archivo(nombre_archivo):
    nombre_upper = nombre_archivo.upper()
    for clave, codigo in MAPA_DIRECTO.items():
        if clave in nombre_upper: return codigo
    return "?"

def limpiar_porcentaje(val):
    if pd.isna(val) or str(val).strip() == "": return 0
    s = str(val).replace('%', '').strip().replace(',', '.')
    try: return float(s)
    except: return 0

def limpiar_numero(val):
    if pd.isna(val) or str(val).strip() == "": return 0
    s = str(val).strip()
    if "," in s: s = s.replace('.', '').replace(',', '.')
    else: s = s.replace('.', '')
    try: return float(s)
    except: return 0

def limpiar_texto(val):
    if pd.isna(val): return ""
    return str(val).strip().replace("\n", " ").replace("\r", " ")

def detectar_encabezados(df):
    for i in range(min(30, len(df))):
        fila = [str(x).upper() for x in df.iloc[i].tolist()]
        if any("INDICADOR" in x for x in fila): return i
    return None

def extraer_nombre_limpio(nombre_archivo):
    # Eliminar extensión
    base = os.path.splitext(nombre_archivo)[0]
    # Eliminar prefijos comunes
    base = base.replace("Planilla SIG - ", "").replace("Planilla SIG- ", "").replace("Planilla SIG ", "")
    base = base.replace("CDC REG ", "").replace("CDC ", "")
    # Eliminar parentesis
    base = re.sub(r"\(.*\)", "", base)
    return base.strip()

def aplicar_estilo_profesional(ruta, hoja):
    try:
        wb = load_workbook(ruta)
        if hoja not in wb.sheetnames: return
        ws = wb[hoja]
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(b=True, color="FFFFFF", size=10)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font; cell.border = border
        for row in ws.iter_rows(min_row=2):
            es_sep = str(row[0].value).startswith("---")
            for c in row:
                c.border = border
                if es_sep: c.fill = PatternFill("solid", fgColor="D9D9D9"); c.font = Font(b=True)
        ws.column_dimensions["B"].width = 15; ws.column_dimensions["C"].width = 50
        wb.save(ruta)
    except: pass

# =============================================================================
# 3. MOTOR DE EXTRACCIÓN (CORE)
# =============================================================================

def procesar_archivo(ruta_archivo):
    nombre_archivo = os.path.basename(ruta_archivo)
    codigo_resp = obtener_codigo_por_archivo(nombre_archivo)
    print(f"   -> Procesando: {nombre_archivo[:35]}... (Resp: {codigo_resp})")
    
    dfs_extraidos = []
    try:
        xls = pd.ExcelFile(ruta_archivo)
        hojas = [h for h in xls.sheet_names if any(x in h.upper() for x in ["CDC", "PMG", "RIESGO"])]
        if not hojas: return None

        for nombre_hoja in hojas:
            df = pd.read_excel(ruta_archivo, sheet_name=nombre_hoja, header=None)
            idx_header = detectar_encabezados(df)
            if idx_header is None: continue
            
            # Mapeo Dinámico
            fila_head = df.iloc[idx_header].astype(str).tolist()
            mapa_cols = {txt.upper().strip(): i for i, txt in enumerate(fila_head)}
            
            IDX_NUM = next((i for txt, i in mapa_cols.items() if "NÚMERO" in txt or "NUMERO" in txt or "CODIGO" in txt), 0)
            IDX_IND = next((i for txt, i in mapa_cols.items() if "INDICADOR" in txt), 1)
            IDX_OP_DESC = next((i for txt, i in mapa_cols.items() if "OPERANDOS" in txt and "ESTIMADO" not in txt), 3)
            IDX_OP_EST = next((i for txt, i in mapa_cols.items() if "ESTIMADO" in txt or "META" in txt and "CUMPLIMIENTO" not in txt), 4)
            idx_unidad = next((i for txt, i in mapa_cols.items() if "UNIDAD" in txt), None)
            idx_medios = next((i for txt, i in mapa_cols.items() if "MEDIO" in txt and "VERIFIC" in txt), None)

            idx_oct = next((i for txt, i in mapa_cols.items() if "OCT" in txt and "ACUM" not in txt), 5)
            idx_nov = next((i for txt, i in mapa_cols.items() if "NOV" in txt and "ACUM" not in txt), 7)
            idx_dic = next((i for txt, i in mapa_cols.items() if "DIC" in txt and "ACUM" not in txt), 9)

            extracted_rows = []
            for i in range(idx_header + 1, len(df) - 5):
                val_num = str(df.iloc[i, IDX_NUM]).strip()
                if pd.notna(df.iloc[i, IDX_NUM]) and "." in val_num and len(val_num) >= 3:
                    try:
                        meta_raw = df.iloc[i+1, IDX_OP_EST]
                        meta = limpiar_porcentaje(meta_raw) if isinstance(meta_raw, str) and "%" in meta_raw else limpiar_numero(meta_raw)
                        
                        op1_desc = limpiar_texto(df.iloc[i, IDX_OP_DESC])
                        if op1_desc.startswith("("): op1_desc = op1_desc[1:]
                        op2_desc = limpiar_texto(df.iloc[i+3, IDX_OP_DESC]).split(")")[0]
                        
                        unidad_val = limpiar_texto(df.iloc[i, idx_unidad]) if idx_unidad is not None else "Número"
                        medios_val = limpiar_texto(df.iloc[i, idx_medios]) if idx_medios is not None else "No aplica"

                        fila = {
                            "ORIGEN_ARCHIVO": nombre_archivo,
                            "NÚMERO": val_num,
                            "INDICADOR": limpiar_texto(df.iloc[i, IDX_IND]),
                            "CODIGO_RESPONSABLE_ASIGNADO": codigo_resp,
                            "Meta 2025 (%)": meta,
                            "Desc. Op1": op1_desc, "Desc. Op2": op2_desc,
                            "Est. Meta Op1": limpiar_numero(df.iloc[i+3, IDX_OP_EST]),
                            "Est. Meta Op2": limpiar_numero(df.iloc[i+5, IDX_OP_EST]),
                            "UNIDAD_EXTRAIDA": unidad_val,
                            "MEDIOS_EXTRAIDOS": medios_val
                        }
                        extracted_rows.append(fila)
                    except: continue

            if extracted_rows:
                dfs_extraidos.append(pd.DataFrame(extracted_rows))

        return pd.concat(dfs_extraidos, ignore_index=True) if dfs_extraidos else None
    except: return None

# =============================================================================
# 4. GENERADORES (F2 y F3 CORREGIDOS)
# =============================================================================

def generar_f2(df_f1):
    rows = []
    col_names = [
        'cod_interno', 'nombre_variable', 'descripcion', 'medio_verificacion',
        'APLICA_DIST_GENERO', 'APLICA_DESP_TERRITORIAL', 'APLICA_SIN_INFORMACION',
        'APLICA_VAL_PERS_JUR', 'requiere_medio', 'texto_ayuda', 'unidad',
        'valor_obligatorio', 'permite_medio_escrito', 'usa_ultimo_valor_ano'
    ]
    archivos = df_f1['ORIGEN_ARCHIVO'].unique()
    for archivo in archivos:
        sep_row = {c: None for c in col_names}
        sep_row['cod_interno'] = f"--- ORIGEN: {archivo} ---"
        rows.append(sep_row)
        subset = df_f1[df_f1['ORIGEN_ARCHIVO'] == archivo]
        for _, r in subset.iterrows():
            cod = str(r["NÚMERO"]).strip()
            medios = r["MEDIOS_EXTRAIDOS"] if r["MEDIOS_EXTRAIDOS"] else "No aplica"
            unidad = r["UNIDAD_EXTRAIDA"] if r["UNIDAD_EXTRAIDA"] else "Número"
            # A
            rows.append({
                "cod_interno": f"{cod}_A", "nombre_variable": r["Desc. Op1"], "descripcion": r["Desc. Op1"],
                "medio_verificacion": medios, "APLICA_DIST_GENERO": "?", "APLICA_DESP_TERRITORIAL": "?",
                "APLICA_SIN_INFORMACION": 1, "APLICA_VAL_PERS_JUR": None, "requiere_medio": None,
                "texto_ayuda": None, "unidad": unidad, "valor_obligatorio": 1, "permite_medio_escrito": 1, "usa_ultimo_valor_ano": 1
            })
            # B
            rows.append({
                "cod_interno": f"{cod}_B", "nombre_variable": r["Desc. Op2"], "descripcion": r["Desc. Op2"],
                "medio_verificacion": medios, "APLICA_DIST_GENERO": "?", "APLICA_DESP_TERRITORIAL": "?",
                "APLICA_SIN_INFORMACION": 1, "APLICA_VAL_PERS_JUR": None, "requiere_medio": None,
                "texto_ayuda": None, "unidad": unidad, "valor_obligatorio": 1, "permite_medio_escrito": 0, "usa_ultimo_valor_ano": 1
            })
    return pd.DataFrame(rows, columns=col_names)

def generar_f3(df_f2):
    rows = []
    # Definición de Columnas Solicitada
    col_names = [
        'cod_variable', 'nombre_variable', 'ano_mes_ini', 'ano_mes_fin', 
        'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 
        'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC',
        'cod_centro_resp_lugar_medicion', 'cod_region',
        'EMAIL_RESPONSABLE_INGRESO_DATO', 'EMAIL_PRIMER_REVISOR', 'EMAIL_SEGUNDO_REVISOR',
        'PERMITE_ADJUNTAR_MEDIO', 'MOSTRAR_TABLA_ANOS', 'FORMULA_VAR_AUTO', 'codigo_var_auto'
    ]
    
    current_file = ""
    
    for _, r in df_f2.iterrows():
        cod = r["cod_interno"]
        
        # Detector de separador para obtener nombre de archivo
        if str(cod).startswith("--- ORIGEN:"):
            # Extraer nombre del archivo del separador
            try:
                current_file = cod.replace("--- ORIGEN: ", "").replace(" ---", "")
            except: current_file = "?"
            
            sep_row = {c: None for c in col_names}
            sep_row['cod_variable'] = cod
            rows.append(sep_row)
            continue
            
        # Lógica de nombre regional
        nombre_region = extraer_nombre_limpio(current_file)
        centro_medicion = f"DIRECCION REGIONAL {nombre_region}"
        
        # Invertir código para var auto
        parts = cod.split('_')
        if len(parts) > 1:
            suffix = parts[-1]
            base = "_".join(parts[:-1])
            var_auto = f"{suffix}_{base}"
        else: var_auto = cod
        
        rows.append({
            "cod_variable": cod,
            "nombre_variable": r["nombre_variable"],
            "ano_mes_ini": 202501, "ano_mes_fin": 202512,
            "ENE": 1, "FEB": 1, "MAR": 1, "ABR": 1, "MAY": 1, "JUN": 1,
            "JUL": 1, "AGO": 1, "SEP": 1, "OCT": 1, "NOV": 1, "DIC": 1,
            "cod_centro_resp_lugar_medicion": centro_medicion,
            "cod_region": "?",
            "EMAIL_RESPONSABLE_INGRESO_DATO": "prueba@arbol-logika.com",
            "EMAIL_PRIMER_REVISOR": None,
            "EMAIL_SEGUNDO_REVISOR": None,
            "PERMITE_ADJUNTAR_MEDIO": 1,
            "MOSTRAR_TABLA_ANOS": 1,
            "FORMULA_VAR_AUTO": "SUMA_ANUAL",
            "codigo_var_auto": var_auto
        })
    return pd.DataFrame(rows, columns=col_names)

def generar_f4(df_f1):
    rows = []
    col_names = ['CODIGO', 'NOMBRE', 'DESCRIPCION', 'ACTIVO', 'UNIDAD', 'RANGO_MINIMO', 'RANGO_MAXIMO', 'FORMULA_COD', 'TIPO_META', 'FACTOR_CUMPLIMIENTO', 'FACTOR_NOCUMPLIMIENTO', 'IND_CDC', 'ANO_ASOCIADO']
    archivos = df_f1['ORIGEN_ARCHIVO'].unique()
    for archivo in archivos:
        sep_row = {c: None for c in col_names}
        sep_row['CODIGO'] = f"--- ORIGEN: {archivo} ---"
        rows.append(sep_row)
        subset = df_f1[df_f1['ORIGEN_ARCHIVO'] == archivo]
        for _, r in subset.iterrows():
            rows.append({
                "CODIGO": str(r["NÚMERO"]).strip(), "NOMBRE": r["INDICADOR"], "DESCRIPCION": r["INDICADOR"],
                "ACTIVO": 1, "UNIDAD": "%", "RANGO_MINIMO": 0, "RANGO_MAXIMO": 100,
                "FORMULA_COD": "PORCENTAJE", "TIPO_META": "TOLERANCIA",
                "FACTOR_CUMPLIMIENTO": 10, "FACTOR_NOCUMPLIMIENTO": 20, "IND_CDC": 1, "ANO_ASOCIADO": 2025
            })
    return pd.DataFrame(rows)

def generar_f5(df_f1):
    rows = []
    col_names = ['INDICADOR_COD', 'NOMBRE_INDICADOR', 'JER_TIPO_COD', 'EMAIL_RESPONSABLE', 'ANO_MES_INI', 'ANO_MES_FIN', 'TIPO_META_ANUAL', 'COMP_A', 'COMP_B', 'META_202512', 'Ponderacion', 'COD_PONDERADO', 'FORMULA_VAR_AUTO', 'COD_VAR_AUTO']
    archivos = df_f1['ORIGEN_ARCHIVO'].unique()
    for archivo in archivos:
        sep_row = {c: None for c in col_names}
        sep_row['INDICADOR_COD'] = f"--- ORIGEN: {archivo} ---"
        rows.append(sep_row)
        subset = df_f1[df_f1['ORIGEN_ARCHIVO'] == archivo]
        for _, r in subset.iterrows():
            cod = str(r["NÚMERO"]).strip()
            cod_pond = r["CODIGO_RESPONSABLE_ASIGNADO"]
            rows.append({
                "INDICADOR_COD": cod, "NOMBRE_INDICADOR": r["INDICADOR"],
                "JER_TIPO_COD": 1, "EMAIL_RESPONSABLE": "prueba@arbol-logika.com",
                "ANO_MES_INI": 202501, "ANO_MES_FIN": 202512, "TIPO_META_ANUAL": "PERIODO_ANUAL",
                "COMP_A": f"{cod}_A", "COMP_B": f"{cod}_B",
                "META_202512": r["Meta 2025 (%)"], "Ponderacion": 0,
                "COD_PONDERADO": cod_pond, "FORMULA_VAR_AUTO": "SUMA_ANUAL", "COD_VAR_AUTO": f"A_{cod_pond}"
            })
    return pd.DataFrame(rows)

# =============================================================================
# 5. EJECUCIÓN
# =============================================================================

def ejecutar_masivo():
    archivos = [f for f in glob.glob("*.xlsx") if not f.startswith("1_") and not f.startswith("2_") and not f.startswith("3_") and not f.startswith("~$")]
    print(f"\n[SIGI 25 v7.3.0] PROCESO MASIVO CON F3 CORREGIDA ({len(archivos)} archivos)")
    
    if not archivos: print("[ERROR] Carpeta vacía."); return

    master_list = []
    for idx, archivo in enumerate(archivos):
        df_ind = procesar_archivo(archivo)
        if df_ind is not None and not df_ind.empty:
            master_list.append(df_ind)
    
    if not master_list: print("\n[ERROR] No se extrajeron datos."); return

    print("\n   -> Generando Archivos Finales...")
    df_full = pd.concat(master_list, ignore_index=True)
    
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F1"], engine='openpyxl') as w:
        df_full.to_excel(w, index=False, sheet_name="DATOS_BRUTOS")
        df_full.to_excel(w, index=False, sheet_name="DATOS_ESTILIZADOS")
    aplicar_estilo_profesional(ARCHIVOS_SALIDA["F1"], "DATOS_ESTILIZADOS")
    
    f2 = generar_f2(df_full); f3 = generar_f3(f2); f4 = generar_f4(df_full); f5 = generar_f5(df_full)
    
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F2"], engine='openpyxl') as w:
        f2.to_excel(w, index=False, sheet_name="F2_VARIABLES")
        f3.to_excel(w, index=False, sheet_name="F3_VAR_APLICADAS")
        f4.to_excel(w, index=False, sheet_name="F4_INDICADORES")
        f5.to_excel(w, index=False, sheet_name="F5_IND_APLICADOS")
        
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F3"], engine='openpyxl') as w:
        f2.to_excel(w, index=False, sheet_name="VISUAL_VARIABLES")
        f3.to_excel(w, index=False, sheet_name="VISUAL_VAR_APP")
        f4.to_excel(w, index=False, sheet_name="VISUAL_INDICADORES")
        f5.to_excel(w, index=False, sheet_name="VISUAL_IND_APP")
    
    for h in ["VISUAL_VARIABLES", "VISUAL_VAR_APP", "VISUAL_INDICADORES", "VISUAL_IND_APP"]:
        aplicar_estilo_profesional(ARCHIVOS_SALIDA["F3"], h)

    print(f"\n   ¡LISTO! Revisa: {ARCHIVOS_SALIDA['F2']}")

if __name__ == "__main__":
    ejecutar_masivo()