"""
=============================================================================
 PROGRAMA: SIGI 25 (Sistema Integral de Gestión de Indicadores)
 VERSIÓN:  v4.0.0 (Fuzzy Matching Engine)
 FECHA:    Febrero 2026
 
 CAMBIOS CRÍTICOS:
 1. MOTOR DE CRUCE: Cambiado de "Igualdad Exacta" a "Contiene Palabra Clave".
    Esto soluciona el problema de archivos con fechas ("2025 oct...") o versiones.
 2. MAPA SIMPLIFICADO: Las claves ahora son las raíces del nombre (ej: "BENEFICIOS").
 3. DEDUPLICACIÓN: Se asegura de que si procesa 2 archivos de Beneficios, solo quede uno.
=============================================================================
"""

import pandas as pd
import os
import re
import glob
import unicodedata
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Silenciar advertencias
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)

# =============================================================================
# 1. CONFIGURACIÓN
# =============================================================================

ANO_PROCESO = "2026"

ARCHIVOS_SALIDA = {
    "F1": f"1_PLANILLA_SIG_CONSOLIDADO_{ANO_PROCESO}.xlsx",
    "F2": f"2_CARGA_BRUTA_CONSOLIDADO_{ANO_PROCESO}.xlsx",
    "F3": f"3_REPORTE_VISUAL_CONSOLIDADO_{ANO_PROCESO}.xlsx"
}

# MAPA INTELIGENTE (CLAVES = PALABRAS CLAVE ÚNICAS)
# El sistema buscará si el nombre del archivo CONTIENE estas claves.
MAPA_PALABRAS_CLAVE = {
    # Divisiones / Deptos
    'BENEFICIOS': 'IP25_712',
    'CLIENTES': 'IP25_713',
    'INFORMATICA': 'IP25_714',
    'JURIDICA': 'IP25_715',
    'PLANIFICACION': 'IP25_716',
    'COMUNICACIONES': 'IP25_717',
    'CONTRALORIA': 'IP25_718',
    'AUDITORIA': 'IP25_738',
    'SISTEMASDEINFORMACION': 'IP25_739', 'SISTINFORM': 'IP25_739',
    'GESTIONPERSONAS': 'IP25_750', 'DESARROLLODEPERSONAS': 'IP25_750',
    'FORMULARIOH': 'IP25_711',
    
    # Regiones (Usamos nombres únicos para evitar confusión)
    'ARICAYPARINACOTA': 'IP25_719', 'ARICA': 'IP25_719',
    'TARAPACA': 'IP25_720',
    'ANTOFAGASTA': 'IP25_721',
    'ATACAMA': 'IP25_722',
    'COQUIMBO': 'IP25_723',
    'VALPARAISO': 'IP25_724',
    'OHIGGINS': 'IP25_725', 'LIBERTADOR': 'IP25_725',
    'MAULE': 'IP25_726',
    'NUBLE': 'IP25_748', 'ÑUBLE': 'IP25_748',
    'BIOBIO': 'IP25_727',
    'ARAUCANIA': 'IP25_728',
    'LOSRIOS': 'IP25_729',
    'LOSLAGOS': 'IP25_730',
    'AYSEN': 'IP25_731', 'AISEN': 'IP25_731',
    'MAGALLANES': 'IP25_732',
    'METROPOLITANA': 'IP25_733',
    
    # PMG (Por si acaso vienen en el nombre)
    'GESTIONEFICAZ': 'IP25_740',
    'EFICIENCIAINSTITUCIONAL': 'IP25_741',
    'CALIDADDELOSSERVICIOS': 'IP25_742',
    'EXPERIENCIAUSUARIA': 'IP25_752'
}

# =============================================================================
# 2. UTILS
# =============================================================================

def norm_txt(txt):
    """Limpia el texto dejando solo letras mayúsculas (A-Z) para comparación."""
    if pd.isna(txt): return ""
    t = str(txt).upper().strip()
    # Eliminar tildes
    t = ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')
    # Dejar solo letras (eliminar espacios, números, puntos, guiones)
    return re.sub(r'[^A-Z]', '', t)

def encontrar_codigo_ponderado(texto_origen):
    """
    MOTOR DE BÚSQUEDA FUZZY:
    Recorre el mapa y verifica si la CLAVE está DENTRO del texto_origen normalizado.
    Ej: Clave 'BENEFICIOS' está en 'CDCBENEFICIOS2025OCT'? -> SÍ -> Retorna IP25_712
    """
    texto_norm = norm_txt(texto_origen)
    
    # Prioridad: Buscar coincidencias más largas primero para evitar falsos positivos
    # (Ej: Evitar que 'BIOBIO' coincida con algo que solo tenga 'BIO')
    claves_ordenadas = sorted(MAPA_PALABRAS_CLAVE.keys(), key=len, reverse=True)
    
    for clave in claves_ordenadas:
        if clave in texto_norm:
            return MAPA_PALABRAS_CLAVE[clave]
            
    return "?"

def inferir_entidad_desde_archivo(filename):
    return os.path.splitext(os.path.basename(filename))[0]

def limpiar_porcentaje_real(val):
    if pd.isna(val) or val == "" or str(val).lower() == "no aplica": return val
    if isinstance(val, str):
        try: return float(val.replace('%', '').replace(',', '.').strip())
        except: return 0
    if isinstance(val, (int, float)): return val * 100
    return 0

def limpiar_op1_inicio(val):
    if pd.isna(val) or val == "": return ""
    t = str(val).strip()
    return t[1:].strip() if t.startswith("(") else t

def limpiar_op2_final(val):
    if pd.isna(val) or val == "": return ""
    t = str(val).strip()
    return re.sub(r'\)\s*\*100$', '', t).strip()

def detectar_encabezados_flex(df):
    for i in range(min(30, len(df))):
        try:
            fila = [str(x).upper().strip() for x in df.iloc[i].tolist()]
            if any("INDICADOR" in x for x in fila):
                return i, df.iloc[i].astype(str).tolist()
        except: continue
    return None, None

def transform_var_code(cod):
    s = str(cod).strip()
    if s.startswith("---"): return None
    if "INDICADOR_NUEVO" in s:
        p = s.split('_')
        if len(p) >= 5: return f"{p[-2]}_{'_'.join(p[:-2])}_{p[-1]}"
        return f"{p[-1]}_{'_'.join(p[:-1])}"
    elif '_' in s:
        p = s.rsplit('_', 1)
        if len(p) == 2: return f"{p[1]}_{p[0]}"
    return s

def parsear_nombre(txt):
    if pd.isna(txt): return "EFICACIA", "PROCESO", ""
    t = str(txt).strip()
    m = re.search(r"^(?:[\d\)\(\s]+)?([a-zA-ZáéíóúñÁÉÍÓÚÑ]+)/([a-zA-ZáéíóúñÁÉÍÓÚÑ]+)\s+(.*)", t, re.DOTALL)
    if m: return m.group(1).upper(), m.group(2).upper(), m.group(3).replace('\n', ' ').strip()
    return "EFICACIA", "PROCESO", t.replace('\n', ' ').strip()

def determinar_unidad(n):
    s = str(n).lower()
    if "porcentaje" in s or "%" in s: return "%"
    if any(x in s for x in ["tiempo", "medidas", "numero", "número", "cantidad", "tasa"]): return "n"
    return "?"

def limpiar_cod_ind(rn, c, e):
    s = str(rn).strip()
    if not rn or pd.isna(rn) or s == "" or s.lower() == "nan" or "NUEVO" in s.upper():
        return f"IND_NUEVO_{c}_{e}", True
    return s, False

def aplicar_estilo_profesional(ruta, hoja):
    try:
        wb = load_workbook(ruta)
        if hoja not in wb.sheetnames: return
        ws = wb[hoja]
        fill = PatternFill("solid", fgColor="1F4E78")
        font_h = Font(b=True, color="FFFFFF", size=10)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws[1]:
            cell.fill = fill; cell.font = font_h; cell.border = border
        for row in ws.iter_rows(min_row=2):
            sep = str(row[0].value).startswith("---")
            for c in row:
                c.border = border
                if sep: c.font = Font(b=True); c.fill = PatternFill("solid", fgColor="D9D9D9")
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        wb.save(ruta)
    except: pass

# =============================================================================
# 3. LÓGICA DE EXTRACCIÓN
# =============================================================================

def procesar_archivo_fase1(ruta, entidad_inferred):
    dfs_archivo = []
    try:
        xls = pd.ExcelFile(ruta)
        hojas_disponibles = xls.sheet_names
        hojas_a_procesar = {}
        for h in hojas_disponibles:
            h_up = h.upper()
            if "CDC" in h_up: hojas_a_procesar[h] = "CDC"
            elif "PMG" in h_up: hojas_a_procesar[h] = "PMG"
            elif "RIESGO" in h_up: hojas_a_procesar[h] = "Riesgos"
            
        if not hojas_a_procesar:
            print(f"      [SKIP] Sin hojas relevantes. Disp: {hojas_disponibles}")
            return None

        for hoja_excel, etiqueta in hojas_a_procesar.items():
            df = pd.read_excel(ruta, sheet_name=hoja_excel, header=None)
            idx_h, heads = detectar_encabezados_flex(df)
            if idx_h is None:
                continue
                
            map_c = {str(n).strip().replace("\n", " ").upper(): i for i, n in enumerate(heads)}
            def gc(k):
                if isinstance(k, str): k=[k]
                for x in k: 
                    x_up = x.upper()
                    for n,i in map_c.items(): 
                        if x_up in n: return i
                return None
            
            idx_num = gc(["NÚMERO", "NUMERO", "N°", "NO.", "CODIGO"])
            if idx_num is None: idx_num = 0
            
            idx_ind = gc("INDICADOR"); idx_resp = gc("RESPONSABLE")
            idx_meta = gc("Meta 2025"); idx_pond = gc("Ponderador")
            idx_op_desc = gc("Operandos"); idx_med = gc("Medios")
            
            starts = []
            for i in range(idx_h+1, len(df)):
                val_num = str(df.iloc[i, idx_num]).strip().upper()
                if pd.notna(df.iloc[i, idx_num]) and val_num not in ["NÚMERO", "NUMERO", "NAN", ""]:
                    starts.append(i)
            
            if not starts: continue

            rows = []
            meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sept", "Oct", "Nov", "Dic"]
            map_m = {m: gc(f"{m}.") for m in meses}
            
            for i in starts:
                r = {}
                r["ORIGEN_ARCHIVO"] = os.path.basename(ruta)
                r["NÚMERO"] = df.iloc[i, idx_num]
                r["INDICADOR"] = df.iloc[i, idx_ind] if idx_ind else ""
                
                # Extracción Responsable: Prioridad Columna > Inferencia Archivo
                resp_val = df.iloc[i, idx_resp] if idx_resp else None
                if pd.isna(resp_val) or str(resp_val).strip() == "":
                    r["RESPONSABLE CENTRO DE RESPONSABILIDAD"] = entidad_inferred
                else:
                    r["RESPONSABLE CENTRO DE RESPONSABILIDAD"] = resp_val
                    
                r["Meta 2025 (%)"] = df.iloc[i, idx_meta] if idx_meta else 100
                r["Ponderador (%)"] = df.iloc[i, idx_pond] if idx_pond else 0
                
                if idx_op_desc:
                    r["Desc. Op1"] = df.iloc[i, idx_op_desc]
                    try: r["Desc. Op2"] = df.iloc[i+3, idx_op_desc]
                    except: r["Desc. Op2"] = ""
                
                for m in meses:
                    ci = map_m.get(m)
                    if ci:
                        r[f"{m} Ind (%)"] = df.iloc[i+1, ci]
                        try: r[f"{m} Op1"] = df.iloc[i+3, ci]; r[f"{m} Op2"] = df.iloc[i+5, ci]
                        except: r[f"{m} Op1"]=0; r[f"{m} Op2"]=0
                    else:
                        r[f"{m} Ind (%)"]="No aplica"; r[f"{m} Op1"]="No aplica"; r[f"{m} Op2"]="No aplica"
                        
                r["Medios Verificación"] = df.iloc[i, idx_med] if idx_med else ""
                rows.append(r)
            
            if rows:
                sub = pd.DataFrame(rows)
                sep = {c: None for c in sub.columns}; sep['NÚMERO'] = f"--- {etiqueta} ({entidad_inferred}) ---"
                dfs_archivo.append(pd.DataFrame([sep]))
                dfs_archivo.append(sub)
    except Exception as e:
        print(f"   [ERROR] {os.path.basename(ruta)}: {e}")
    return pd.concat(dfs_archivo, ignore_index=True) if dfs_archivo else None

def rearmar(df_master):
    dic = {"CDC": [], "PMG": [], "Riesgos": []}
    curr = "CDC"
    for _, row in df_master.iterrows():
        v = str(row['NÚMERO']).upper()
        if v.startswith("---"):
            if "CDC" in v: curr = "CDC"
            elif "PMG" in v: curr = "PMG"
            elif "RIESGO" in v: curr = "Riesgos"
        else:
            dic[curr].append(row)
    return {k: pd.DataFrame(v) for k,v in dic.items() if v}

# --- GENERADORES F2-F5 ---

def gen_f2(inputs):
    res = []; cg = 1
    cols = ['cod_interno', 'nombre_variable', 'descripcion', 'medio_verificacion', 'APLICA_DIST_GENERO', 'APLICA_DESP_TERRITORIAL', 'APLICA_SIN_INFORMACION', 'APLICA_VAL_PERS_JUR', 'requiere_medio', 'texto_ayuda', 'unidad', 'valor_obligatorio', 'permite_medio_escrito', 'usa_ultimo_valor_ano']
    for tag, df in inputs.items():
        if df.empty: continue
        rows = []
        for _, r in df.iterrows():
            rn = r.get('NÚMERO', ''); c_str = str(rn).strip()
            ci, nw = limpiar_cod_ind(rn, cg, tag)
            if nw: cg += 1
            cA, cB = f"{ci}_A", f"{ci}_B"
            nA = limpiar_op1_inicio(r.get('Desc. Op1', '')); nB = limpiar_op2_final(r.get('Desc. Op2', ''))
            med = str(r.get('Medios Verificación', '')).strip()
            base = {'APLICA_DIST_GENERO': 0, 'APLICA_DESP_TERRITORIAL': 0, 'APLICA_SIN_INFORMACION': 1, 'APLICA_VAL_PERS_JUR': 0, 'requiere_medio': 0, 'texto_ayuda': None, 'unidad': None, 'valor_obligatorio': 1, 'permite_medio_escrito': 1, 'usa_ultimo_valor_ano': 1}
            rA = base.copy(); rA.update({'cod_interno': cA, 'nombre_variable': nA, 'descripcion': nA, 'medio_verificacion': med})
            rB = base.copy(); rB.update({'cod_interno': cB, 'nombre_variable': nB, 'descripcion': nB, 'medio_verificacion': med})
            rows.extend([rA, rB])
        if rows:
            t = {c: None for c in cols}; t['cod_interno'] = f"--- {tag} VARIABLES ---"
            res.append(pd.DataFrame([t])); res.append(pd.DataFrame(rows))
    final = pd.concat(res, ignore_index=True).reindex(columns=cols) if res else pd.DataFrame()
    mask = ~final['cod_interno'].astype(str).str.startswith("---")
    return final[mask].drop_duplicates(subset=['cod_interno'], keep='last')

def gen_f3(f2_df):
    rows = []; cols = ['cod_variable', 'nombre_variable', 'ano_mes_ini', 'ano_mes_fin', 'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC', 'cod_centro_resp_lugar_medicion', 'cod_region', 'EMAIL_RESPONSABLE_INGRESO_DATO', 'EMAIL_PRIMER_REVISOR', 'EMAIL_SEGUNDO_REVISOR', 'PERMITE_ADJUNTAR_MEDIO', 'MOSTRAR_TABLA_ANOS', 'FORMULA_VAR_AUTO', 'codigo_var_auto']
    seen = set()
    for _, r in f2_df.iterrows():
        c = str(r.get('cod_interno', '')).strip()
        if c.startswith("---") or c in seen: continue
        seen.add(c)
        rows.append({
            'cod_variable': c, 'nombre_variable': r.get('nombre_variable', ''),
            'ano_mes_ini': 202501, 'ano_mes_fin': 202512,
            'ENE': 1, 'FEB': 1, 'MAR': 1, 'ABR': 1, 'MAY': 1, 'JUN': 1, 'JUL': 1, 'AGO': 1, 'SEP': 1, 'OCT': 1, 'NOV': 1, 'DIC': 1,
            'cod_centro_resp_lugar_medicion': None, 'cod_region': None,
            'EMAIL_RESPONSABLE_INGRESO_DATO': 'prueba@arbol-logika.com', 'EMAIL_PRIMER_REVISOR': None, 'EMAIL_SEGUNDO_REVISOR': None,
            'PERMITE_ADJUNTAR_MEDIO': 1, 'MOSTRAR_TABLA_ANOS': 1, 'FORMULA_VAR_AUTO': 'SUMA_ANUAL',
            'codigo_var_auto': transform_var_code(c)
        })
    return pd.DataFrame(rows).reindex(columns=cols)

def gen_f4(inputs):
    res = []; cg = 1
    cols = ['CODIGO', 'NOMBRE', 'DESCRIPCION', 'ACTIVO', 'UNIDAD', 'RANGO_MINIMO', 'RANGO_MAXIMO', 'APLICA_DIST_GENERO', 'APLICA_SIN_INFORMACION', 'APLICA_VAL_PERS_JUR', 'APLICA_DESP_TERRITORIAL', 'VALOR_DEFECTO', 'AMBITO_COD', 'DIMENSION_COD', 'PERSPECTIVA_COD', 'FORMULA_COD', 'PROD_ESTRATEGICO_COD', 'OBJ_SERVICIO_COD', 'SENTIDO_META', 'TIPO_META', 'FACTOR_CUMPLIMIENTO', 'FACTOR_NOCUMPLIMIENTO', 'FACTOR_SOBRECUMPLIMIENTO', 'IND_BGI', 'IND_CDC', 'IND_PROP', 'IND_DISC', 'IND_H', 'IND_INT', 'IND_H_NO_PMG', 'IND_PRIO', 'IND_PLAC', 'IND_PMG', 'IND_RIESGO', 'IND_TRANS', 'ANO_ASOCIADO']
    for tag, df in inputs.items():
        if df.empty: continue
        rows = []
        for _, r in df.iterrows():
            rn = r.get('NÚMERO', '')
            ci, nw = limpiar_cod_ind(rn, cg, tag)
            if nw: cg += 1
            amb, dim, nom = parsear_nombre(r.get('INDICADOR', ''))
            fl = {'IND_BGI': 0, 'IND_CDC': 0, 'IND_PROP': 0, 'IND_DISC': 0, 'IND_H': 0, 'IND_INT': 0, 'IND_H_NO_PMG': 0, 'IND_PRIO': 0, 'IND_PLAC': 0, 'IND_PMG': 0, 'IND_RIESGO': 0, 'IND_TRANS': 0}
            if tag == "CDC": fl['IND_CDC'] = 1
            elif tag == "Riesgos": fl['IND_RIESGO'] = 1
            elif tag == "PMG": fl['IND_PMG'] = 1
            rows.append({
                'CODIGO': ci, 'NOMBRE': nom, 'DESCRIPCION': nom, 'ACTIVO': 1, 'UNIDAD': determinar_unidad(nom), 'RANGO_MINIMO': 0, 'RANGO_MAXIMO': 100,
                'APLICA_DIST_GENERO': 0, 'APLICA_SIN_INFORMACION': 0, 'APLICA_VAL_PERS_JUR': 0, 'APLICA_DESP_TERRITORIAL': 0, 'VALOR_DEFECTO': 0,
                'AMBITO_COD': amb, 'DIMENSION_COD': dim, 'PERSPECTIVA_COD': None, 'FORMULA_COD': "PORCENTAJE",
                'PROD_ESTRATEGICO_COD': None, 'OBJ_SERVICIO_COD': None, 'SENTIDO_META': 1, 'TIPO_META': "TOLERANCIA",
                'FACTOR_CUMPLIMIENTO': 10, 'FACTOR_NOCUMPLIMIENTO': 20, 'FACTOR_SOBRECUMPLIMIENTO': 0, **fl, 'ANO_ASOCIADO': 2025
            })
        if rows:
            t = {c: None for c in cols}; t['CODIGO'] = f"--- {tag} INDICADORES ---"
            res.append(pd.DataFrame([t])); res.append(pd.DataFrame(rows))
    final = pd.concat(res, ignore_index=True).reindex(columns=cols) if res else pd.DataFrame()
    mask = ~final['CODIGO'].astype(str).str.startswith("---")
    return final[mask].drop_duplicates(subset=['CODIGO'], keep='last')

def gen_f5(inputs):
    res = []; cg = 1
    cols = ['INDICADOR_COD', 'NOMBRE_INDICADOR', 'JER_TIPO_COD', 'CENTRO_RESP_COD', 'COD_REGION', 'EMAIL_RESPONSABLE', 'COD_GENERO', 'ANO_MES_INI', 'ANO_MES_FIN', 'COD_ANALISIS_CAUSA', 'EMAIL_RESP_ANALISIS_CAUSA', 'CREAR_COMENTARIO_FORM', 'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC', 'ORIGEN', 'NOTAS_SUPUESTOS', 'MOSTRAR_PANEL', 'APLICA_RIESGO', 'OBJ_ESTRATEGICO_COD', 'OBJ_ESPECIFICO_COD', 'PROD_ESTRATEGICO_COD', 'COD_PROGRAMA', 'COD_COMPONENTE_PROG', 'TIPO_META_ANUAL', 'COMP_A', 'COMP_A_CR', 'COMP_A_GEN', 'COMP_A_REGION', 'COMP_B', 'COMP_B_CR', 'COMP_B_GEN', 'COMP_B_REGION', 'CONST_A', 'META_202512', 'Ponderacion', 'COD_PONDERADO', 'FORMULA_VAR_AUTO', 'COD_VAR_AUTO']
    for tag, df in inputs.items():
        if df.empty: continue
        rows = []
        for _, r in df.iterrows():
            rn = r.get('NÚMERO', '')
            ci, nw = limpiar_cod_ind(rn, cg, tag)
            if nw: cg += 1
            _, _, nom = parsear_nombre(r.get('INDICADOR', ''))
            
            # --- MOTOR DE CRUCE FUZZY (NUEVO) ---
            # 1. Probar con nombre de columna si existe
            raw_resp = r.get('RESPONSABLE CENTRO DE RESPONSABILIDAD', '')
            cod_pond = encontrar_codigo_ponderado(raw_resp)
            
            # 2. Si falló o es el nombre del archivo (que puede estar sucio), reintentar
            # usando el nombre de archivo limpio de "Planilla SIG", etc.
            if cod_pond == "?":
                fname = r.get("ORIGEN_ARCHIVO", "")
                cod_pond = encontrar_codigo_ponderado(fname)
            
            c_auto = f"A_{cod_pond}" if cod_pond != "?" else "?"
            
            rows.append({
                'INDICADOR_COD': ci, 'NOMBRE_INDICADOR': nom, 'JER_TIPO_COD': 1, 'CENTRO_RESP_COD': None, 'COD_REGION': None,
                'EMAIL_RESPONSABLE': 'prueba@arbol-logika.com', 'COD_GENERO': None, 'ANO_MES_INI': 202501, 'ANO_MES_FIN': 202512,
                'COD_ANALISIS_CAUSA': 'RESP_INDICADOR', 'EMAIL_RESP_ANALISIS_CAUSA': None, 'CREAR_COMENTARIO_FORM': None,
                'ENE': 1, 'FEB': 1, 'MAR': 1, 'ABR': 1, 'MAY': 1, 'JUN': 1, 'JUL': 1, 'AGO': 1, 'SEP': 1, 'OCT': 1, 'NOV': 1, 'DIC': 1,
                'ORIGEN': None, 'NOTAS_SUPUESTOS': None, 'MOSTRAR_PANEL': None, 'APLICA_RIESGO': None,
                'OBJ_ESTRATEGICO_COD': None, 'OBJ_ESPECIFICO_COD': None, 'PROD_ESTRATEGICO_COD': None, 'COD_PROGRAMA': None, 'COD_COMPONENTE_PROG': None,
                'TIPO_META_ANUAL': 'PERIODO_ANUAL',
                'COMP_A': f"{ci}_A", 'COMP_A_CR': None, 'COMP_A_GEN': None, 'COMP_A_REGION': None,
                'COMP_B': f"{ci}_B", 'COMP_B_CR': None, 'COMP_B_GEN': None, 'COMP_B_REGION': None,
                'CONST_A': None, 'META_202512': r.get('Meta 2025 (%)', 0),
                'Ponderacion': r.get('Ponderador (%)', 0) if tag == "CDC" else None,
                'COD_PONDERADO': cod_pond, 'FORMULA_VAR_AUTO': 'SUMA_ANUAL', 'COD_VAR_AUTO': c_auto
            })
        if rows:
            t = {c: None for c in cols}; t['INDICADOR_COD'] = f"--- {tag} APLICADOS ---"
            res.append(pd.DataFrame([t])); res.append(pd.DataFrame(rows))
    final = pd.concat(res, ignore_index=True).reindex(columns=cols) if res else pd.DataFrame()
    mask = ~final['INDICADOR_COD'].astype(str).str.startswith("---")
    return final[mask].drop_duplicates(subset=['INDICADOR_COD'], keep='last')

# =============================================================================
# 4. ORQUESTACIÓN
# =============================================================================

def ejecutar_masivo():
    archivos = [f for f in glob.glob("*.xlsx") if not f.startswith("1_") and not f.startswith("2_") and not f.startswith("3_") and not f.startswith("~$")]
    print(f"\n[SIGI 25 v4.0.0] INICIANDO PROCESAMIENTO (FUZZY MATCH)")
    print(f"   -> Archivos detectados: {len(archivos)}")
    
    if not archivos: print("   [ERROR] No hay archivos Excel."); return

    master_list = []
    
    for idx, archivo in enumerate(archivos):
        entidad = inferir_entidad_desde_archivo(archivo)
        print(f"   [{idx+1}/{len(archivos)}] Procesando: {archivo} ...", end=" ")
        
        df_ind = procesar_archivo_fase1(archivo, entidad)
        if df_ind is not None and not df_ind.empty:
            master_list.append(df_ind)
            print("OK")
        else:
            print("VACÍO/SKIP")

    if not master_list: print("\n[ERROR] No se pudo extraer nada."); return

    print("\n   -> Consolidando Fases...")
    df_consolidado = pd.concat(master_list, ignore_index=True)
    dic_consolidado = rearmar(df_consolidado)
    
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F1"], engine='openpyxl') as w:
        df_consolidado.to_excel(w, "DATOS_BRUTOS", index=False)
        df_consolidado.to_excel(w, "DATOS_ESTILIZADOS", index=False)
    aplicar_estilo_profesional(ARCHIVOS_SALIDA["F1"], "DATOS_ESTILIZADOS")

    print("   -> Generando Paquetes de Carga...")
    f2 = gen_f2(dic_consolidado); f3 = gen_f3(f2); f4 = gen_f4(dic_consolidado); f5 = gen_f5(dic_consolidado)
    
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F2"], engine='openpyxl') as w:
        f2.to_excel(w, "F2_VARIABLES", index=False)
        f3.to_excel(w, "F3_VAR_APLICADAS", index=False)
        f4.to_excel(w, "F4_INDICADORES", index=False)
        f5.to_excel(w, "F5_IND_APLICADOS", index=False)
        
    with pd.ExcelWriter(ARCHIVOS_SALIDA["F3"], engine='openpyxl') as w:
        f2.to_excel(w, "VISUAL_VARIABLES", index=False)
        f3.to_excel(w, "VISUAL_VAR_APP", index=False)
        f4.to_excel(w, "VISUAL_INDICADORES", index=False)
        f5.to_excel(w, "VISUAL_IND_APP", index=False)
    for h in ["VISUAL_VARIABLES", "VISUAL_VAR_APP", "VISUAL_INDICADORES", "VISUAL_IND_APP"]:
        aplicar_estilo_profesional(ARCHIVOS_SALIDA["F3"], h)

    print("\n   ¡PROCESO COMPLETADO! Archivos listos:")
    print(f"   1. {ARCHIVOS_SALIDA['F1']}")
    print(f"   2. {ARCHIVOS_SALIDA['F2']}")
    print(f"   3. {ARCHIVOS_SALIDA['F3']}")

if __name__ == "__main__":
    ejecutar_masivo()