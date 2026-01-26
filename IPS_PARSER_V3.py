import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# ==========================================
# 1. MOTOR DE PROCESAMIENTO (RETORNA DATAFRAMES)
# ==========================================

def limpiar_porcentaje_real(val):
    """Convierte texto '20%' o número 0.2 a valor limpio 20."""
    if pd.isna(val) or val == "": return 0
    if isinstance(val, str):
        limpio = val.replace('%', '').replace(',', '.').strip()
        try: return float(limpio)
        except: return 0
    if isinstance(val, (int, float)): return val * 100
    return 0

def detectar_encabezados(df):
    """Escanea las primeras 25 filas buscando dónde empieza la tabla."""
    for i in range(25):
        try:
            fila = df.iloc[i].astype(str).tolist()
            if "NÚMERO" in fila and any("INDICADOR" in s for s in fila):
                return i, fila
        except:
            continue
    raise ValueError("ERROR: No se encontró la fila de encabezados.")

def obtener_dataframe_hoja(ruta_archivo, nombre_hoja_excel, etiqueta_log):
    """
    Lee una hoja específica y devuelve el DataFrame procesado y limpio.
    NO guarda archivos, solo retorna los datos.
    """
    print(f"   -> Procesando datos de: {etiqueta_log}...")
    
    try:
        df_raw = pd.read_excel(ruta_archivo, sheet_name=nombre_hoja_excel, header=None)
    except Exception as e:
        print(f"      [ERROR] No se pudo leer la hoja '{nombre_hoja_excel}': {e}")
        return None

    # A. Detección Estructura
    try:
        idx_header, encabezados = detectar_encabezados(df_raw)
    except ValueError as e:
        print(f"      [ERROR] {e}")
        return None

    mapa_cols = {str(nombre).strip().replace("\n", " "): i for i, nombre in enumerate(encabezados)}
    
    def buscar_col(claves):
        if isinstance(claves, str): claves = [claves]
        for c in claves:
            for nombre_real, idx in mapa_cols.items():
                if c.lower() in nombre_real.lower():
                    return idx
        return None

    # Índices
    idx_num = buscar_col("NÚMERO")
    idx_meta = buscar_col("Meta 2025")
    idx_ponderador = buscar_col("Ponderador")
    idx_op_desc = buscar_col("Operandos")
    idx_op_est = buscar_col(["Operandos Estimados", "Operandos  Estimados"])
    
    # Columnas finales
    idx_medios = buscar_col("Medios de Verificación")
    idx_control = buscar_col("Control de Cambios")
    idx_inst = buscar_col(["Instrumentos de Gestión", "Instrumentos Gestión"])
    idx_cump_meta = buscar_col("% Cumplimiento de Meta")

    # B. Extracción
    indices_inicio = []
    for i in range(idx_header + 1, len(df_raw)):
        val = df_raw.iloc[i, idx_num]
        if pd.notna(val) and str(val).strip() != "" and str(val) != "NÚMERO":
            indices_inicio.append(i)
            
    lista_datos = []
    
    # Mapa Meses
    meses_std = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sept", "Oct", "Nov", "Dic"]
    mapa_meses_indices = {}
    for mes in meses_std:
        idx = buscar_col(f"{mes}.") 
        if idx: mapa_meses_indices[mes] = idx
    
    idx_proy = buscar_col(["Cumplimiento Proyectado", "Cumplimiento\nProyectado"])
    if idx_proy: mapa_meses_indices["Cump. Proy."] = idx_proy

    # Bucle Principal
    for idx in indices_inicio:
        fila = {}
        # Identificación
        fila["NÚMERO"] = df_raw.iloc[idx, idx_num]
        
        col_prod = buscar_col("PRODUCTO")
        if col_prod: fila["PRODUCTO O PROCESO ESPECÍFICO"] = df_raw.iloc[idx, col_prod]
        
        col_ind = buscar_col("INDICADOR")
        if col_ind: fila["INDICADOR"] = df_raw.iloc[idx, col_ind]
        
        col_form = buscar_col("FORMULA")
        if col_form: fila["FORMULA"] = df_raw.iloc[idx, col_form]
        
        col_unidad = buscar_col("UNIDAD")
        if col_unidad: fila["UNIDAD"] = df_raw.iloc[idx, col_unidad]
        
        # Responsables
        col_resp = buscar_col("RESPONSABLE CENTRO")
        if col_resp: fila["RESPONSABLE CENTRO DE RESPONSABILIDAD"] = df_raw.iloc[idx, col_resp]
        col_gestor = buscar_col("GESTOR")
        if col_gestor: fila["GESTOR"] = df_raw.iloc[idx, col_gestor]
        col_sup = buscar_col("SUPERVISORES")
        if col_sup: fila["SUPERVISORES"] = df_raw.iloc[idx, col_sup]
        
        # Metas
        if idx_meta: fila["Meta 2025 (%)"] = df_raw.iloc[idx, idx_meta]
        if idx_ponderador: fila["Ponderador (%)"] = df_raw.iloc[idx, idx_ponderador]
        else: fila["Ponderador (%)"] = 0 
        
        # Operandos
        if idx_op_desc:
            fila["Desc. Op1"] = df_raw.iloc[idx, idx_op_desc]
            fila["Desc. Op2"] = df_raw.iloc[idx+3, idx_op_desc]
        if idx_op_est:
            fila["Est. Meta Op1"] = df_raw.iloc[idx+3, idx_op_est]
            fila["Est. Meta Op2"] = df_raw.iloc[idx+5, idx_op_est]
            
        # Meses
        for nombre_mes, col_idx in mapa_meses_indices.items():
            fila[f"{nombre_mes} Ind (%)"] = df_raw.iloc[idx+1, col_idx]
            fila[f"{nombre_mes} Op1"] = df_raw.iloc[idx+3, col_idx]
            fila[f"{nombre_mes} Op2"] = df_raw.iloc[idx+5, col_idx]
            
        # Finales
        if idx_cump_meta: fila["Cumplimiento Meta (%)"] = df_raw.iloc[idx+3, idx_cump_meta]
        if idx_medios: fila["Medios Verificación"] = df_raw.iloc[idx, idx_medios]
        if idx_control: fila["Control Cambios"] = df_raw.iloc[idx, idx_control]
        if idx_inst: fila["Instrumentos Gestión"] = df_raw.iloc[idx, idx_inst]
        
        lista_datos.append(fila)

    # DataFrame Final
    df_final = pd.DataFrame(lista_datos).fillna(0)
    for col in df_final.columns:
        if "(%)" in col:
            df_final[col] = df_final[col].apply(limpiar_porcentaje_real)
            
    return df_final

def aplicar_estilos_global(nombre_archivo):
    """Aplica estilos a TODAS las hojas del archivo Excel."""
    print(f"   -> Aplicando maquillaje visual a: {nombre_archivo}...")
    try:
        wb = load_workbook(nombre_archivo)
        
        # Estilos definidos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        cols_texto = ["PRODUCTO O PROCESO ESPECÍFICO", "INDICADOR", "FORMULA", "Desc. Op1", "Desc. Op2", "Medios Verificación", "Control Cambios", "Instrumentos Gestión"]
        cols_resp = ["UNIDAD", "RESPONSABLE CENTRO DE RESPONSABILIDAD", "GESTOR", "SUPERVISORES"]

        # Iterar por cada hoja del libro
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for col_cells in ws.columns:
                col_name = col_cells[0].value
                col_letter = col_cells[0].column_letter
                
                # Anchos
                if col_name in cols_texto: ws.column_dimensions[col_letter].width = 40 
                elif col_name in cols_resp: ws.column_dimensions[col_letter].width = 30
                elif col_name and len(str(col_name)) < 6: ws.column_dimensions[col_letter].width = 10 
                else: ws.column_dimensions[col_letter].width = 18

                # Celdas
                for cell in col_cells:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        if col_name in cols_texto or col_name in cols_resp:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(nombre_archivo)
        print("      [OK] Estilos aplicados correctamente.")
        
    except Exception as e:
        print(f"      [ERROR] Falló el estilizado: {e}")

# ==========================================
# 2. GESTOR DE MENÚ Y SALIDA
# ==========================================

def menu_principal():
    print("="*60)
    print("   GENERADOR DE PLANILLAS MAESTRAS 2025")
    print("   (CDC / RIESGOS / PMG)")
    print("="*60)
    
    archivo_maestro = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
    if not os.path.exists(archivo_maestro):
        print(f"[ERROR] No encuentro el archivo: {archivo_maestro}")
        return

    # --- PASO 1: SELECCIÓN DE FORMATO ---
    print("\n[PASO 1] ¿Qué tipo de planilla(s) deseas generar?")
    print("  1. Solo Datos Brutos (Planilla_Bruta_2025.xlsx)")
    print("  2. Solo Estilizada (Planilla_Estilizada_2025.xlsx)")
    print("  3. Ambas")
    opcion_formato = input("  >> Elige (1-3): ").strip()
    
    if opcion_formato not in ["1", "2", "3"]:
        print("Opción inválida. Saliendo.")
        return

    # --- PASO 2: SELECCIÓN DE HOJAS ---
    print("\n[PASO 2] ¿Qué hojas quieres incluir en la(s) planilla(s)?")
    # Diccionario de Hojas Disponibles (Nombre Excel : Etiqueta Salida)
    hojas_config = {
        "CDC 2025": "CDC",
        "PMG 2025": "PMG",
        "Riesgos 2025": "Riesgos"
    }
    
    hojas_seleccionadas = {}
    
    # Preguntamos por cada una para permitir cualquier combinación
    print("  Responde 's' para incluir o 'n' para omitir:")
    for hoja_excel, etiqueta in hojas_config.items():
        resp = input(f"  - ¿Incluir {etiqueta}? (s/n): ").lower().strip()
        if resp == 's':
            hojas_seleccionadas[hoja_excel] = etiqueta
            
    if not hojas_seleccionadas:
        print("[AVISO] No seleccionaste ninguna hoja. Fin del proceso.")
        return

    # --- PASO 3: PROCESAMIENTO Y EXTRACCIÓN ---
    print("\n" + "-"*40)
    print("   INICIANDO EXTRACCIÓN DE DATOS")
    print("-"*40)
    
    dataframes_listos = {}
    
    for hoja_excel, etiqueta in hojas_seleccionadas.items():
        df = obtener_dataframe_hoja(archivo_maestro, hoja_excel, etiqueta)
        if df is not None:
            dataframes_listos[etiqueta] = df # Guardamos con la etiqueta corta (CDC, PMG, etc.)

    if not dataframes_listos:
        print("[ERROR] No se pudo extraer información de ninguna hoja.")
        return

    # --- PASO 4: GENERACIÓN DE ARCHIVOS ---
    print("\n" + "-"*40)
    print("   GENERANDO ARCHIVOS FINALES")
    print("-"*40)

    archivos_a_generar = []
    if opcion_formato in ["1", "3"]: archivos_a_generar.append(("Planilla_Bruta_2025.xlsx", False))
    if opcion_formato in ["2", "3"]: archivos_a_generar.append(("Planilla_Estilizada_2025.xlsx", True))

    for nombre_archivo, aplicar_estilo in archivos_a_generar:
        print(f"Generando: {nombre_archivo}...")
        
        # Usamos ExcelWriter para crear múltiples pestañas
        try:
            with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
                for etiqueta, df in dataframes_listos.items():
                    df.to_excel(writer, sheet_name=etiqueta, index=False)
            
            print(f"   [OK] Archivo creado con {len(dataframes_listos)} pestañas.")
            
            if aplicar_estilo:
                aplicar_estilos_global(nombre_archivo)
                
        except Exception as e:
            print(f"   [ERROR] No se pudo crear {nombre_archivo}: {e}")

    print("\n" + "="*60)
    print("   ¡PROCESO COMPLETADO EXITOSAMENTE!")
    print("="*60)

if __name__ == "__main__":
    menu_principal()