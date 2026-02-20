import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v1.4.0 - FINAL RECOVERY
# =============================================================================

class IPSParserV140:
    def __init__(self, master_files):
        self.master_files = master_files if isinstance(master_files, list) else [master_files]
        self.output_file = "IPS_PARSER_v1.4.0_Output.xlsx"
        self.sheets_to_process = ["CDC 2025", "PMG 2025", "Riesgos 2025"]
        self.data_tree = {} # {Archivo: {Hoja: [Filas]}}
        self.flat_data = [] # Para la carga bruta
        self.new_indicator_count = 1

    def alert_user(self, message, critical=False):
        print(f"\n[ALERTA]: {message}")
        if critical:
            choice = input("¿Desea detener el proceso (d) o continuar ignorando el error (c)? ").lower()
            if choice == 'd': sys.exit()
            return False
        return True

    def transform_percentage(self, val, col_name):
        if pd.isna(val) or val == "" or val == "No aplica": return val
        try:
            num = float(val)
            if 0 < abs(num) <= 1:
                new_val = round(num * 100, 2)
                # print(f"[INFO]: {col_name} {num} -> {new_val}") # Reducir ruido en consola
                return new_val
            return num
        except:
            return val

    def process_all(self):
        for file_path in self.master_files:
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            
            for sheet in self.sheets_to_process:
                print(f"Procesando: {file_name} > {sheet}")
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except Exception as e:
                    print(f"  -> Saltando hoja (no encontrada o error): {e}")
                    continue

                # 1. Buscar Encabezado
                h_idx = None
                for idx, row in df.iterrows():
                    if "NÚMERO" in [str(x).upper().strip() for x in row.values if pd.notna(x)]:
                        h_idx = idx
                        break
                
                if h_idx is None:
                    self.alert_user(f"No se encontró 'NÚMERO' en {file_name} > {sheet}", critical=True)
                    continue

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                # 2. Mapeo de Columnas
                def find_c(names):
                    for i, h in enumerate(headers):
                        if any(n.lower() in h.lower().replace('\n', ' ') for n in names): return i
                    return None

                # Mapa base
                c_map = {
                    "num": find_c(["NÚMERO"]),
                    "prod": find_c(["PRODUCTO O PROCESO ESPECÍFICO"]),
                    "ind": find_c(["INDICADOR"]),
                    "form": find_c(["FORMULA"]),
                    "uni": find_c(["UNIDAD"]),
                    "resp": find_c(["RESPONSABLE CENTRO"]),
                    "gest": find_c(["GESTOR"]),
                    "sup": find_c(["SUPERVISORES"]),
                    "meta": find_c(["Meta 2025"]),
                    "pond": find_c(["Ponderador"]),
                    "op_desc": find_c(["Operandos"]), # K
                    "op_est": find_c(["Operandos Estimados Meta", "Operandos  Estimados"]), # L
                    "proy": find_c(["Cumplimiento Proyectado"]), # AI
                    "cump_meta": find_c(["% Cumplimiento de Meta"]), # AJ
                    "medios": find_c(["Medios de Verificación"]), # AK
                    "inst": find_c(["Instrumentos de Gestión"]) # AM
                }

                # Mapa de Meses (M - AH)
                months = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                          "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                          "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                month_map = {m: find_c([m]) for m in months}

                missing = [k for k, v in c_map.items() if v is None]
                if missing: self.alert_user(f"Faltan columnas en {sheet}: {missing}", critical=True)

                # 3. Extracción
                sheet_rows = []
                for i in range(h_idx + 1, len(df)):
                    def get_val(col_idx, row_offset=0):
                        if col_idx is None or (i + row_offset) >= len(df): return "No aplica"
                        val = df.iloc[i + row_offset, col_idx]
                        return val if pd.notna(val) else ""

                    raw_num = get_val(c_map["num"])
                    str_num = str(raw_num).strip()

                    # Detector de Indicador Nuevo o Basura
                    if str_num == "" or str_num.lower() == "nan" or "NUEVO" in str_num.upper():
                        # Si tiene descripción de indicador pero no número -> Es nuevo
                        ind_val = get_val(c_map["ind"])
                        if ind_val and str(ind_val).strip() != "":
                            final_code = f"INDICADOR_NUEVO_{self.new_indicator_count}_{sheet.split()[0]}"
                            print(f"  [AVISO] Asignando código auto: {final_code}")
                            self.new_indicator_count += 1
                        else:
                            continue # Fila vacía
                    else:
                        if not any(c.isdigit() for c in str_num): continue # Título intermedio o basura
                        final_code = str_num

                    # Construcción del Diccionario de Fila
                    row_data = {
                        # Metadata para Carga Bruta
                        "ARCHIVO": file_name,
                        "HOJA": sheet,
                        
                        # Datos Base
                        "NÚMERO": final_code,
                        "PRODUCTO": get_val(c_map["prod"]),
                        "INDICADOR": get_val(c_map["ind"]),
                        "FORMULA": get_val(c_map["form"]),
                        "UNIDAD": get_val(c_map["uni"]),
                        "RESPONSABLE": get_val(c_map["resp"]),
                        "GESTOR": get_val(c_map["gest"]),
                        "SUPERVISORES": get_val(c_map["sup"]),
                        "Meta 2025": self.transform_percentage(get_val(c_map["meta"]), "Meta 2025"),
                        "Ponderador": self.transform_percentage(get_val(c_map["pond"]), "Ponderador"),
                        
                        # Operandos (K, L)
                        "Desc Op 1": get_val(c_map["op_desc"], 0),
                        "Desc Op 2": get_val(c_map["op_desc"], 3),
                        "Meta Op 1": get_val(c_map["op_est"], 3),
                        "Meta Op 2": get_val(c_map["op_est"], 5),
                    }

                    # Datos Mensuales (Loop dinámico)
                    for m_name, m_idx in month_map.items():
                        row_data[f"{m_name} Op 1"] = get_val(m_idx, 3)
                        row_data[f"{m_name} Op 2"] = get_val(m_idx, 5)

                    # Cierre (AI, AJ, AK, AM)
                    row_data["Cump Proy Op 1"] = get_val(c_map["proy"], 3)
                    row_data["Cump Proy Op 2"] = get_val(c_map["proy"], 5)
                    row_data["% Cump Meta"] = self.transform_percentage(get_val(c_map["cump_meta"], 3), "% Cump Meta")
                    row_data["Medios Verif"] = get_val(c_map["medios"], 0)
                    row_data["Instrumentos"] = get_val(c_map["inst"], 0)

                    sheet_rows.append(row_data)
                    self.flat_data.append(row_data)

                self.data_tree[file_name][sheet] = sheet_rows

    def export_excel(self):
        wb = Workbook()
        
        # --- HOJA 1: CARGA BRUTA ---
        ws_bruta = wb.active
        ws_bruta.title = "Carga Bruta"
        
        if self.flat_data:
            # Encabezados
            headers = list(self.flat_data[0].keys())
            ws_bruta.append(headers)
            # Datos
            for row in self.flat_data:
                ws_bruta.append(list(row.values()))
        
        # --- HOJA 2: PLANILLA ESTILIZADA ---
        ws_style = wb.create_sheet("Planilla Estilizada")
        
        # Estilos
        file_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Negro
        sheet_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Azul Oscuro
        header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid") # Gris
        white_font = Font(color="FFFFFF", bold=True, size=11)
        black_bold = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        current_row = 1
        max_col_idx = 0

        for file_name, sheets in self.data_tree.items():
            # SEPARADOR ARCHIVO
            ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
            c = ws_style.cell(row=current_row, column=1, value=f"ARCHIVO: {file_name}")
            c.fill = file_fill
            c.font = white_font
            c.alignment = Alignment(horizontal='center')
            current_row += 1

            for sheet_name, rows in sheets.items():
                # SEPARADOR HOJA
                ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
                c = ws_style.cell(row=current_row, column=1, value=f"HOJA: {sheet_name}")
                c.fill = sheet_fill
                c.font = white_font
                current_row += 1

                if not rows: continue

                # Filtrar columnas de metadata para la vista estilizada (opcional, pero pedido por usuario)
                # El usuario dijo: "primera fila indicando de que planilla...". 
                # Ya lo hicimos con los separadores. Quitamos ARCHIVO y HOJA de las columnas de datos visuales.
                keys_to_show = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                max_col_idx = max(max_col_idx, len(keys_to_show))

                # Encabezados Tabla
                for idx, k in enumerate(keys_to_show, 1):
                    c = ws_style.cell(row=current_row, column=idx, value=k)
                    c.fill = header_fill
                    c.font = black_bold
                    c.border = thin_border
                current_row += 1

                # Datos
                for row_dict in rows:
                    for idx, k in enumerate(keys_to_show, 1):
                        val = row_dict[k]
                        c = ws_style.cell(row=current_row, column=idx, value=val)
                        c.border = thin_border
                        c.alignment = Alignment(wrap_text=True, vertical='top')
                    current_row += 1
                current_row += 1

        # Ajuste Ancho
        for i in range(1, max_col_idx + 2):
            col_let = get_column_letter(i)
            ws_style.column_dimensions[col_let].width = 20

        wb.save(self.output_file)
        print(f"\n[EXITO] Archivo generado: {self.output_file}")

if __name__ == "__main__":
    archivo = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
    if os.path.exists(archivo):
        parser = IPSParserV140(archivo)
        parser.process_all()
        parser.export_excel()
    else:
        print(f"No existe el archivo: {archivo}")