import pandas as pd
import os
import re
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Silenciar alertas
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

ARCHIVO_INPUT = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
ARCHIVO_OUTPUT = "1_PLANILLA_SIG_CONSOLIDADO_2025.xlsx"

HOJAS = {
    "CDC 2025": "CDC 2025", 
    "PMG 2025": "PMG 2025", 
    "Riesgos 2025": "RIESGOS 2025"
}

# ORDEN ESTRICTO DE COLUMNAS (Basado en la planilla original)
COLUMNAS = [
    "NÚMERO", 
    "PRODUCTO O PROCESO ESPECÍFICO", 
    "INDICADOR", 
    "FORMULA", 
    "UNIDAD", 
    "RESPONSABLE CENTRO DE RESPONSABILIDAD", 
    "GESTOR", 
    "SUPERVISORES", 
    "Meta 2025 (%)", 
    "Ponderador (%)",
    # Bloque Operandos (Aplanado)
    "Desc. Op1", 
    "Est. Meta Op1",
    "Desc. Op2", 
    "Est. Meta Op2",
    # Bloque Meses (Aplanado)
    "Ene Ind (%)", "Ene Op1", "Ene Op2",
    "Feb Ind (%)", "Feb Op1", "Feb Op2",
    "Mar Ind (%)", "Mar Op1", "Mar Op2",
    "Abr Ind (%)", "Abr Op1", "Abr Op2",
    "May Ind (%)", "May Op1", "May Op2",
    "Jun Ind (%)", "Jun Op1", "Jun Op2",
    "Jul Ind (%)", "Jul Op1", "Jul Op2",
    "Ago Ind (%)", "Ago Op1", "Ago Op2",
    "Sept Ind (%)", "Sept Op1", "Sept Op2",
    "Oct Ind (%)", "Oct Op1", "Oct Op2",
    "Nov Ind (%)", "Nov Op1", "Nov Op2",
    "Dic Ind (%)", "Dic Op1", "Dic Op2",
    # Bloque Final
    "Medios de Verificación", 
    "Control de Cambios", 
    "Instrumentos de Gestión Asociados"
]

# =============================================================================
# UTILS
# =============================================================================

def limpiar_num(val):
    if pd.isna(val) or str(val).strip() == "": return 0
    s = str(val).strip()
    if "," in s: s = s.replace('.', '').replace(',', '.')
    else: s = s.replace('.', '')
    try: return float(s)
    except: return 0

def limpiar_porc(val):
    v = limpiar_num(val)
    if 0 < v < 1.05: return v * 100
    return v

def limpiar_txt(val):
    if pd.isna(val): return ""
    return str(val).strip().replace("\n", " ").replace("\r", " ")

def buscar_header(df):
    for i in range(min(30, len(df))):
        try:
            fila = [str(x).upper() for x in df.iloc[i].tolist()]
            if "INDICADOR" in str(fila) and ("NÚMERO" in str(fila) or "CODIGO" in str(fila)):
                return i, df.iloc[i].astype(str).tolist()
        except: continue
    return None, None

# =============================================================================
# MOTOR
# =============================================================================

def procesar_hoja(ruta, hoja, etiqueta):
    print(f"-> Procesando: {hoja}...")
    try:
        df = pd.read_excel(ruta, sheet_name=hoja, header=None)
    except:
        print(f"   [ERROR] No se encontró la hoja {hoja}")
        return pd.DataFrame()

    idx_h, heads = buscar_header(df)
    if idx_h is None: return pd.DataFrame()

    mapa = {str(h).upper().strip(): i for i, h in enumerate(heads)}
    
    def get_col(kw):
        for k, v in mapa.items():
            if kw in k: return v
        return None

    # Indices Base (Mapeo directo)
    I_NUM = get_col("NÚMERO"); I_PROD = get_col("PRODUCTO"); I_IND = get_col("INDICADOR")
    I_FORM = get_col("FORMULA"); I_UNIDAD = get_col("UNIDAD"); I_RESP = get_col("RESPONSABLE")
    I_GEST = get_col("GESTOR"); I_SUP = get_col("SUPERVISORES")
    I_META = get_col("META 2025"); I_POND = get_col("PONDERADOR")
    I_OP_DESC = get_col("OPERANDOS")
    I_OP_EST = get_col("ESTIMADOS") 
    if I_OP_EST is None: I_OP_EST = get_col("META") # Fallback para Riesgos/PMG a veces

    I_MEDIOS = get_col("MEDIOS"); I_CONTROL = get_col("CONTROL")
    I_INST = get_col("INSTRUMENTOS")

    # Meses
    meses = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEPT", "OCT", "NOV", "DIC"]
    mapa_meses = {}
    for m in meses:
        c = get_col(f"{m}.")
        if c is None: c = get_col(m)
        if c is not None: mapa_meses[m.title()] = c

    filas = []
    # Fila Separadora
    sep = {c: None for c in COLUMNAS}
    sep["NÚMERO"] = f"--- {etiqueta} ---"
    filas.append(sep)

    for i in range(idx_h + 1, len(df) - 5):
        val_num = str(df.iloc[i, I_NUM]).strip()
        if pd.notna(df.iloc[i, I_NUM]) and val_num != "NÚMERO" and len(val_num) >= 3:
            try:
                d = {}
                # Bloque 1: Identificación (Tal cual la planilla)
                d["NÚMERO"] = val_num
                d["PRODUCTO O PROCESO ESPECÍFICO"] = limpiar_txt(df.iloc[i, I_PROD]) if I_PROD else ""
                d["INDICADOR"] = limpiar_txt(df.iloc[i, I_IND])
                d["FORMULA"] = limpiar_txt(df.iloc[i, I_FORM]) if I_FORM else ""
                d["UNIDAD"] = limpiar_txt(df.iloc[i, I_UNIDAD]) if I_UNIDAD else ""
                d["RESPONSABLE CENTRO DE RESPONSABILIDAD"] = limpiar_txt(df.iloc[i, I_RESP]) if I_RESP else ""
                d["GESTOR"] = limpiar_txt(df.iloc[i, I_GEST]) if I_GEST else ""
                d["SUPERVISORES"] = limpiar_txt(df.iloc[i, I_SUP]) if I_SUP else ""
                
                # Bloque 2: Metas
                if I_META: d["Meta 2025 (%)"] = limpiar_porc(df.iloc[i, I_META])
                else: d["Meta 2025 (%)"] = limpiar_porc(df.iloc[i+1, I_OP_EST]) # Fallback

                d["Ponderador (%)"] = limpiar_porc(df.iloc[i, I_POND]) if I_POND else 0

                # Bloque 3: Operandos (Aplanado)
                d["Desc. Op1"] = limpiar_txt(df.iloc[i, I_OP_DESC]).lstrip('(')
                d["Est. Meta Op1"] = limpiar_num(df.iloc[i+3, I_OP_EST])
                
                d["Desc. Op2"] = limpiar_txt(df.iloc[i+3, I_OP_DESC]).split(')')[0]
                d["Est. Meta Op2"] = limpiar_num(df.iloc[i+5, I_OP_EST])

                # Bloque 4: Meses
                for m_nom, m_idx in mapa_meses.items():
                    d[f"{m_nom} Ind (%)"] = limpiar_porc(df.iloc[i+1, m_idx])
                    d[f"{m_nom} Op1"] = limpiar_num(df.iloc[i+3, m_idx])
                    d[f"{m_nom} Op2"] = limpiar_num(df.iloc[i+5, m_idx])

                # Bloque 5: Final
                d["Medios de Verificación"] = limpiar_txt(df.iloc[i, I_MEDIOS]) if I_MEDIOS else ""
                d["Control de Cambios"] = limpiar_txt(df.iloc[i, I_CONTROL]) if I_CONTROL else ""
                d["Instrumentos de Gestión Asociados"] = limpiar_txt(df.iloc[i, I_INST]) if I_INST else ""

                filas.append(d)
            except: continue

    return pd.DataFrame(filas)

def estilizar(ruta):
    wb = load_workbook(ruta)
    ws = wb["DATOS_ESTILIZADOS"]
    
    # Estilos
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(b=True, color="FFFFFF")
    sep_fill = PatternFill("solid", fgColor="D9D9D9")
    sep_font = Font(b=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Cabecera
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    # Filas
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).startswith("---"):
            for c in row:
                c.fill = sep_fill
                c.font = sep_font
        else:
            for c in row:
                c.border = border
                c.alignment = Alignment(vertical='top', wrap_text=False)

    # Anchos críticos
    ws.column_dimensions['A'].width = 12 # Numero
    ws.column_dimensions['B'].width = 30 # Producto
    ws.column_dimensions['C'].width = 50 # Indicador
    ws.column_dimensions['D'].width = 40 # Formula
    ws.column_dimensions['K'].width = 40 # Desc Op1
    ws.column_dimensions['M'].width = 40 # Desc Op2
    ws.column_dimensions['AW'].width = 50 # Medios

    ws.freeze_panes = "D2"
    wb.save(ruta)

# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    if not os.path.exists(ARCHIVO_INPUT):
        print("ERROR: No está el archivo de entrada.")
        exit()

    dfs = []
    for hoja, etiqueta in HOJAS.items():
        res = procesar_hoja(ARCHIVO_INPUT, hoja, etiqueta)
        if not res.empty:
            dfs.append(res)

    if dfs:
        full = pd.concat(dfs, ignore_index=True)
        # Asegurar orden columnas
        cols_finales = [c for c in COLUMNAS if c in full.columns]
        full = full[cols_finales]

        print(f"Guardando {ARCHIVO_OUTPUT}...")
        with pd.ExcelWriter(ARCHIVO_OUTPUT, engine='openpyxl') as w:
            full.to_excel(w, sheet_name="DATOS_BRUTOS", index=False)
            full.to_excel(w, sheet_name="DATOS_ESTILIZADOS", index=False)
        
        estilizar(ARCHIVO_OUTPUT)
        print("¡Listo!")
    else:
        print("No se extrajeron datos.")