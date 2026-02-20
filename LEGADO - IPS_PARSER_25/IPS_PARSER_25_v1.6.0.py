import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v1.6.0 - INCLUYE CONTROL DE CAMBIOS
# =============================================================================

class IPSParserV160:
    def __init__(self, master_files):
        self.master_files = master_files if isinstance(master_files, list) else [master_files]
        self.output_file = "IPS_PARSER_v1.6.0_Output.xlsx"
        self.sheets_to_process = ["CDC 2025", "PMG 2025", "Riesgos 2025"]
        self.data_tree = {} # {Archivo: {Hoja: [Filas]}}
        self.flat_data = [] # Para la carga bruta
        self.new_indicator_count = 1

    def alert_user(self, message, critical=False):
        """Manejo de errores interactivo con el usuario."""
        print(f"\n[ALERTA]: {message}")
        if critical:
            choice = input("¿Desea detener el proceso (d) o continuar ignorando el error (c)? ").lower()
            if choice == 'd':
                print("Proceso detenido por el usuario.")
                sys.exit()
            print(" -> Continuando proceso bajo riesgo del usuario...")
            return False
        return True

    def transform_percentage(self, val, col_name):
        """Transforma decimales (0.2) a enteros (20.0) y avisa en consola."""
        if pd.isna(val) or val == "" or val == "No aplica": return val
        try:
            num = float(val)
            # Si es decimal entre 0 y 1, asumimos porcentaje
            if 0 < abs(num) <= 1:
                new_val = round(num * 100, 2)
                print(f"    [TRANSFORMACIÓN] {col_name}: {val} -> {new_val}") 
                return new_val
            return num
        except:
            return val

    def process_all(self):
        for file_path in self.master_files:
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            
            print(f"\n{'='*60}")
            print(f"PROCESANDO ARCHIVO: {file_name}")
            print(f"{'='*60}")
            
            for sheet in self.sheets_to_process:
                print(f"\n>>> Analizando Hoja: {sheet}")
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except Exception as e:
                    print(f"  [ERROR] No se pudo leer la hoja (¿Existe?): {e}")
                    continue

                # 1. Buscar Encabezado (Fila con "NÚMERO")
                h_idx = None
                for idx, row in df.iterrows():
                    if "NÚMERO" in [str(x).upper().strip() for x in row.values if pd.notna(x)]:
                        h_idx = idx
                        break
                
                if h_idx is None:
                    self.alert_user(f"No se encontró la columna 'NÚMERO' en {sheet}", critical=True)
                    continue
                else:
                    print(f"  [INFO] Encabezados detectados en fila {h_idx + 1}")

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                # 2. Mapeo de Columnas Dinámico
                def find_c(names):
                    for i, h in enumerate(headers):
                        if any(n.lower() in h.lower().replace('\n', ' ') for n in names): return i
                    return None

                c_map = {
                    "num": find_c(["NÚMERO"]),
                    "prod": find_c(["PRODUCTO O PROCESO ESPECÍFICO"]),
                    "ind": find_c(["INDICADOR"]),
                    "form": find_c(["FORMULA"]),
                    "uni": find_c(["UNIDAD"]),
                    "resp": find_c(["RESPONSABLE CENTRO"]),
                    "gest": find_c(["GESTOR"]),
                    "sup": find_c(["SUPERVISORES"]),
                    "meta": find_c(["Meta 2025"]),
                    "pond": find_c(["Ponderador"]),
                    "op_desc": find_c(["Operandos"]), 
                    "op_est": find_c(["Operandos Estimados Meta", "Operandos  Estimados"]),
                    "proy": find_c(["Cumplimiento Proyectado"]),
                    "cump_meta": find_c(["% Cumplimiento de Meta"]),
                    "medios": find_c(["Medios de Verificación"]),
                    "control": find_c(["Control de Cambios"]), # <--- AGREGADO AQUÍ
                    "inst": find_c(["Instrumentos de Gestión"])
                }

                # Validar columnas críticas
                missing = [k for k, v in c_map.items() if v is None]
                if missing:
                    self.alert_user(f"Faltan columnas en {sheet}: {missing}", critical=True)
                else:
                    print(f"  [INFO] Estructura de columnas completa.")

                # Meses
                months = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                          "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                          "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                month_map = {m: find_c([m]) for m in months}

                # 3. Extracción de Datos
                print("  [INFO] Extrayendo filas...")
                sheet_rows = []
                count_rows = 0
                
                for i in range(h_idx + 1, len(df)):
                    # Función interna para obtener valor seguro
                    def get_val(col_idx, row_offset=0):
                        if col_idx is None or (i + row_offset) >= len(df): return "No aplica"
                        val = df.iloc[i + row_offset, col_idx]
                        return val if pd.notna(val) else ""

                    raw_num = get_val(c_map["num"])
                    str_num = str(raw_num).strip()

                    # Lógica para detectar si es un Indicador válido o "NUEVO"
                    if str_num == "" or str_num.lower() == "nan" or "NUEVO" in str_num.upper():
                        ind_val = get_val(c_map["ind"])
                        # Si tiene texto en indicador pero no tiene número, es un indicador nuevo
                        if ind_val and str(ind_val).strip() != "":
                            final_code = f"INDICADOR_NUEVO_{self.new_indicator_count}_{sheet.split()[0]}"
                            print(f"    [AVISO] Fila {i+1}: Sin código numérico. Asignando: {final_code}")
                            self.new_indicator_count += 1
                        else:
                            continue # Fila vacía
                    else:
                        if not any(c.isdigit() for c in str_num): continue # Título o basura
                        final_code = str_num

                    count_rows += 1
                    
                    # Construcción del registro
                    row_data = {
                        "ARCHIVO": file_name,
                        "HOJA": sheet,
                        "NÚMERO": final_code,
                        "PRODUCTO": get_val(c_map["prod"]),
                        "INDICADOR": get_val(c_map["ind"]),
                        "FORMULA": get_val(c_map["form"]),
                        "UNIDAD": get_val(c_map["uni"]),
                        "RESPONSABLE": get_val(c_map["resp"]),
                        "GESTOR": get_val(c_map["gest"]),
                        "SUPERVISORES": get_val(c_map["sup"]),
                        "Meta 2025": self.transform_percentage(get_val(c_map["meta"]), "Meta 2025"),
                        "Ponderador": self.transform_percentage(get_val(c_map["pond"]), "Ponderador"),
                        "Desc Op 1": get_val(c_map["op_desc"], 0),
                        "Desc Op 2": get_val(c_map["op_desc"], 3),
                        "Meta Op 1": get_val(c_map["op_est"], 3),
                        "Meta Op 2": get_val(c_map["op_est"], 5),
                    }

                    # Meses dinámicos
                    for m_name, m_idx in month_map.items():
                        row_data[f"{m_name} Op 1"] = get_val(m_idx, 3)
                        row_data[f"{m_name} Op 2"] = get_val(m_idx, 5)

                    # Columnas finales
                    row_data["Cump Proy Op 1"] = get_val(c_map["proy"], 3)
                    row_data["Cump Proy Op 2"] = get_val(c_map["proy"], 5)
                    row_data["% Cump Meta"] = self.transform_percentage(get_val(c_map["cump_meta"], 3), "% Cump Meta")
                    row_data["Medios Verif"] = get_val(c_map["medios"], 0)
                    row_data["Control Cambios"] = get_val(c_map["control"], 0) # <--- AGREGADO AQUÍ
                    row_data["Instrumentos"] = get_val(c_map["inst"], 0)

                    sheet_rows.append(row_data)
                    self.flat_data.append(row_data)

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  [FIN HOJA] Se procesaron {count_rows} indicadores.")

    def export_excel(self):
        print(f"\n{'='*60}")
        print("GENERANDO ARCHIVO EXCEL DE SALIDA...")
        print(f"{'='*60}")
        
        wb = Workbook()
        
        # --- HOJA 1: CARGA BRUTA ---
        print("  -> Creando hoja 'Carga Bruta'...")
        ws_bruta = wb.active
        ws_bruta.title = "Carga Bruta"
        
        if self.flat_data:
            headers = list(self.flat_data[0].keys())
            ws_bruta.append(headers)
            for row in self.flat_data:
                ws_bruta.append(list(row.values()))
        
        # --- HOJA 2: ESTILIZADA ---
        print("  -> Creando hoja 'Planilla Estilizada'...")
        ws_style = wb.create_sheet("Planilla Estilizada")
        
        # Configuración de Estilos
        file_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        sheet_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=11)
        black_bold = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        current_row = 1
        max_col_idx = 0

        for file_name, sheets in self.data_tree.items():
            # SEPARADOR: NOMBRE DEL ARCHIVO
            ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
            c = ws_style.cell(row=current_row, column=1, value=f"ARCHIVO: {file_name}")
            c.fill = file_fill
            c.font = white_font
            c.alignment = Alignment(horizontal='center')
            current_row += 1

            for sheet_name, rows in sheets.items():
                # SEPARADOR: NOMBRE DE LA HOJA
                ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
                c = ws_style.cell(row=current_row, column=1, value=f"HOJA: {sheet_name}")
                c.fill = sheet_fill
                c.font = white_font
                current_row += 1

                if not rows: continue

                # Ocultar metadata (ARCHIVO, HOJA) en la vista estilizada
                keys_to_show = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                max_col_idx = max(max_col_idx, len(keys_to_show))

                # Encabezados
                for idx, k in enumerate(keys_to_show, 1):
                    c = ws_style.cell(row=current_row, column=idx, value=k)
                    c.fill = header_fill
                    c.font = black_bold
                    c.border = thin_border
                current_row += 1

                # Datos
                for row_dict in rows:
                    for idx, k in enumerate(keys_to_show, 1):
                        val = row_dict[k]
                        c = ws_style.cell(row=current_row, column=idx, value=val)
                        c.border = thin_border
                        c.alignment = Alignment(wrap_text=True, vertical='top')
                    current_row += 1
                current_row += 1 # Espacio entre tablas

        # Ajuste de ancho de columnas
        for i in range(1, max_col_idx + 2):
            col_let = get_column_letter(i)
            ws_style.column_dimensions[col_let].width = 20

        wb.save(self.output_file)
        print(f"\n[EXITO FINAL] Archivo generado: {self.output_file}")

if __name__ == "__main__":
    # Nombre del archivo a procesar
    archivo = "Proyecciones Indicadores 2025 - División Planificación (1).xlsx"
    
    if os.path.exists(archivo):
        parser = IPSParserV160(archivo)
        parser.process_all()
        parser.export_excel()
    else:
        print(f"[ERROR CRÍTICO] No existe el archivo: {archivo}")