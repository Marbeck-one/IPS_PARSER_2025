import pandas as pd
import os
import sys
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v2.0.0 - CONTROL TOTAL (Hojas y Columnas)
# =============================================================================

class IPSParserV200:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_UNIVERSAL_v2.0.xlsx")
        self.data_tree = {} # {NombreArchivo: {NombreHoja: [Filas]}}
        self.flat_data = [] 
        self.new_indicator_count = 1

    def ask_user_action(self, context_msg):
        """
        Manejo interactivo cuando falta la columna NÚMERO.
        """
        print(f"\n[ATENCIÓN REQUERIDA] {context_msg}")
        print("  Opciones:")
        print("   [s] Saltar esta hoja (No procesar).")
        print("   [d] Detener TODO el proceso.")
        
        while True:
            choice = input("  >> Elija opción (s/d): ").lower().strip()
            if choice == 's':
                print("  -> Saltando hoja...")
                return 'skip'
            elif choice == 'd':
                print("  -> Proceso detenido por el usuario.")
                sys.exit()
            else:
                print("  Opción no válida.")

    def ask_column_action(self, missing_cols, sheet_name):
        """
        Manejo interactivo cuando FALTAN COLUMNAS de datos.
        """
        print(f"\n[FALTAN DATOS] En la hoja '{sheet_name}' no se encontraron las columnas: {missing_cols}")
        print("  Opciones:")
        print("   [c] Continuar (Rellenar estas columnas con 'No aplica').")
        print("   [s] Saltar esta hoja (No incluirla en el reporte).")
        print("   [d] Detener TODO el proceso.")
        
        while True:
            choice = input("  >> Elija opción (c/s/d): ").lower().strip()
            if choice == 'c':
                print("  -> Continuando (rellenando vacíos)...")
                return 'continue'
            elif choice == 's':
                print("  -> Saltando hoja...")
                return 'skip'
            elif choice == 'd':
                print("  -> Proceso detenido por el usuario.")
                sys.exit()
            else:
                print("  Opción no válida.")

    def transform_percentage(self, val, col_name):
        if pd.isna(val) or val == "" or val == "No aplica": return val
        try:
            num = float(val)
            if 0 < abs(num) <= 1:
                new_val = round(num * 100, 2)
                return new_val
            return num
        except:
            return val

    def get_excel_files(self):
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$") and "IPS_CONSOLIDADO" not in f]
        
        if not valid_files:
            print(f"[ERROR] No se encontraron archivos Excel válidos en: {self.folder_path}")
            sys.exit()
        
        print(f"\n[INFO] Se encontraron {len(valid_files)} archivos para procesar.")
        return valid_files

    def process_folder(self):
        files = self.get_excel_files()
        
        for file_path in files:
            file_name = os.path.basename(file_path)
            self.data_tree[file_name] = {}
            
            print(f"\n{'='*60}")
            print(f"PROCESANDO ARCHIVO: {file_name}")
            print(f"{'='*60}")
            
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except Exception as e:
                print(f"  [ERROR] No se pudo abrir el archivo: {e}")
                continue

            for sheet in sheet_names:
                print(f"\n>>> Analizando Hoja: '{sheet}'")
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except Exception as e:
                    print(f"  [ERROR] Lectura fallida: {e}")
                    continue

                # 1. BUSCAR ENCABEZADO "NÚMERO"
                h_idx = None
                for idx, row in df.iterrows():
                    row_vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    if any(x in row_vals for x in ["NÚMERO", "NUMERO", "N°"]):
                        h_idx = idx
                        break
                
                if h_idx is None:
                    action = self.ask_user_action(f"No se encontró la columna 'NÚMERO' en '{sheet}'.")
                    if action == 'skip': continue
                
                print(f"  [INFO] Encabezados detectados en fila {h_idx + 1}.")
                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                # 2. MAPEO DE COLUMNAS
                def find_c(names):
                    for i, h in enumerate(headers):
                        if any(n.lower() in h.lower().replace('\n', ' ') for n in names): return i
                    return None

                c_map = {
                    "num": find_c(["NÚMERO", "NUMERO", "N°"]),
                    "prod": find_c(["PRODUCTO"]),
                    "ind": find_c(["INDICADOR"]),
                    "form": find_c(["FORMULA", "FÓRMULA"]),
                    "uni": find_c(["UNIDAD"]),
                    "resp": find_c(["RESPONSABLE"]),
                    "gest": find_c(["GESTOR"]),
                    "sup": find_c(["SUPERVISORES"]),
                    "meta": find_c(["Meta 2025", "Meta 2026", "Meta"]),
                    "pond": find_c(["Ponderador"]),
                    "op_desc": find_c(["Operandos"]), 
                    "op_est": find_c(["Operandos Estimados"]),
                    "proy": find_c(["Cumplimiento Proyectado"]),
                    "cump_meta": find_c(["% Cumplimiento"]),
                    "medios": find_c(["Medios"]),
                    "control": find_c(["Control de Cambios"]),
                    "inst": find_c(["Instrumentos"])
                }

                # 3. VERIFICAR COLUMNAS FALTANTES E INTERVENIR
                # Excluimos 'pond' y 'control' de ser críticas, pero incluimos 'op_est', 'proy', etc.
                missing = [k for k, v in c_map.items() if v is None and k not in ["pond", "control"]]
                
                if missing:
                    action = self.ask_column_action(missing, sheet)
                    if action == 'skip': continue
                    # Si action == 'continue', el código sigue y get_val rellenará con "No aplica"
                
                # 4. EXTRACCIÓN
                months = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                          "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                          "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                month_map = {m: find_c([m]) for m in months}

                sheet_rows = []
                count_rows = 0
                
                for i in range(h_idx + 1, len(df)):
                    def get_val(col_idx, row_offset=0):
                        if col_idx is None or (i + row_offset) >= len(df): return "No aplica"
                        val = df.iloc[i + row_offset, col_idx]
                        return val if pd.notna(val) else ""

                    raw_num = get_val(c_map["num"])
                    str_num = str(raw_num).strip()

                    # Detector de Indicadores (incluye Nuevos)
                    if str_num == "" or str_num.lower() == "nan" or "NUEVO" in str_num.upper():
                        ind_val = get_val(c_map["ind"])
                        if ind_val and str(ind_val).strip() not in ["", "0", "No aplica"]:
                            prefix_file = file_name.split()[0][:8]
                            clean_sheet = ''.join(e for e in sheet if e.isalnum())
                            final_code = f"NUEVO_{self.new_indicator_count}_{prefix_file}_{clean_sheet}"
                            print(f"    [AVISO] Fila {i+1}: Asignando código auto: {final_code}")
                            self.new_indicator_count += 1
                        else:
                            continue 
                    else:
                        if not any(c.isdigit() for c in str_num): continue 
                        final_code = str_num

                    count_rows += 1
                    
                    row_data = {
                        "ARCHIVO": file_name,
                        "HOJA": sheet,
                        "NÚMERO": final_code,
                        "PRODUCTO": get_val(c_map["prod"]),
                        "INDICADOR": get_val(c_map["ind"]),
                        "FORMULA": get_val(c_map["form"]),
                        "UNIDAD": get_val(c_map["uni"]),
                        "RESPONSABLE": get_val(c_map["resp"]),
                        "GESTOR": get_val(c_map["gest"]),
                        "SUPERVISORES": get_val(c_map["sup"]),
                        "Meta": self.transform_percentage(get_val(c_map["meta"]), "Meta"),
                        "Ponderador": self.transform_percentage(get_val(c_map["pond"]), "Ponderador"),
                        "Desc Op 1": get_val(c_map["op_desc"], 0),
                        "Desc Op 2": get_val(c_map["op_desc"], 3),
                        "Meta Op 1": get_val(c_map["op_est"], 3),
                        "Meta Op 2": get_val(c_map["op_est"], 5),
                    }

                    for m_name, m_idx in month_map.items():
                        row_data[f"{m_name} Op 1"] = get_val(m_idx, 3)
                        row_data[f"{m_name} Op 2"] = get_val(m_idx, 5)

                    row_data["Cump Proy Op 1"] = get_val(c_map["proy"], 3)
                    row_data["Cump Proy Op 2"] = get_val(c_map["proy"], 5)
                    row_data["% Cump Meta"] = self.transform_percentage(get_val(c_map["cump_meta"], 3), "% Cump Meta")
                    row_data["Medios Verif"] = get_val(c_map["medios"], 0)
                    row_data["Control Cambios"] = get_val(c_map["control"], 0)
                    row_data["Instrumentos"] = get_val(c_map["inst"], 0)

                    sheet_rows.append(row_data)
                    self.flat_data.append(row_data)

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  [OK] {count_rows} registros extraídos.")

    def export_excel(self):
        if not self.flat_data:
            print("\n[AVISO FINAL] No se generaron datos para exportar.")
            return

        print(f"\n{'='*60}")
        print("CONSOLIDANDO DATOS...")
        print(f"{'='*60}")
        
        wb = Workbook()
        
        # HOJA 1: BRUTA
        ws_bruta = wb.active
        ws_bruta.title = "Carga Bruta"
        headers = list(self.flat_data[0].keys())
        ws_bruta.append(headers)
        for row in self.flat_data:
            ws_bruta.append(list(row.values()))
        
        # HOJA 2: ESTILIZADA
        ws_style = wb.create_sheet("Planilla Estilizada")
        
        file_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        sheet_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=11)
        black_bold = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        current_row = 1
        max_col_idx = 0

        # Mantenemos orden: Archivo -> Hoja
        for file_name, sheets in self.data_tree.items():
            if not any(sheets.values()): continue

            # SEPARADOR ARCHIVO
            ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
            c = ws_style.cell(row=current_row, column=1, value=f"ARCHIVO: {file_name}")
            c.fill = file_fill
            c.font = white_font
            c.alignment = Alignment(horizontal='center')
            current_row += 1

            for sheet_name, rows in sheets.items():
                if not rows: continue

                # SEPARADOR HOJA
                ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
                c = ws_style.cell(row=current_row, column=1, value=f"PLANILLA: {sheet_name}")
                c.fill = sheet_fill
                c.font = white_font
                current_row += 1

                # Ocultar columnas de metadata en la visualización
                keys_to_show = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                max_col_idx = max(max_col_idx, len(keys_to_show))

                # Encabezados
                for idx, k in enumerate(keys_to_show, 1):
                    c = ws_style.cell(row=current_row, column=idx, value=k)
                    c.fill = header_fill
                    c.font = black_bold
                    c.border = thin_border
                current_row += 1

                # Datos
                for row_dict in rows:
                    for idx, k in enumerate(keys_to_show, 1):
                        val = row_dict[k]
                        c = ws_style.cell(row=current_row, column=idx, value=val)
                        c.border = thin_border
                        c.alignment = Alignment(wrap_text=True, vertical='top')
                    current_row += 1
                current_row += 1

        for i in range(1, max_col_idx + 2):
            col_let = get_column_letter(i)
            ws_style.column_dimensions[col_let].width = 20

        wb.save(self.output_file)
        print(f"\n[ÉXITO] Archivo generado: {self.output_file}")

if __name__ == "__main__":
    try:
        print("--- IPS PARSER v2.0.0 (Control Total) ---")
        path = input("Ruta de la carpeta (Enter para actual): ").strip()
        if not path: path = os.getcwd()
        
        if os.path.isdir(path):
            parser = IPSParserV200(path)
            parser.process_folder()
            parser.export_excel()
        else:
            print("[ERROR] Ruta no válida.")
    except Exception as e:
        print(f"Error fatal: {e}")
        input("Enter para salir.")