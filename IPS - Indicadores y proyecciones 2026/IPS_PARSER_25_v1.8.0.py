import pandas as pd
import os
import sys
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# IPS_PARSER_v1.8.0 - MODO UNIVERSAL (Detección Automática de Hojas)
# =============================================================================

class IPSParserV180:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.output_file = os.path.join(folder_path, "IPS_CONSOLIDADO_UNIVERSAL.xlsx")
        self.data_tree = {} # {NombreArchivo: {NombreHoja: [Filas]}}
        self.flat_data = [] 
        self.new_indicator_count = 1

    def alert_user(self, message, critical=False):
        """Manejo de errores interactivo."""
        print(f"\n[ALERTA]: {message}")
        if critical:
            choice = input("¿Desea detener el proceso (d) o continuar ignorando el error (c)? ").lower()
            if choice == 'd':
                print("Proceso detenido por el usuario.")
                sys.exit()
            print(" -> Continuando proceso bajo riesgo del usuario...")
            return False
        return True

    def transform_percentage(self, val, col_name):
        """Transforma decimales (0.2) a enteros (20.0)."""
        if pd.isna(val) or val == "" or val == "No aplica": return val
        try:
            num = float(val)
            if 0 < abs(num) <= 1:
                new_val = round(num * 100, 2)
                # print(f"    [TRANSFORMACIÓN] {col_name}: {val} -> {new_val}") 
                return new_val
            return num
        except:
            return val

    def get_excel_files(self):
        """Obtiene lista de archivos Excel ignorando temporales."""
        all_files = glob.glob(os.path.join(self.folder_path, "*.xlsx")) + glob.glob(os.path.join(self.folder_path, "*.xls"))
        valid_files = [f for f in all_files if not os.path.basename(f).startswith("~$")]
        
        if not valid_files:
            print(f"[ERROR] No se encontraron archivos Excel en: {self.folder_path}")
            sys.exit()
        
        print(f"\n[INFO] Se encontraron {len(valid_files)} archivos para procesar.")
        return valid_files

    def process_folder(self):
        files = self.get_excel_files()
        
        for file_path in files:
            file_name = os.path.basename(file_path)
            if "IPS_CONSOLIDADO" in file_name: continue # Ignorar archivo de salida propio

            self.data_tree[file_name] = {}
            
            print(f"\n{'='*60}")
            print(f"PROCESANDO ARCHIVO: {file_name}")
            print(f"{'='*60}")
            
            try:
                # Cargar el archivo Excel para inspeccionar nombres de hojas
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
            except Exception as e:
                print(f"  [ERROR] No se pudo abrir el archivo Excel: {e}")
                continue

            # ITERAR SOBRE TODAS LAS HOJAS DISPONIBLES
            for sheet in sheet_names:
                print(f"\n>>> Analizando Hoja: '{sheet}'")
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                except Exception as e:
                    print(f"  [AVISO] Error al leer hoja '{sheet}': {e}")
                    continue

                # 1. BUSCAR ENCABEZADO "NÚMERO" PARA VALIDAR SI ES UNA HOJA DE DATOS
                h_idx = None
                for idx, row in df.iterrows():
                    # Buscamos la palabra clave en la fila
                    row_vals = [str(x).upper().strip() for x in row.values if pd.notna(x)]
                    if "NÚMERO" in row_vals or "NUMERO" in row_vals:
                        h_idx = idx
                        break
                
                if h_idx is None:
                    # Si no tiene la columna NÚMERO, asumimos que no es una hoja de indicadores
                    print(f"  [OMITIDO] La hoja '{sheet}' no parece contener una tabla de indicadores (Falta columna 'NÚMERO').")
                    continue
                else:
                    print(f"  [INFO] Encabezados detectados en fila {h_idx + 1}. Procesando...")

                headers = [str(h).strip() for h in df.iloc[h_idx]]
                
                # 2. Mapeo de Columnas
                def find_c(names):
                    for i, h in enumerate(headers):
                        if any(n.lower() in h.lower().replace('\n', ' ') for n in names): return i
                    return None

                c_map = {
                    "num": find_c(["NÚMERO", "NUMERO"]),
                    "prod": find_c(["PRODUCTO O PROCESO ESPECÍFICO", "PRODUCTO"]),
                    "ind": find_c(["INDICADOR"]),
                    "form": find_c(["FORMULA", "FÓRMULA"]),
                    "uni": find_c(["UNIDAD"]),
                    "resp": find_c(["RESPONSABLE CENTRO", "RESPONSABLE"]),
                    "gest": find_c(["GESTOR"]),
                    "sup": find_c(["SUPERVISORES"]),
                    # Buscamos Meta del año que sea (flexible 2025, 2026, etc)
                    "meta": find_c(["Meta 2025", "Meta 2026", "Meta"]),
                    "pond": find_c(["Ponderador"]),
                    "op_desc": find_c(["Operandos"]), 
                    "op_est": find_c(["Operandos Estimados Meta", "Operandos  Estimados", "Operandos Estimados"]),
                    "proy": find_c(["Cumplimiento Proyectado"]),
                    "cump_meta": find_c(["% Cumplimiento de Meta", "% Cumplimiento"]),
                    "medios": find_c(["Medios de Verificación", "Medios"]),
                    "control": find_c(["Control de Cambios"]),
                    "inst": find_c(["Instrumentos de Gestión"])
                }

                # Reportar columnas faltantes críticas pero permitir continuar
                missing = [k for k, v in c_map.items() if v is None and k not in ["pond", "control"]]
                if missing:
                    print(f"  [ADVERTENCIA] Faltan columnas opcionales o críticas: {missing}")
                    # No detenemos el script por defecto en batch, salvo que sea crítico crítico

                # Meses
                months = ["Ene.", "Feb.", "Acum Feb.", "Mar.", "Acum Mar.", "Abr.", "Acum Abr.", 
                          "May.", "Acum May.", "Jun.", "Acum Jun.", "Jul.", "Acum Jul.", "Ago.", 
                          "Acum Ago", "Sept.", "Acum Sept", "Oct.", "Acum Oct.", "Nov.", "Acum Nov.", "Dic."]
                month_map = {m: find_c([m]) for m in months}

                # 3. Extracción
                sheet_rows = []
                count_rows = 0
                
                for i in range(h_idx + 1, len(df)):
                    def get_val(col_idx, row_offset=0):
                        if col_idx is None or (i + row_offset) >= len(df): return "No aplica"
                        val = df.iloc[i + row_offset, col_idx]
                        return val if pd.notna(val) else ""

                    raw_num = get_val(c_map["num"])
                    str_num = str(raw_num).strip()

                    # Detector de Indicador
                    if str_num == "" or str_num.lower() == "nan" or "NUEVO" in str_num.upper():
                        ind_val = get_val(c_map["ind"])
                        # Validar que tenga texto el indicador para considerarlo nuevo
                        if ind_val and str(ind_val).strip() != "" and str(ind_val).strip() != "0":
                            prefix_file = file_name.split()[0][:8] # Primeros caracteres del nombre archivo
                            # Limpiar nombre de hoja para codigo
                            clean_sheet = ''.join(e for e in sheet if e.isalnum())
                            final_code = f"NUEVO_{self.new_indicator_count}_{prefix_file}_{clean_sheet}"
                            print(f"    [AVISO] Fila {i+1}: Sin código. Generado: {final_code}")
                            self.new_indicator_count += 1
                        else:
                            continue 
                    else:
                        # Debe empezar con dígito para ser código real
                        if not any(c.isdigit() for c in str_num): continue 
                        final_code = str_num

                    count_rows += 1
                    
                    row_data = {
                        "ARCHIVO": file_name,
                        "HOJA": sheet,
                        "NÚMERO": final_code,
                        "PRODUCTO": get_val(c_map["prod"]),
                        "INDICADOR": get_val(c_map["ind"]),
                        "FORMULA": get_val(c_map["form"]),
                        "UNIDAD": get_val(c_map["uni"]),
                        "RESPONSABLE": get_val(c_map["resp"]),
                        "GESTOR": get_val(c_map["gest"]),
                        "SUPERVISORES": get_val(c_map["sup"]),
                        "Meta": self.transform_percentage(get_val(c_map["meta"]), "Meta"),
                        "Ponderador": self.transform_percentage(get_val(c_map["pond"]), "Ponderador"),
                        "Desc Op 1": get_val(c_map["op_desc"], 0),
                        "Desc Op 2": get_val(c_map["op_desc"], 3),
                        "Meta Op 1": get_val(c_map["op_est"], 3),
                        "Meta Op 2": get_val(c_map["op_est"], 5),
                    }

                    for m_name, m_idx in month_map.items():
                        row_data[f"{m_name} Op 1"] = get_val(m_idx, 3)
                        row_data[f"{m_name} Op 2"] = get_val(m_idx, 5)

                    row_data["Cump Proy Op 1"] = get_val(c_map["proy"], 3)
                    row_data["Cump Proy Op 2"] = get_val(c_map["proy"], 5)
                    row_data["% Cump Meta"] = self.transform_percentage(get_val(c_map["cump_meta"], 3), "% Cump Meta")
                    row_data["Medios Verif"] = get_val(c_map["medios"], 0)
                    row_data["Control Cambios"] = get_val(c_map["control"], 0)
                    row_data["Instrumentos"] = get_val(c_map["inst"], 0)

                    sheet_rows.append(row_data)
                    self.flat_data.append(row_data)

                self.data_tree[file_name][sheet] = sheet_rows
                print(f"  [OK] {count_rows} indicadores extraídos.")

    def export_excel(self):
        if not self.flat_data:
            print("\n[AVISO] No se extrajeron datos. Verifica que las hojas tengan la columna 'NÚMERO'.")
            return

        print(f"\n{'='*60}")
        print("CONSOLIDANDO Y GENERANDO EXCEL MAESTRO...")
        print(f"{'='*60}")
        
        wb = Workbook()
        
        # --- HOJA 1: CARGA BRUTA ---
        ws_bruta = wb.active
        ws_bruta.title = "Carga Bruta Consolidada"
        if self.flat_data:
            headers = list(self.flat_data[0].keys())
            ws_bruta.append(headers)
            for row in self.flat_data:
                ws_bruta.append(list(row.values()))
        
        # --- HOJA 2: ESTILIZADA ---
        ws_style = wb.create_sheet("Planilla Estilizada")
        
        file_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        sheet_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=11)
        black_bold = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        current_row = 1
        max_col_idx = 0

        # Iterar sobre archivos
        for file_name, sheets in self.data_tree.items():
            if not any(sheets.values()): continue 

            # SEPARADOR DE ARCHIVO
            ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
            c = ws_style.cell(row=current_row, column=1, value=f"ARCHIVO FUENTE: {file_name}")
            c.fill = file_fill
            c.font = white_font
            c.alignment = Alignment(horizontal='center')
            current_row += 1

            # Iterar sobre hojas
            for sheet_name, rows in sheets.items():
                if not rows: continue

                # SEPARADOR DE HOJA
                ws_style.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
                c = ws_style.cell(row=current_row, column=1, value=f"PLANILLA: {sheet_name}")
                c.fill = sheet_fill
                c.font = white_font
                current_row += 1

                keys_to_show = [k for k in rows[0].keys() if k not in ["ARCHIVO", "HOJA"]]
                max_col_idx = max(max_col_idx, len(keys_to_show))

                # Encabezados
                for idx, k in enumerate(keys_to_show, 1):
                    c = ws_style.cell(row=current_row, column=idx, value=k)
                    c.fill = header_fill
                    c.font = black_bold
                    c.border = thin_border
                current_row += 1

                # Datos
                for row_dict in rows:
                    for idx, k in enumerate(keys_to_show, 1):
                        val = row_dict[k]
                        c = ws_style.cell(row=current_row, column=idx, value=val)
                        c.border = thin_border
                        c.alignment = Alignment(wrap_text=True, vertical='top')
                    current_row += 1
                current_row += 1

        for i in range(1, max_col_idx + 2):
            col_let = get_column_letter(i)
            ws_style.column_dimensions[col_let].width = 20

        wb.save(self.output_file)
        print(f"\n[EXITO FINAL] Archivo maestro generado en: {self.output_file}")

if __name__ == "__main__":
    # =========================================================================
    # RUTA DE LA CARPETA
    # =========================================================================
    try:
        RUTA_CARPETA = input("Ingresa la ruta de la carpeta (o presiona Enter): ").strip()
        if not RUTA_CARPETA:
            RUTA_CARPETA = os.getcwd() 
        
        if os.path.isdir(RUTA_CARPETA):
            parser = IPSParserV180(RUTA_CARPETA)
            parser.process_folder()
            parser.export_excel()
        else:
            print(f"[ERROR] Ruta inválida: {RUTA_CARPETA}")
    except Exception as e:
        print(f"Error fatal: {e}")
        input("Presiona Enter para salir.")